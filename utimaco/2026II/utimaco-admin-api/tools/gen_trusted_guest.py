#!/usr/bin/env python3
"""
生成 trusted / guest sheet — 覆盖全部 7.1/7.2 接口的参数校验和业务逻辑测试。

每个接口的标准测试维度:
  1. 正常调用 (happy path)
  2. requestId="" → 400
  3. requestId 缺失 → 400
  4. requestId 非UUID格式 → 200/400 (边界)
  5. oprType="" / 缺失 / 无效 (oprType 接口)
  6. 各必填字段缺失/空/非法
  7. 业务边界
"""

import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

CALLBACK_URL = "${env.callbackUrl}"
UPLOAD_URL = "${env.uploadUrl}"
ALARM_URL = "${env.alarmUrl}"
MONITORING_URL = "${env.monitoringUrl}"
LOG_SERVER_URL = "${env.logServerUrl}"
CHSM_IMAGE_URL = "${env.chsmImageUrl}"
CHSM_PACK_URL = "${env.chsmPackUrl}"
CHSM_BACKUP_URL = "${env.chsmBackupUrl}"
VSM_IMAGE_URL = "${env.vsmImageUrl}"
VSM_PACK_URL = "${env.vsmPackUrl}"
SIGN_BASE64 = "${env.signBase64}"
SIGN_HEX = "${env.signHex}"

HEADERS = [
    "case_id", "description", "host", "endpoint", "method",
    "params", "json_data", "expected_status", "assert_rules",
    "scenario_id", "step", "step_type", "save_vars",
    "enabled", "section", "ref_case_id", "auth", "approved",
]

OK = json.dumps([
    {"type": "status_code", "expected_code": 200},
    {"type": "json_contains", "key": "message", "value": "success"},
], ensure_ascii=False)

ERR400 = json.dumps([
    {"type": "status_code", "expected_code": 400},
], ensure_ascii=False)


def ok_with(*extra_rules):
    base = [
        {"type": "status_code", "expected_code": 200},
        {"type": "json_contains", "key": "message", "value": "success"},
    ]
    base.extend(extra_rules)
    return json.dumps(base, ensure_ascii=False)


def jd(d):
    return json.dumps(d, ensure_ascii=False)


def row(case_id, desc, endpoint, expected_status, json_data=None, params=None,
        assert_rules=None, ref_case_id=None, enabled="yes", section="",
        scenario_id=None, step=None, step_type=None, save_vars=None,
        method="POST", auth="yes", approved=""):
    if assert_rules is None:
        assert_rules = OK if expected_status == 200 else ERR400
    return {
        "case_id": case_id,
        "description": desc,
        "host": None,
        "endpoint": endpoint,
        "method": method,
        "params": params,
        "json_data": json_data,
        "expected_status": expected_status,
        "assert_rules": assert_rules,
        "scenario_id": scenario_id,
        "step": step,
        "step_type": step_type,
        "save_vars": save_vars,
        "enabled": enabled,
        "section": section,
        "ref_case_id": ref_case_id or "",
        "auth": auth,
        "approved": approved,
    }



# ================================================================
# CHSM 查询类
# ================================================================

def gen_getCHSMInfo():
    EP = "/api/1.0/chsm"
    return [
        row("CHSM_INFO_001", "getCHSMInfo 正常获取", EP, 200,
            jd({"requestId": "uuid", "oprType": "getinfo"}),
            assert_rules=ok_with({"type": "json_contains", "key": "result.id"}),
            ref_case_id="CHSM_INFO_001", section="7.1.1"),
        row("CHSM_INFO_002", "getCHSMInfo oprType空字符串", EP, 400,
            jd({"requestId": "uuid", "oprType": ""}),
            ref_case_id="CHSM_INFO_002", section="7.1.1"),
        row("CHSM_INFO_003", "getCHSMInfo oprType无效值", EP, 400,
            jd({"requestId": "uuid", "oprType": "invalid_type"}),
            ref_case_id="CHSM_INFO_003", section="7.1.1"),
        row("CHSM_INFO_004", "getCHSMInfo 缺requestId", EP, 400,
            jd({"oprType": "getinfo"}),
            ref_case_id="CHSM_INFO_004", section="7.1.1"),
        row("CHSM_INFO_007", "getCHSMInfo requestId非UUID格式", EP, 200,
            jd({"requestId": "abc123", "oprType": "getinfo"}),
            assert_rules=ok_with({"type": "json_contains", "key": "result.id"}),
            ref_case_id="CHSM_INFO_007", section="7.1.1"),
        row("CHSM_INFO_T01", "getCHSMInfo requestId空字符串", EP, 400,
            jd({"requestId": "", "oprType": "getinfo"}),
            section="7.1.1"),
        row("CHSM_INFO_T02", "getCHSMInfo oprType缺失", EP, 400,
            jd({"requestId": "uuid"}),
            section="7.1.1"),
        row("CHSM_INFO_T03", "getCHSMInfo 空body", EP, 400,
            jd({}),
            section="7.1.1"),
    ]


def gen_getCHSMDebugInfo():
    EP = "/api/1.0/chsm/debuginfo"
    return [
        row("CHSM_DEBUG_001", "getCHSMDebugInfo 正常获取", EP, 200,
            params="requestId=uuid",
            assert_rules=ok_with({"type": "json_contains", "key": "result"}),
            ref_case_id="CHSM_DEBUG_001", section="7.1.17", method="GET", auth="no"),
        row("CHSM_DEBUG_T01", "getCHSMDebugInfo requestId空字符串", EP, 400,
            params="requestId=",
            section="7.1.17", method="GET", auth="no"),
        row("CHSM_DEBUG_T02", "getCHSMDebugInfo 缺requestId", EP, 400,
            method="GET", section="7.1.17", auth="no"),
        row("CHSM_DEBUG_T03", "getCHSMDebugInfo requestId非UUID", EP, 200,
            params="requestId=not-a-uuid",
            assert_rules=ok_with({"type": "json_contains", "key": "result"}),
            section="7.1.17", method="GET", auth="no"),
    ]


def gen_getCHSMDeviceInfo():
    EP = "/api/1.0/chsm"
    return [
        row("CHSM_DEVICE_001", "getCHSMDeviceInfo 正常获取", EP, 200,
            jd({"requestId": "uuid", "oprType": "getDeviceInfo"}),
            assert_rules=ok_with(
                {"type": "json_contains", "key": "result.version"},
                {"type": "json_contains", "key": "result.sn"},
            ),
            ref_case_id="CHSM_DEVICE_001", section="7.1.18"),
        row("CHSM_DEVICE_T01", "getCHSMDeviceInfo requestId空字符串", EP, 400,
            jd({"requestId": "", "oprType": "getDeviceInfo"}),
            section="7.1.18"),
        row("CHSM_DEVICE_T02", "getCHSMDeviceInfo 缺requestId", EP, 400,
            jd({"oprType": "getDeviceInfo"}),
            section="7.1.18"),
        row("CHSM_DEVICE_T03", "getCHSMDeviceInfo oprType缺失", EP, 400,
            jd({"requestId": "uuid"}),
            section="7.1.18"),
        row("CHSM_DEVICE_T04", "getCHSMDeviceInfo oprType空字符串", EP, 400,
            jd({"requestId": "uuid", "oprType": ""}),
            section="7.1.18"),
    ]


# ================================================================
# CHSM 配置类
# ================================================================

def gen_configCHSMNet():
    EP = "/api/1.0/chsm/network"
    FULL = {
        "requestId": "uuid",
        "dnsList": ["114.114.114.114", "192.168.8.1"],
        "netAddrs": [{"name": "eth0", "ip": "192.168.8.120", "mask": "255.255.255.0", "gateway": "192.168.8.1"}],
    }
    return [
        row("CHSM_NET_001", "configCHSMNet 正常配置全参数", EP, 200,
            jd(FULL),
            ref_case_id="CHSM_NET_001", section="7.1.4"),
        row("CHSM_NET_002", "configCHSMNet 重复配置(幂等)", EP, 200,
            jd(FULL),
            ref_case_id="CHSM_NET_002", section="7.1.4"),
        row("CHSM_NET_003", "configCHSMNet IPv6地址", EP, 200,
            jd({"requestId": "uuid", "dnsList": ["8.8.8.8"], "netAddrs": [{"name": "eth0", "ip": "fd00::100", "mask": "64", "gateway": "fd00::1"}]}),
            ref_case_id="CHSM_NET_003", section="7.1.4"),
        row("CHSM_NET_T01", "configCHSMNet requestId空字符串", EP, 400,
            jd({**FULL, "requestId": ""}),
            section="7.1.4"),
        row("CHSM_NET_T02", "configCHSMNet 缺requestId", EP, 400,
            jd({k: v for k, v in FULL.items() if k != "requestId"}),
            section="7.1.4"),
        row("CHSM_NET_T03", "configCHSMNet 缺dnsList", EP, 400,
            jd({k: v for k, v in FULL.items() if k != "dnsList"}),
            section="7.1.4"),
        row("CHSM_NET_T04", "configCHSMNet 缺netAddrs", EP, 400,
            jd({k: v for k, v in FULL.items() if k != "netAddrs"}),
            section="7.1.4"),
        row("CHSM_NET_T05", "configCHSMNet dnsList空数组", EP, 400,
            jd({**FULL, "dnsList": []}),
            section="7.1.4"),
        row("CHSM_NET_T06", "configCHSMNet netAddrs空数组", EP, 400,
            jd({**FULL, "netAddrs": []}),
            section="7.1.4"),
        row("CHSM_NET_T07", "configCHSMNet netAddrs缺ip字段", EP, 400,
            jd({**FULL, "netAddrs": [{"name": "eth0", "mask": "255.255.255.0", "gateway": "192.168.1.1"}]}),
            section="7.1.4"),
        row("CHSM_NET_T08", "configCHSMNet netAddrs缺mask字段", EP, 400,
            jd({**FULL, "netAddrs": [{"name": "eth0", "ip": "192.168.1.100", "gateway": "192.168.1.1"}]}),
            section="7.1.4"),
        row("CHSM_NET_T09", "configCHSMNet netAddrs缺gateway字段", EP, 400,
            jd({**FULL, "netAddrs": [{"name": "eth0", "ip": "192.168.1.100", "mask": "255.255.255.0"}]}),
            section="7.1.4"),
        row("CHSM_NET_T10", "configCHSMNet netAddrs缺name字段", EP, 400,
            jd({**FULL, "netAddrs": [{"ip": "192.168.1.100", "mask": "255.255.255.0", "gateway": "192.168.1.1"}]}),
            section="7.1.4"),
        row("CHSM_NET_T11", "configCHSMNet 无效IP", EP, 400,
            jd({**FULL, "netAddrs": [{"name": "eth0", "ip": "999.999.999.999", "mask": "255.255.255.0", "gateway": "192.168.1.1"}]}),
            section="7.1.4"),
        row("CHSM_NET_T12", "configCHSMNet dnsList含非法DNS格式", EP, 400,
            jd({**FULL, "dnsList": ["not_an_ip"]}),
            section="7.1.4"),
    ]


def gen_configCHSMNtp():
    EP = "/api/1.0/chsm/ntp"
    return [
        row("CHSM_NTP_001", "configCHSMNtp 正常配置", EP, 200,
            jd({"requestId": "uuid", "addr": "ntp.aliyun.com", "syncPeriod": 60}),
            ref_case_id="CHSM_NTP_001", section="7.1.5"),
        row("CHSM_NTP_002", "configCHSMNtp syncPeriod超大值", EP, 200,
            jd({"requestId": "uuid", "addr": "ntp.aliyun.com", "syncPeriod": 99999}),
            ref_case_id="CHSM_NTP_002", section="7.1.5"),
        row("CHSM_NTP_003", "configCHSMNtp addr空字符串", EP, 400,
            jd({"requestId": "uuid", "addr": "", "syncPeriod": 60}),
            ref_case_id="CHSM_NTP_003", section="7.1.5"),
        row("CHSM_NTP_004", "configCHSMNtp syncPeriod负数", EP, 400,
            jd({"requestId": "uuid", "addr": "10.10.1.1", "syncPeriod": -1}),
            ref_case_id="CHSM_NTP_004", section="7.1.5"),
        row("CHSM_NTP_T01", "configCHSMNtp requestId空字符串", EP, 400,
            jd({"requestId": "", "addr": "10.10.1.1", "syncPeriod": 60}),
            section="7.1.5"),
        row("CHSM_NTP_T02", "configCHSMNtp 缺requestId", EP, 400,
            jd({"addr": "10.10.1.1", "syncPeriod": 60}),
            section="7.1.5"),
        row("CHSM_NTP_T03", "configCHSMNtp 缺addr", EP, 400,
            jd({"requestId": "uuid", "syncPeriod": 60}),
            section="7.1.5"),
        row("CHSM_NTP_T04", "configCHSMNtp 缺syncPeriod", EP, 400,
            jd({"requestId": "uuid", "addr": "10.10.1.1"}),
            section="7.1.5"),
        row("CHSM_NTP_T05", "configCHSMNtp syncPeriod=0", EP, 400,
            jd({"requestId": "uuid", "addr": "10.10.1.1", "syncPeriod": 0}),
            section="7.1.5"),
        row("CHSM_NTP_T06", "configCHSMNtp addr非法格式", EP, 400,
            jd({"requestId": "uuid", "addr": "not_an_ip", "syncPeriod": 60}),
            section="7.1.5"),
    ]


def gen_configCHSMUploadAddress():
    EP = "/api/1.0/chsm/imageuploader"
    return [
        row("CHSM_UPLOAD_001", "configCHSMUploadAddress 正常配置", EP, 200,
            jd({"requestId": "uuid", "url": UPLOAD_URL}),
            ref_case_id="CHSM_UPLOAD_001", section="7.1.6"),
        row("CHSM_UPLOAD_002", "configCHSMUploadAddress url空字符串", EP, 400,
            jd({"requestId": "uuid", "url": ""}),
            ref_case_id="CHSM_UPLOAD_002", section="7.1.6"),
        row("CHSM_UPLOAD_T01", "configCHSMUploadAddress requestId空字符串", EP, 400,
            jd({"requestId": "", "url": UPLOAD_URL}),
            section="7.1.6"),
        row("CHSM_UPLOAD_T02", "configCHSMUploadAddress 缺requestId", EP, 400,
            jd({"url": UPLOAD_URL}),
            section="7.1.6"),
        row("CHSM_UPLOAD_T03", "configCHSMUploadAddress 缺url", EP, 400,
            jd({"requestId": "uuid"}),
            section="7.1.6"),
        row("CHSM_UPLOAD_T04", "configCHSMUploadAddress url非法格式", EP, 400,
            jd({"requestId": "uuid", "url": "not_a_url"}),
            section="7.1.6"),
    ]


def gen_configSyslogAddr():
    EP = "/api/1.0/chsm/loguploader"
    return [
        row("CHSM_SYSLOG_001", "configSyslogAddr type=syslog", EP, 200,
            jd({"requestId": "uuid", "logServerType": "syslog", "logServerAddress": "10.10.1.100:514"}),
            ref_case_id="CHSM_SYSLOG_001", section="7.1.7"),
        row("CHSM_SYSLOG_002", "configSyslogAddr type=logserver", EP, 200,
            jd({"requestId": "uuid", "logServerType": "logserver", "logServerAddress": LOG_SERVER_URL}),
            ref_case_id="CHSM_SYSLOG_002", section="7.1.7"),
        row("CHSM_SYSLOG_003", "configSyslogAddr type无效值", EP, 400,
            jd({"requestId": "uuid", "logServerType": "invalid", "logServerAddress": "10.10.1.100:514"}),
            ref_case_id="CHSM_SYSLOG_003", section="7.1.7"),
        row("CHSM_SYSLOG_T01", "configSyslogAddr requestId空字符串", EP, 400,
            jd({"requestId": "", "logServerType": "syslog", "logServerAddress": "10.10.1.100:514"}),
            section="7.1.7"),
        row("CHSM_SYSLOG_T02", "configSyslogAddr 缺requestId", EP, 400,
            jd({"logServerType": "syslog", "logServerAddress": "10.10.1.100:514"}),
            section="7.1.7"),
        row("CHSM_SYSLOG_T03", "configSyslogAddr 缺logServerType", EP, 400,
            jd({"requestId": "uuid", "logServerAddress": "10.10.1.100:514"}),
            section="7.1.7"),
        row("CHSM_SYSLOG_T04", "configSyslogAddr 缺logServerAddress", EP, 400,
            jd({"requestId": "uuid", "logServerType": "syslog"}),
            section="7.1.7"),
        row("CHSM_SYSLOG_T05", "configSyslogAddr logServerAddress空字符串", EP, 400,
            jd({"requestId": "uuid", "logServerType": "syslog", "logServerAddress": ""}),
            section="7.1.7"),
        row("CHSM_SYSLOG_T06", "configSyslogAddr logServerType空字符串", EP, 400,
            jd({"requestId": "uuid", "logServerType": "", "logServerAddress": "10.10.1.100:514"}),
            section="7.1.7"),
    ]


def gen_configCHSMAlarmAddress():
    EP = "/api/1.0/chsm/alarmaddress"
    FULL = {"requestId": "uuid", "url": ALARM_URL, "monitoringUrl": MONITORING_URL}
    return [
        row("CHSM_ALARM_001", "configCHSMAlarmAddress 正常配置全参数", EP, 200,
            jd(FULL),
            ref_case_id="CHSM_ALARM_001", section="7.1.14"),
        row("CHSM_ALARM_002", "configCHSMAlarmAddress 重复配置(幂等)", EP, 200,
            jd(FULL),
            ref_case_id="CHSM_ALARM_002", section="7.1.14"),
        row("CHSM_ALARM_T01", "configCHSMAlarmAddress requestId空字符串", EP, 400,
            jd({**FULL, "requestId": ""}),
            section="7.1.14"),
        row("CHSM_ALARM_T02", "configCHSMAlarmAddress 缺requestId", EP, 400,
            jd({k: v for k, v in FULL.items() if k != "requestId"}),
            section="7.1.14"),
        row("CHSM_ALARM_T03", "configCHSMAlarmAddress 缺url", EP, 400,
            jd({k: v for k, v in FULL.items() if k != "url"}),
            section="7.1.14"),
        row("CHSM_ALARM_T04", "configCHSMAlarmAddress url空字符串", EP, 400,
            jd({**FULL, "url": ""}),
            section="7.1.14"),
        row("CHSM_ALARM_T05", "configCHSMAlarmAddress 缺monitoringUrl", EP, 400,
            jd({k: v for k, v in FULL.items() if k != "monitoringUrl"}),
            section="7.1.14"),
        row("CHSM_ALARM_T06", "configCHSMAlarmAddress monitoringUrl空字符串", EP, 400,
            jd({**FULL, "monitoringUrl": ""}),
            section="7.1.14"),
    ]


def gen_configCHSMToken():
    EP = "/api/1.0/chsm/cloudtoken"
    return [
        row("CHSM_TOKEN_001", "configCHSMToken 正常配置", EP, 200,
            jd({"requestId": "uuid", "cloudToken": "eyJhbGciOiJSUzI1NiIsInR5cCI6IkpXVCJ9.test_token_value"}),
            ref_case_id="CHSM_TOKEN_001", section="7.1.15"),
        row("CHSM_TOKEN_002", "configCHSMToken cloudToken空字符串", EP, 400,
            jd({"requestId": "uuid", "cloudToken": ""}),
            ref_case_id="CHSM_TOKEN_002", section="7.1.15"),
        row("CHSM_TOKEN_T01", "configCHSMToken requestId空字符串", EP, 400,
            jd({"requestId": "", "cloudToken": "test_token"}),
            section="7.1.15"),
        row("CHSM_TOKEN_T02", "configCHSMToken 缺requestId", EP, 400,
            jd({"cloudToken": "test_token"}),
            section="7.1.15"),
        row("CHSM_TOKEN_T03", "configCHSMToken 缺cloudToken", EP, 400,
            jd({"requestId": "uuid"}),
            section="7.1.15"),
    ]


def gen_configCHSMMOOCAddress():
    EP = "/api/1.0/chsm/moocaddress"
    FULL = {
        "requestId": "uuid", "ocIp": "10.10.10.1", "ocProtocol": "https",
        "ocPort": "26635", "resourceUrl": "/rest/cloudInfra/v1/resource",
        "performanceUrl": "/rest/cloudInfra/v1/performance",
        "bizRegionNativeId": "region-001", "cloudInfraType": "FusionSphere",
        "dewServiceHost": "kms.cn-north-1.myhuaweicloud.com",
    }
    return [
        row("CHSM_MOOC_001", "configCHSMMOOCAddress 正常配置全参数", EP, 200,
            jd(FULL),
            ref_case_id="CHSM_MOOC_001", section="7.1.16"),
        row("CHSM_MOOC_002", "configCHSMMOOCAddress 缺ocIp", EP, 400,
            jd({k: v for k, v in FULL.items() if k != "ocIp"}),
            ref_case_id="CHSM_MOOC_002", section="7.1.16"),
        row("CHSM_MOOC_T01", "configCHSMMOOCAddress requestId空字符串", EP, 400,
            jd({**FULL, "requestId": ""}),
            section="7.1.16"),
        row("CHSM_MOOC_T02", "configCHSMMOOCAddress 缺requestId", EP, 400,
            jd({k: v for k, v in FULL.items() if k != "requestId"}),
            section="7.1.16"),
        row("CHSM_MOOC_T03", "configCHSMMOOCAddress ocIp空字符串", EP, 400,
            jd({**FULL, "ocIp": ""}),
            section="7.1.16"),
        row("CHSM_MOOC_T04", "configCHSMMOOCAddress 缺ocPort", EP, 400,
            jd({k: v for k, v in FULL.items() if k != "ocPort"}),
            section="7.1.16"),
        row("CHSM_MOOC_T05", "configCHSMMOOCAddress 缺resourceUrl", EP, 400,
            jd({k: v for k, v in FULL.items() if k != "resourceUrl"}),
            section="7.1.16"),
        row("CHSM_MOOC_T06", "configCHSMMOOCAddress 缺performanceUrl", EP, 400,
            jd({k: v for k, v in FULL.items() if k != "performanceUrl"}),
            section="7.1.16"),
        row("CHSM_MOOC_T07", "configCHSMMOOCAddress 缺bizRegionNativeId", EP, 400,
            jd({k: v for k, v in FULL.items() if k != "bizRegionNativeId"}),
            section="7.1.16"),
        row("CHSM_MOOC_T08", "configCHSMMOOCAddress 缺cloudInfraType", EP, 400,
            jd({k: v for k, v in FULL.items() if k != "cloudInfraType"}),
            section="7.1.16"),
        row("CHSM_MOOC_T09", "configCHSMMOOCAddress 缺dewServiceHost", EP, 400,
            jd({k: v for k, v in FULL.items() if k != "dewServiceHost"}),
            section="7.1.16"),
        row("CHSM_MOOC_T10", "configCHSMMOOCAddress ocPort非数值", EP, 400,
            jd({**FULL, "ocPort": "abc"}),
            section="7.1.16"),
    ]


# ================================================================
# CHSM 异步类
# ================================================================

def gen_exportCHSM():
    EP = "/api/1.0/chsm/image"
    return [
        row("CHSM_EXPORT_001", "exportCHSM 正常导出", EP, 200,
            jd({"requestId": "uuid", "oprType": "export", "callbackUrl": CALLBACK_URL}),
            ref_case_id="CHSM_EXPORT_001", section="7.1.8"),
        row("CHSM_EXPORT_002", "exportCHSM 缺callbackUrl", EP, 400,
            jd({"requestId": "uuid", "oprType": "export"}),
            ref_case_id="CHSM_EXPORT_002", section="7.1.8"),
        row("CHSM_EXPORT_T01", "exportCHSM requestId空字符串", EP, 400,
            jd({"requestId": "", "oprType": "export", "callbackUrl": CALLBACK_URL}),
            section="7.1.8"),
        row("CHSM_EXPORT_T02", "exportCHSM 缺requestId", EP, 400,
            jd({"oprType": "export", "callbackUrl": CALLBACK_URL}),
            section="7.1.8"),
        row("CHSM_EXPORT_T03", "exportCHSM oprType缺失", EP, 400,
            jd({"requestId": "uuid", "callbackUrl": CALLBACK_URL}),
            section="7.1.8"),
        row("CHSM_EXPORT_T04", "exportCHSM callbackUrl空字符串", EP, 400,
            jd({"requestId": "uuid", "oprType": "export", "callbackUrl": ""}),
            section="7.1.8"),
        row("CHSM_EXPORT_T05", "exportCHSM oprType空字符串", EP, 400,
            jd({"requestId": "uuid", "oprType": "", "callbackUrl": CALLBACK_URL}),
            section="7.1.8"),
    ]


def gen_importCHSM():
    EP = "/api/1.0/chsm/image"
    BASE = {
        "requestId": "uuid", "oprType": "import",
        "imageUrl": CHSM_IMAGE_URL,
        "alg": "RSAWithSHA256", "sign": SIGN_BASE64,
        "callbackUrl": CALLBACK_URL,
    }
    return [
        row("CHSM_IMPORT_001", "importCHSM 正常导入", EP, 200,
            jd(BASE), ref_case_id="CHSM_IMPORT_001", section="7.1.9"),
        row("CHSM_IMPORT_003", "importCHSM imageUrl不可达", EP, 200,
            jd({**BASE, "imageUrl": "http://192.168.99.99/no_exist.zip"}),
            ref_case_id="CHSM_IMPORT_003", section="7.1.9"),
        row("CHSM_IMPORT_004", "importCHSM sign篡改", EP, 200,
            jd({**BASE, "sign": "TAMPERED_SIGN_VALUE"}),
            ref_case_id="CHSM_IMPORT_004", section="7.1.9"),
        row("CHSM_IMPORT_005", "importCHSM 缺alg", EP, 400,
            jd({k: v for k, v in BASE.items() if k != "alg"}),
            ref_case_id="CHSM_IMPORT_005", section="7.1.9"),
        row("CHSM_IMPORT_T01", "importCHSM requestId空字符串", EP, 400,
            jd({**BASE, "requestId": ""}),
            section="7.1.9"),
        row("CHSM_IMPORT_T02", "importCHSM 缺requestId", EP, 400,
            jd({k: v for k, v in BASE.items() if k != "requestId"}),
            section="7.1.9"),
        row("CHSM_IMPORT_T03", "importCHSM 缺imageUrl", EP, 400,
            jd({k: v for k, v in BASE.items() if k != "imageUrl"}),
            section="7.1.9"),
        row("CHSM_IMPORT_T04", "importCHSM 缺sign", EP, 400,
            jd({k: v for k, v in BASE.items() if k != "sign"}),
            section="7.1.9"),
        row("CHSM_IMPORT_T05", "importCHSM 缺callbackUrl", EP, 400,
            jd({k: v for k, v in BASE.items() if k != "callbackUrl"}),
            section="7.1.9"),
        row("CHSM_IMPORT_T06", "importCHSM oprType缺失", EP, 400,
            jd({k: v for k, v in BASE.items() if k != "oprType"}),
            section="7.1.9"),
        row("CHSM_IMPORT_T07", "importCHSM alg无效值", EP, 400,
            jd({**BASE, "alg": "INVALID_ALG"}),
            section="7.1.9"),
    ]


def gen_upgradeCHSM():
    EP = "/api/1.0/chsm"
    BASE = {
        "requestId": "uuid", "oprType": "upgrade",
        "packVersion": "2.0.1", "packUrl": CHSM_PACK_URL,
        "alg": "RSAWithSHA256", "sign": SIGN_HEX,
        "callbackUrl": CALLBACK_URL,
    }
    return [
        row("CHSM_UPGRADE_001", "upgradeCHSM 正常升级", EP, 200,
            jd(BASE), ref_case_id="CHSM_UPGRADE_001", section="7.1.10"),
        row("CHSM_UPGRADE_003", "upgradeCHSM packUrl无效", EP, 200,
            jd({**BASE, "packUrl": "http://192.168.99.99/no_exist.bin"}),
            ref_case_id="CHSM_UPGRADE_003", section="7.1.10"),
        row("CHSM_UPGRADE_T01", "upgradeCHSM requestId空字符串", EP, 400,
            jd({**BASE, "requestId": ""}),
            section="7.1.10"),
        row("CHSM_UPGRADE_T02", "upgradeCHSM 缺requestId", EP, 400,
            jd({k: v for k, v in BASE.items() if k != "requestId"}),
            section="7.1.10"),
        row("CHSM_UPGRADE_T03", "upgradeCHSM 缺packVersion", EP, 400,
            jd({k: v for k, v in BASE.items() if k != "packVersion"}),
            section="7.1.10"),
        row("CHSM_UPGRADE_T04", "upgradeCHSM 缺packUrl", EP, 400,
            jd({k: v for k, v in BASE.items() if k != "packUrl"}),
            section="7.1.10"),
        row("CHSM_UPGRADE_T05", "upgradeCHSM 缺alg", EP, 400,
            jd({k: v for k, v in BASE.items() if k != "alg"}),
            section="7.1.10"),
        row("CHSM_UPGRADE_T06", "upgradeCHSM 缺callbackUrl", EP, 400,
            jd({k: v for k, v in BASE.items() if k != "callbackUrl"}),
            section="7.1.10"),
        row("CHSM_UPGRADE_T07", "upgradeCHSM 缺sign", EP, 400,
            jd({k: v for k, v in BASE.items() if k != "sign"}),
            section="7.1.10"),
        row("CHSM_UPGRADE_T08", "upgradeCHSM 缺oprType", EP, 400,
            jd({k: v for k, v in BASE.items() if k != "oprType"}),
            section="7.1.10"),
        row("CHSM_UPGRADE_T09", "upgradeCHSM oprType空字符串", EP, 400,
            jd({**BASE, "oprType": ""}),
            section="7.1.10"),
    ]


def gen_restartCHSM():
    EP = "/api/1.0/chsm"
    return [
        row("CHSM_RESTART_001", "restartCHSM 正常重启", EP, 200,
            jd({"requestId": "uuid", "oprType": "restart", "callbackUrl": CALLBACK_URL}),
            ref_case_id="CHSM_RESTART_001", section="7.1.11"),
        row("CHSM_RESTART_T01", "restartCHSM requestId空字符串", EP, 400,
            jd({"requestId": "", "oprType": "restart", "callbackUrl": CALLBACK_URL}),
            section="7.1.11"),
        row("CHSM_RESTART_T02", "restartCHSM 缺requestId", EP, 400,
            jd({"oprType": "restart", "callbackUrl": CALLBACK_URL}),
            section="7.1.11"),
        row("CHSM_RESTART_T03", "restartCHSM 缺callbackUrl", EP, 400,
            jd({"requestId": "uuid", "oprType": "restart"}),
            section="7.1.11"),
        row("CHSM_RESTART_T04", "restartCHSM oprType缺失", EP, 400,
            jd({"requestId": "uuid", "callbackUrl": CALLBACK_URL}),
            section="7.1.11"),
        row("CHSM_RESTART_T05", "restartCHSM callbackUrl空字符串", EP, 400,
            jd({"requestId": "uuid", "oprType": "restart", "callbackUrl": ""}),
            section="7.1.11"),
        row("CHSM_RESTART_T06", "restartCHSM oprType空字符串", EP, 400,
            jd({"requestId": "uuid", "oprType": "", "callbackUrl": CALLBACK_URL}),
            section="7.1.11"),
    ]


def gen_backupCHSM():
    EP = "/api/1.0/chsm"
    return [
        row("CHSM_BACKUP_001", "backupCHSM 正常备份", EP, 200,
            jd({"requestId": "uuid", "oprType": "backup", "callbackUrl": CALLBACK_URL}),
            ref_case_id="CHSM_BACKUP_001", section="7.1.12"),
        row("CHSM_BACKUP_T01", "backupCHSM requestId空字符串", EP, 400,
            jd({"requestId": "", "oprType": "backup", "callbackUrl": CALLBACK_URL}),
            section="7.1.12"),
        row("CHSM_BACKUP_T02", "backupCHSM 缺requestId", EP, 400,
            jd({"oprType": "backup", "callbackUrl": CALLBACK_URL}),
            section="7.1.12"),
        row("CHSM_BACKUP_T03", "backupCHSM 缺callbackUrl", EP, 400,
            jd({"requestId": "uuid", "oprType": "backup"}),
            section="7.1.12"),
        row("CHSM_BACKUP_T04", "backupCHSM oprType缺失", EP, 400,
            jd({"requestId": "uuid", "callbackUrl": CALLBACK_URL}),
            section="7.1.12"),
        row("CHSM_BACKUP_T05", "backupCHSM callbackUrl空字符串", EP, 400,
            jd({"requestId": "uuid", "oprType": "backup", "callbackUrl": ""}),
            section="7.1.12"),
        row("CHSM_BACKUP_T06", "backupCHSM oprType空字符串", EP, 400,
            jd({"requestId": "uuid", "oprType": "", "callbackUrl": CALLBACK_URL}),
            section="7.1.12"),
    ]


def gen_restoreCHSM():
    EP = "/api/1.0/chsm"
    BASE = {
        "requestId": "uuid", "oprType": "restore",
        "backupUrl": CHSM_BACKUP_URL,
        "alg": "RSAWithSHA256", "sign": SIGN_BASE64,
        "callbackUrl": CALLBACK_URL,
    }
    return [
        row("CHSM_RESTORE_001", "restoreCHSM 正常恢复", EP, 200,
            jd(BASE), ref_case_id="CHSM_RESTORE_001", section="7.1.13"),
        row("CHSM_RESTORE_T01", "restoreCHSM requestId空字符串", EP, 400,
            jd({**BASE, "requestId": ""}),
            section="7.1.13"),
        row("CHSM_RESTORE_T02", "restoreCHSM 缺requestId", EP, 400,
            jd({k: v for k, v in BASE.items() if k != "requestId"}),
            section="7.1.13"),
        row("CHSM_RESTORE_T03", "restoreCHSM 缺backupUrl", EP, 400,
            jd({k: v for k, v in BASE.items() if k != "backupUrl"}),
            section="7.1.13"),
        row("CHSM_RESTORE_T04", "restoreCHSM 缺alg", EP, 400,
            jd({k: v for k, v in BASE.items() if k != "alg"}),
            section="7.1.13"),
        row("CHSM_RESTORE_T05", "restoreCHSM 缺callbackUrl", EP, 400,
            jd({k: v for k, v in BASE.items() if k != "callbackUrl"}),
            section="7.1.13"),
        row("CHSM_RESTORE_T06", "restoreCHSM 缺oprType", EP, 400,
            jd({k: v for k, v in BASE.items() if k != "oprType"}),
            section="7.1.13"),
        row("CHSM_RESTORE_T07", "restoreCHSM 缺sign", EP, 400,
            jd({k: v for k, v in BASE.items() if k != "sign"}),
            section="7.1.13"),
        row("CHSM_RESTORE_T08", "restoreCHSM oprType空字符串", EP, 400,
            jd({**BASE, "oprType": ""}),
            section="7.1.13"),
    ]


# ================================================================
# VSM 场景 + 独立用例 (按接口拆分，降低耦合)
# ================================================================

EP_VSM = "/api/1.0/vsm"
EP_VNET = "/api/1.0/vsm/network"
EP_VTOK = "/api/1.0/vsm/token"
EP_VIMG = "/api/1.0/vsm/image"

VSM_IMP_BASE = {
    "requestId": "uuid", "oprType": "import", "vsmId": "${env.vsmId}",
    "imageUrl": VSM_IMAGE_URL,
    "alg": "RSAWithSHA256", "sign": SIGN_BASE64,
    "callbackUrl": CALLBACK_URL,
}

VSM_UPG_BASE = {
    "requestId": "uuid", "oprType": "upgrade", "vsmId": "${env.vsmId}",
    "packVersion": "1.5.0", "packUrl": VSM_PACK_URL,
    "alg": "RSAWithSHA256", "sign": SIGN_HEX,
    "callbackUrl": CALLBACK_URL,
}


def _scn(scenario_id, suffix, test_rows):
    result = []
    for i, r in enumerate(test_rows, 1):
        r["scenario_id"] = scenario_id
        r["step"] = i
        if not r.get("step_type"):
            r["step_type"] = "test"
        result.append(r)
    return result


def gen_vsm_info_scenario():
    return _scn("SCN_VSM_INFO", "INFO", [
        row("VSM_INFO_001", "getVSMInfo 正常获取", EP_VSM, 200,
            jd({"requestId": "uuid", "oprType": "getinfo", "vsmId": "${env.vsmId}"}),
            assert_rules=ok_with(
                {"type": "json_contains", "key": "result.id"},
                {"type": "json_contains", "key": "result.version"},
            ),
            ref_case_id="VSM_INFO_001", section="7.2.1"),
        row("VSM_INFO_T01", "getVSMInfo requestId空字符串", EP_VSM, 400,
            jd({"requestId": "", "oprType": "getinfo", "vsmId": "${env.vsmId}"}),
            section="7.2.1"),
        row("VSM_INFO_T02", "getVSMInfo 缺requestId", EP_VSM, 400,
            jd({"oprType": "getinfo", "vsmId": "${env.vsmId}"}),
            section="7.2.1"),
        row("VSM_INFO_T03", "getVSMInfo oprType缺失", EP_VSM, 400,
            jd({"requestId": "uuid", "vsmId": "${env.vsmId}"}),
            section="7.2.1"),
    ])


def gen_vsm_net_scenario():
    return _scn("SCN_VSM_NET", "NET", [
        row("VSM_NET_001", "configVSMNet 正常配置IPv4", EP_VNET, 200,
            jd({"requestId": "uuid", "vsmId": "${env.vsmId}", "ip": "192.168.8.201", "mask": "24", "gateway": "192.168.8.1"}),
            ref_case_id="VSM_NET_001", section="7.2.3"),
        row("VSM_NET_002", "configVSMNet IPv6", EP_VNET, 200,
            jd({"requestId": "uuid", "vsmId": "${env.vsmId}", "ip": "fd00::100", "mask": "64", "gateway": "fd00::1"}),
            ref_case_id="VSM_NET_002", section="7.2.3"),
        row("VSM_NET_T01", "configVSMNet requestId空字符串", EP_VNET, 400,
            jd({"requestId": "", "vsmId": "${env.vsmId}", "ip": "192.168.10.100", "mask": "255.255.255.0", "gateway": "192.168.10.1"}),
            section="7.2.3"),
        row("VSM_NET_T02", "configVSMNet 缺requestId", EP_VNET, 400,
            jd({"vsmId": "${env.vsmId}", "ip": "192.168.10.100", "mask": "255.255.255.0", "gateway": "192.168.10.1"}),
            section="7.2.3"),
        row("VSM_NET_T04", "configVSMNet 缺ip", EP_VNET, 400,
            jd({"requestId": "uuid", "vsmId": "${env.vsmId}", "mask": "255.255.255.0", "gateway": "192.168.10.1"}),
            section="7.2.3"),
        row("VSM_NET_T05", "configVSMNet ip无效格式", EP_VNET, 400,
            jd({"requestId": "uuid", "vsmId": "${env.vsmId}", "ip": "999.999.999.999", "mask": "255.255.255.0", "gateway": "192.168.10.1"}),
            section="7.2.3"),
    ])


def gen_vsm_token_scenario():
    return _scn("SCN_VSM_TOKEN", "TOKEN", [
        row("VSM_TOKEN_001", "configVSMToken 正常配置", EP_VTOK, 200,
            jd({"requestId": "uuid", "vsmId": "${env.vsmId}", "token": "test_token_123", "tenantId": "tenant_001"}),
            ref_case_id="VSM_TOKEN_001", section="7.2.4"),
        row("VSM_TOKEN_002", "configVSMToken token=0释放", EP_VTOK, 200,
            jd({"requestId": "uuid", "vsmId": "${env.vsmId}", "token": "0", "tenantId": "tenant_001"}),
            ref_case_id="VSM_TOKEN_002", section="7.2.4"),
        row("VSM_TOKEN_T01", "configVSMToken requestId空字符串", EP_VTOK, 400,
            jd({"requestId": "", "vsmId": "${env.vsmId}", "token": "test_token", "tenantId": "tenant_001"}),
            section="7.2.4"),
        row("VSM_TOKEN_T02", "configVSMToken 缺requestId", EP_VTOK, 400,
            jd({"vsmId": "${env.vsmId}", "token": "test_token", "tenantId": "tenant_001"}),
            section="7.2.4"),
        row("VSM_TOKEN_T04", "configVSMToken token空字符串", EP_VTOK, 400,
            jd({"requestId": "uuid", "vsmId": "${env.vsmId}", "token": "", "tenantId": "tenant_001"}),
            section="7.2.4"),
        row("VSM_TOKEN_T05", "configVSMToken 缺token字段", EP_VTOK, 400,
            jd({"requestId": "uuid", "vsmId": "${env.vsmId}", "tenantId": "tenant_001"}),
            section="7.2.4"),
    ])


def gen_vsm_verify_token():
    """配置token后通过getVSMInfo回查验证"""
    return [
        row("VERIFY_VTOKEN_001", "配置VSM Token用特定值", EP_VTOK, 200,
            jd({"requestId": "uuid", "vsmId": "${env.vsmId}", "token": "VERIFY_TOKEN_V2", "tenantId": "verify_tenant_001"}),
            scenario_id="SCN_VERIFY_VTOKEN", step=1, step_type="setup", section="7.2.4"),
        row("VERIFY_VTOKEN_002", "getVSMInfo查询验证token生效", EP_VSM, 200,
            jd({"requestId": "uuid", "oprType": "getinfo", "vsmId": "${env.vsmId}"}),
            assert_rules=ok_with(
                {"type": "json_contains", "key": "result.token", "value": "VERIFY_TOKEN_V2"},
            ),
            scenario_id="SCN_VERIFY_VTOKEN", step=2, step_type="verify", section="7.2.1"),
    ]


def gen_vsm_export_scenario():
    return _scn("SCN_VSM_EXPORT", "EXPORT", [
        row("VSM_EXPORT_001", "exportVSM 正常导出", EP_VIMG, 200,
            jd({"requestId": "uuid", "oprType": "export", "vsmId": "${env.vsmId}", "callbackUrl": CALLBACK_URL}),
            ref_case_id="VSM_EXPORT_001", section="7.2.5"),
        row("VSM_EXPORT_T01", "exportVSM requestId空字符串", EP_VIMG, 400,
            jd({"requestId": "", "oprType": "export", "vsmId": "${env.vsmId}", "callbackUrl": CALLBACK_URL}),
            section="7.2.5"),
        row("VSM_EXPORT_T03", "exportVSM 缺callbackUrl", EP_VIMG, 400,
            jd({"requestId": "uuid", "oprType": "export", "vsmId": "${env.vsmId}"}),
            section="7.2.5"),
        row("VSM_EXPORT_T04", "exportVSM 缺requestId", EP_VIMG, 400,
            jd({"oprType": "export", "vsmId": "${env.vsmId}", "callbackUrl": CALLBACK_URL}),
            section="7.2.5"),
        row("VSM_EXPORT_T05", "exportVSM oprType缺失", EP_VIMG, 400,
            jd({"requestId": "uuid", "vsmId": "${env.vsmId}", "callbackUrl": CALLBACK_URL}),
            section="7.2.5"),
        row("VSM_EXPORT_T06", "exportVSM callbackUrl空字符串", EP_VIMG, 400,
            jd({"requestId": "uuid", "oprType": "export", "vsmId": "${env.vsmId}", "callbackUrl": ""}),
            section="7.2.5"),
    ])


def gen_vsm_import_scenario():
    return _scn("SCN_VSM_IMPORT", "IMPORT", [
        row("VSM_IMPORT_001", "importVSM 正常导入", EP_VIMG, 200,
            jd(VSM_IMP_BASE), ref_case_id="VSM_IMPORT_001", section="7.2.6"),
        row("VSM_IMPORT_T01", "importVSM requestId空字符串", EP_VIMG, 400,
            jd({**VSM_IMP_BASE, "requestId": ""}),
            section="7.2.6"),
        row("VSM_IMPORT_T03", "importVSM 缺alg", EP_VIMG, 400,
            jd({k: v for k, v in VSM_IMP_BASE.items() if k != "alg"}),
            section="7.2.6"),
        row("VSM_IMPORT_T04", "importVSM 缺requestId", EP_VIMG, 400,
            jd({k: v for k, v in VSM_IMP_BASE.items() if k != "requestId"}),
            section="7.2.6"),
        row("VSM_IMPORT_T05", "importVSM 缺imageUrl", EP_VIMG, 400,
            jd({k: v for k, v in VSM_IMP_BASE.items() if k != "imageUrl"}),
            section="7.2.6"),
        row("VSM_IMPORT_T06", "importVSM 缺sign", EP_VIMG, 400,
            jd({k: v for k, v in VSM_IMP_BASE.items() if k != "sign"}),
            section="7.2.6"),
        row("VSM_IMPORT_T07", "importVSM 缺callbackUrl", EP_VIMG, 400,
            jd({k: v for k, v in VSM_IMP_BASE.items() if k != "callbackUrl"}),
            section="7.2.6"),
        row("VSM_IMPORT_T08", "importVSM oprType缺失", EP_VIMG, 400,
            jd({k: v for k, v in VSM_IMP_BASE.items() if k != "oprType"}),
            section="7.2.6"),
        row("VSM_IMPORT_T09", "importVSM alg无效值", EP_VIMG, 400,
            jd({**VSM_IMP_BASE, "alg": "INVALID_ALG"}),
            section="7.2.6"),
    ])


def gen_vsm_start_scenario():
    return _scn("SCN_VSM_START", "START", [
        row("VSM_START_001", "startVSM 正常启动", EP_VSM, 200,
            jd({"requestId": "uuid", "oprType": "start", "vsmId": "${env.vsmId}", "callbackUrl": CALLBACK_URL}),
            ref_case_id="VSM_START_001", section="7.2.7"),
        row("VSM_START_T01", "startVSM requestId空字符串", EP_VSM, 400,
            jd({"requestId": "", "oprType": "start", "vsmId": "${env.vsmId}", "callbackUrl": CALLBACK_URL}),
            section="7.2.7"),
        row("VSM_START_T03", "startVSM oprType缺失", EP_VSM, 400,
            jd({"requestId": "uuid", "vsmId": "${env.vsmId}", "callbackUrl": CALLBACK_URL}),
            section="7.2.7"),
        row("VSM_START_T04", "startVSM 缺requestId", EP_VSM, 400,
            jd({"oprType": "start", "vsmId": "${env.vsmId}", "callbackUrl": CALLBACK_URL}),
            section="7.2.7"),
        row("VSM_START_T05", "startVSM 缺callbackUrl", EP_VSM, 400,
            jd({"requestId": "uuid", "oprType": "start", "vsmId": "${env.vsmId}"}),
            section="7.2.7"),
    ])


def gen_vsm_stop_scenario():
    return _scn("SCN_VSM_STOP", "STOP", [
        row("VSM_STOP_001", "stopVSM 正常停止", EP_VSM, 200,
            jd({"requestId": "uuid", "oprType": "stop", "vsmId": "${env.vsmId}", "callbackUrl": CALLBACK_URL}),
            ref_case_id="VSM_STOP_001", section="7.2.8"),
        row("VSM_STOP_T01", "stopVSM requestId空字符串", EP_VSM, 400,
            jd({"requestId": "", "oprType": "stop", "vsmId": "${env.vsmId}", "callbackUrl": CALLBACK_URL}),
            section="7.2.8"),
        row("VSM_STOP_T03", "stopVSM 缺requestId", EP_VSM, 400,
            jd({"oprType": "stop", "vsmId": "${env.vsmId}", "callbackUrl": CALLBACK_URL}),
            section="7.2.8"),
        row("VSM_STOP_T04", "stopVSM oprType缺失", EP_VSM, 400,
            jd({"requestId": "uuid", "vsmId": "${env.vsmId}", "callbackUrl": CALLBACK_URL}),
            section="7.2.8"),
        row("VSM_STOP_T05", "stopVSM 缺callbackUrl", EP_VSM, 400,
            jd({"requestId": "uuid", "oprType": "stop", "vsmId": "${env.vsmId}"}),
            section="7.2.8"),
    ])


def gen_vsm_restart_scenario():
    return _scn("SCN_VSM_RESTART", "RESTART", [
        row("VSM_RESTART_001", "restartVSM 正常重启", EP_VSM, 200,
            jd({"requestId": "uuid", "oprType": "restart", "vsmId": "${env.vsmId}", "callbackUrl": CALLBACK_URL}),
            ref_case_id="VSM_RESTART_001", section="7.2.9"),
        row("VSM_RESTART_T01", "restartVSM requestId空字符串", EP_VSM, 400,
            jd({"requestId": "", "oprType": "restart", "vsmId": "${env.vsmId}", "callbackUrl": CALLBACK_URL}),
            section="7.2.9"),
        row("VSM_RESTART_T03", "restartVSM 缺requestId", EP_VSM, 400,
            jd({"oprType": "restart", "vsmId": "${env.vsmId}", "callbackUrl": CALLBACK_URL}),
            section="7.2.9"),
        row("VSM_RESTART_T04", "restartVSM oprType缺失", EP_VSM, 400,
            jd({"requestId": "uuid", "vsmId": "${env.vsmId}", "callbackUrl": CALLBACK_URL}),
            section="7.2.9"),
        row("VSM_RESTART_T05", "restartVSM 缺callbackUrl", EP_VSM, 400,
            jd({"requestId": "uuid", "oprType": "restart", "vsmId": "${env.vsmId}"}),
            section="7.2.9"),
    ])


def gen_vsm_reset_scenario():
    return _scn("SCN_VSM_RESET", "RESET", [
        row("VSM_RESET_001", "resetVSM 正常重置", EP_VSM, 200,
            jd({"requestId": "uuid", "oprType": "reset", "vsmId": "${env.vsmId}", "callbackUrl": CALLBACK_URL}),
            ref_case_id="VSM_RESET_001", section="7.2.10"),
        row("VSM_RESET_T01", "resetVSM requestId空字符串", EP_VSM, 400,
            jd({"requestId": "", "oprType": "reset", "vsmId": "${env.vsmId}", "callbackUrl": CALLBACK_URL}),
            section="7.2.10"),
        row("VSM_RESET_T03", "resetVSM 缺requestId", EP_VSM, 400,
            jd({"oprType": "reset", "vsmId": "${env.vsmId}", "callbackUrl": CALLBACK_URL}),
            section="7.2.10"),
        row("VSM_RESET_T04", "resetVSM oprType缺失", EP_VSM, 400,
            jd({"requestId": "uuid", "vsmId": "${env.vsmId}", "callbackUrl": CALLBACK_URL}),
            section="7.2.10"),
        row("VSM_RESET_T05", "resetVSM 缺callbackUrl", EP_VSM, 400,
            jd({"requestId": "uuid", "oprType": "reset", "vsmId": "${env.vsmId}"}),
            section="7.2.10"),
    ])


def gen_vsm_upgrade_scenario():
    return _scn("SCN_VSM_UPGRADE", "UPGRADE", [
        row("VSM_UPGRADE_001", "upgradeVSM 正常升级", EP_VSM, 200,
            jd(VSM_UPG_BASE), ref_case_id="VSM_UPGRADE_001", section="7.2.11"),
        row("VSM_UPGRADE_T01", "upgradeVSM requestId空字符串", EP_VSM, 400,
            jd({**VSM_UPG_BASE, "requestId": ""}),
            section="7.2.11"),
        row("VSM_UPGRADE_T03", "upgradeVSM 缺packVersion", EP_VSM, 400,
            jd({k: v for k, v in VSM_UPG_BASE.items() if k != "packVersion"}),
            section="7.2.11"),
        row("VSM_UPGRADE_T04", "upgradeVSM 缺callbackUrl", EP_VSM, 400,
            jd({k: v for k, v in VSM_UPG_BASE.items() if k != "callbackUrl"}),
            section="7.2.11"),
        row("VSM_UPGRADE_T05", "upgradeVSM 缺requestId", EP_VSM, 400,
            jd({k: v for k, v in VSM_UPG_BASE.items() if k != "requestId"}),
            section="7.2.11"),
        row("VSM_UPGRADE_T06", "upgradeVSM 缺packUrl", EP_VSM, 400,
            jd({k: v for k, v in VSM_UPG_BASE.items() if k != "packUrl"}),
            section="7.2.11"),
        row("VSM_UPGRADE_T07", "upgradeVSM 缺alg", EP_VSM, 400,
            jd({k: v for k, v in VSM_UPG_BASE.items() if k != "alg"}),
            section="7.2.11"),
        row("VSM_UPGRADE_T08", "upgradeVSM 缺sign", EP_VSM, 400,
            jd({k: v for k, v in VSM_UPG_BASE.items() if k != "sign"}),
            section="7.2.11"),
        row("VSM_UPGRADE_T09", "upgradeVSM oprType缺失", EP_VSM, 400,
            jd({k: v for k, v in VSM_UPG_BASE.items() if k != "oprType"}),
            section="7.2.11"),
    ])


# ================================================================
# 7.2.12-7.2.16 预留接口（简单用例保证调通）
# ================================================================

def gen_vsm_create():
    """7.2.12 createVSM（预留）"""
    EP = "/api/1.0/vsm"
    return [
        row("VSM_CREATE_001", "createVSM 正常创建(预留)", EP, 200,
            jd({"requestId": "uuid", "oprType": "create", "flavor": 1, "callbackUrl": CALLBACK_URL}),
            section="7.2.12"),
        row("VSM_CREATE_T01", "createVSM 缺requestId", EP, 400,
            jd({"oprType": "create", "flavor": 1, "callbackUrl": CALLBACK_URL}),
            section="7.2.12"),
    ]


def gen_vsm_destroy():
    """7.2.13 deleteVSM（预留）"""
    EP = "/api/1.0/vsm"
    return [
        row("VSM_DESTROY_001", "deleteVSM 正常删除(预留)", EP, 200,
            jd({"requestId": "uuid", "oprType": "destroy", "vsmId": "${env.vsmId}", "callbackUrl": CALLBACK_URL}),
            section="7.2.13"),
        row("VSM_DESTROY_T01", "deleteVSM 缺vsmId", EP, 400,
            jd({"requestId": "uuid", "oprType": "destroy", "callbackUrl": CALLBACK_URL}),
            section="7.2.13"),
    ]


def gen_vsm_mac():
    """7.2.14 configVSMMac（预留）"""
    EP = "/api/1.0/vsm/mac"
    return [
        row("VSM_MAC_001", "configVSMMac 正常配置(预留)", EP, 200,
            jd({"requestId": "uuid", "vsmId": "${env.vsmId}", "mac": "ba:ca:c1:d2:e3:e4"}),
            section="7.2.14"),
        row("VSM_MAC_T01", "configVSMMac 缺mac", EP, 400,
            jd({"requestId": "uuid", "vsmId": "${env.vsmId}"}),
            section="7.2.14"),
    ]


def gen_vsm_vlan():
    """7.2.15 configVSMVlan（预留）"""
    EP = "/api/1.0/vsm/vlan"
    return [
        row("VSM_VLAN_001", "configVSMVlan modify(预留)", EP, 200,
            jd({"requestId": "uuid", "oprType": "modify", "vsmId": "${env.vsmId}", "vlanType": "0", "vlanId": "100"}),
            section="7.2.15"),
        row("VSM_VLAN_002", "configVSMVlan delete(预留)", EP, 200,
            jd({"requestId": "uuid", "oprType": "delete", "vsmId": "${env.vsmId}", "vlanType": "0", "vlanId": "100"}),
            section="7.2.15"),
    ]


def gen_vsm_del_network():
    """7.2.16 deleteVSMNetwork（预留）"""
    EP = "/api/1.0/vsm/network"
    return [
        row("VSM_DELNET_001", "deleteVSMNetwork 正常删除(预留)", EP, 200,
            jd({"requestId": "uuid", "vsmId": "${env.vsmId}"}), method="DELETE",
            section="7.2.16"),
        row("VSM_DELNET_T01", "deleteVSMNetwork 缺vsmId", EP, 400,
            jd({"requestId": "uuid"}), method="DELETE",
            section="7.2.16"),
    ]


# ================================================================
# 7.2.17 查看设备详情
# ================================================================

def gen_vsm_device_info():
    """7.2.17 查看设备详情 — 与7.1.18类似，VSM获取CHSM的sn/mac/hwAuthCode"""
    EP = "/api/1.0/vsm"
    return [
        row("VSM_DEVICE_001", "getDeviceInfo 正常获取设备详情", EP, 200,
            jd({"requestId": "uuid", "oprType": "getDeviceinfo"}),
            assert_rules=ok_with(
                {"type": "json_contains", "key": "result.version"},
                {"type": "json_contains", "key": "result.sn"},
                {"type": "json_contains", "key": "result.mac"},
            ),
            ref_case_id="VSM_DEVICE_001", section="7.2.17"),
        row("VSM_DEVICE_T01", "getDeviceInfo requestId空字符串", EP, 400,
            jd({"requestId": "", "oprType": "getDeviceinfo"}),
            section="7.2.17"),
        row("VSM_DEVICE_T02", "getDeviceInfo 缺requestId", EP, 400,
            jd({"oprType": "getDeviceinfo"}),
            section="7.2.17"),
        row("VSM_DEVICE_T03", "getDeviceInfo oprType缺失", EP, 400,
            jd({"requestId": "uuid"}),
            section="7.2.17"),
        row("VSM_DEVICE_T04", "getDeviceInfo oprType空字符串", EP, 400,
            jd({"requestId": "uuid", "oprType": ""}),
            section="7.2.17"),
    ]


def gen_vsm_standalone():
    """不需要 vsmId 的 VSM 独立用例 (无场景编排)"""
    return [
        row("VSM_INFO_002", "getVSMInfo vsmId不存在", EP_VSM, 400,
            jd({"requestId": "uuid", "oprType": "getinfo", "vsmId": "non_existent_id"}),
            ref_case_id="VSM_INFO_002", section="7.2.1"),
        row("VSM_INFO_003", "getVSMInfo 缺vsmId", EP_VSM, 400,
            jd({"requestId": "uuid", "oprType": "getinfo"}),
            ref_case_id="VSM_INFO_003", section="7.2.1"),
        row("VSM_INFO_T04", "getVSMInfo vsmId空字符串", EP_VSM, 400,
            jd({"requestId": "uuid", "oprType": "getinfo", "vsmId": ""}),
            section="7.2.1"),
        row("VSM_NET_T03", "configVSMNet 缺vsmId", EP_VNET, 400,
            jd({"requestId": "uuid", "ip": "192.168.10.100", "mask": "255.255.255.0", "gateway": "192.168.10.1"}),
            section="7.2.3"),
        row("VSM_TOKEN_T03", "configVSMToken 缺vsmId", EP_VTOK, 400,
            jd({"requestId": "uuid", "token": "test_token", "tenantId": "tenant_001"}),
            section="7.2.4"),
        row("VSM_EXPORT_T02", "exportVSM 缺vsmId", EP_VIMG, 400,
            jd({"requestId": "uuid", "oprType": "export", "callbackUrl": CALLBACK_URL}),
            section="7.2.5"),
        row("VSM_IMPORT_T02", "importVSM 缺vsmId", EP_VIMG, 400,
            jd({k: v for k, v in VSM_IMP_BASE.items() if k != "vsmId"}),
            section="7.2.6"),
        row("VSM_START_002", "startVSM vsmId不存在", EP_VSM, 400,
            jd({"requestId": "uuid", "oprType": "start", "vsmId": "non_existent_id", "callbackUrl": CALLBACK_URL}),
            ref_case_id="VSM_START_002", section="7.2.7"),
        row("VSM_START_T02", "startVSM 缺vsmId", EP_VSM, 400,
            jd({"requestId": "uuid", "oprType": "start", "callbackUrl": CALLBACK_URL}),
            section="7.2.7"),
        row("VSM_STOP_002", "stopVSM vsmId不存在", EP_VSM, 400,
            jd({"requestId": "uuid", "oprType": "stop", "vsmId": "non_existent_id", "callbackUrl": CALLBACK_URL}),
            ref_case_id="VSM_STOP_002", section="7.2.8"),
        row("VSM_STOP_T02", "stopVSM 缺vsmId", EP_VSM, 400,
            jd({"requestId": "uuid", "oprType": "stop", "callbackUrl": CALLBACK_URL}),
            section="7.2.8"),
        row("VSM_RESTART_002", "restartVSM vsmId不存在", EP_VSM, 400,
            jd({"requestId": "uuid", "oprType": "restart", "vsmId": "non_existent_id", "callbackUrl": CALLBACK_URL}),
            ref_case_id="VSM_RESTART_002", section="7.2.9"),
        row("VSM_RESTART_T02", "restartVSM 缺vsmId", EP_VSM, 400,
            jd({"requestId": "uuid", "oprType": "restart", "callbackUrl": CALLBACK_URL}),
            section="7.2.9"),
        row("VSM_RESET_002", "resetVSM vsmId不存在", EP_VSM, 400,
            jd({"requestId": "uuid", "oprType": "reset", "vsmId": "non_existent_id", "callbackUrl": CALLBACK_URL}),
            ref_case_id="VSM_RESET_002", section="7.2.10"),
        row("VSM_RESET_T02", "resetVSM 缺vsmId", EP_VSM, 400,
            jd({"requestId": "uuid", "oprType": "reset", "callbackUrl": CALLBACK_URL}),
            section="7.2.10"),
        row("VSM_UPGRADE_T02", "upgradeVSM 缺vsmId", EP_VSM, 400,
            jd({k: v for k, v in VSM_UPG_BASE.items() if k != "vsmId"}),
            section="7.2.11"),
    ]


# ================================================================
# 7.1.2 getCHSMStatus (guest)
# ================================================================

def gen_getCHSMStatus():
    EP = "/api/1.0/chsm/status"
    return [
        row("CHSM_STATUS_001", "getCHSMStatus 正常获取", EP, 200,
            params="requestId=uuid", method="GET",
            assert_rules=ok_with({"type": "json_contains", "key": "result.status"}),
            ref_case_id="CHSM_STATUS_001", section="7.1.2", auth="no"),
        row("CHSM_STATUS_T01", "getCHSMStatus requestId空字符串", EP, 400,
            params="requestId=", method="GET",
            section="7.1.2", auth="no"),
        row("CHSM_STATUS_T02", "getCHSMStatus 缺requestId", EP, 400,
            method="GET", ref_case_id="CHSM_STATUS_002",
            section="7.1.2", auth="no"),
    ]


# ================================================================
# 7.1.3 getCHSMAllStatus (guest)
# ================================================================

def gen_getCHSMAllStatus():
    EP = "/api/1.0/chsm/allstatus"
    return [
        row("CHSM_ALLSTATUS_001", "getCHSMAllStatus 正常获取", EP, 200,
            params="requestId=uuid", method="GET",
            assert_rules=ok_with({"type": "json_contains", "key": "result.hsmStatus"}),
            ref_case_id="CHSM_ALLSTATUS_001", section="7.1.3", auth="no"),
        row("CHSM_ALLSTATUS_T01", "getCHSMAllStatus requestId空字符串", EP, 400,
            params="requestId=", method="GET",
            section="7.1.3", auth="no"),
        row("CHSM_ALLSTATUS_T02", "getCHSMAllStatus 缺requestId", EP, 400,
            method="GET",
            section="7.1.3", auth="no"),
    ]


# ================================================================
# 7.2.2 getVSMStatus (guest)
# ================================================================

def gen_getVSMStatus():
    EP = "/api/1.0/vsm/status"
    return [
        row("VSM_STATUS_001", "getVSMStatus 正常获取", EP, 200,
            params="requestId=uuid&vsmId=${env.vsmId}", method="GET",
            assert_rules=ok_with({"type": "json_contains", "key": "result.status"}),
            ref_case_id="VSM_STATUS_001", section="7.2.2", auth="no"),
        row("VSM_STATUS_002", "getVSMStatus vsmId不存在", EP, 400,
            params="requestId=uuid&vsmId=non_existent_id", method="GET",
            ref_case_id="VSM_STATUS_002", section="7.2.2", auth="no"),
        row("VSM_STATUS_T01", "getVSMStatus requestId空字符串", EP, 400,
            params="requestId=&vsmId=${env.vsmId}", method="GET",
            section="7.2.2", auth="no"),
        row("VSM_STATUS_T02", "getVSMStatus 缺requestId", EP, 400,
            params="vsmId=${env.vsmId}", method="GET",
            section="7.2.2", auth="no"),
        row("VSM_STATUS_T03", "getVSMStatus 缺vsmId", EP, 400,
            params="requestId=uuid", method="GET",
            section="7.2.2", auth="no"),
        row("VSM_STATUS_T04", "getVSMStatus vsmId空字符串", EP, 400,
            params="requestId=uuid&vsmId=", method="GET",
            section="7.2.2", auth="no"),
    ]


# ================================================================
# 配置→查询验证场景（配置后通过getCHSMInfo回查确认生效）
# ================================================================

def gen_config_verify_scenarios():
    """配置接口修改值后，通过getCHSMInfo查询验证字段已变更"""
    EP_CHSM = "/api/1.0/chsm"
    rows = []

    # --- SCN_VERIFY_NET: 配网络→查验证（用真实有效的网络配置）---
    rows.extend([
        row("VERIFY_NET_001", "配置网络(DNS+网口)用真实值", "/api/1.0/chsm/network", 200,
            jd({"requestId": "uuid",
                "dnsList": ["114.114.114.114", "192.168.8.1"],
                "netAddrs": [{"name": "eth0", "ip": "192.168.8.121", "mask": "255.255.255.0", "gateway": "192.168.8.1"}]}),
            scenario_id="SCN_VERIFY_NET", step=1, step_type="setup", section="7.1.4"),
        row("VERIFY_NET_002", "getCHSMInfo查询验证网络配置生效", EP_CHSM, 200,
            jd({"requestId": "uuid", "oprType": "getinfo"}),
            assert_rules=ok_with(
                {"type": "json_contains", "key": "result.dnsList[*]", "value": "114.114.114.114"},
                {"type": "json_contains", "key": "result.netAddrs[0].ip", "value": "192.168.8.121"},
            ),
            scenario_id="SCN_VERIFY_NET", step=2, step_type="verify", section="7.1.1"),
    ])

    # --- SCN_VERIFY_NTP: 配NTP不同syncPeriod→回查验证 ---
    rows.extend([
        row("VERIFY_NTP_001", "配置NTP syncPeriod=1(最小值)", "/api/1.0/chsm/ntp", 200,
            jd({"requestId": "uuid", "addr": "ntp.aliyun.com", "syncPeriod": 1}),
            scenario_id="SCN_VERIFY_NTP", step=1, step_type="test", section="7.1.5"),
        row("VERIFY_NTP_002", "查询验证syncPeriod=1", EP_CHSM, 200,
            jd({"requestId": "uuid", "oprType": "getinfo"}),
            assert_rules=ok_with(
                {"type": "json_contains", "key": "result.ntpAddr", "value": "ntp.aliyun.com"},
                {"type": "json_contains", "key": "result.ntpSyncPeriod", "value": 1},
            ),
            scenario_id="SCN_VERIFY_NTP", step=2, step_type="verify", section="7.1.1"),
        row("VERIFY_NTP_003", "配置NTP syncPeriod=30(常规值)", "/api/1.0/chsm/ntp", 200,
            jd({"requestId": "uuid", "addr": "ntp.aliyun.com", "syncPeriod": 30}),
            scenario_id="SCN_VERIFY_NTP", step=3, step_type="test", section="7.1.5"),
        row("VERIFY_NTP_004", "查询验证syncPeriod=30", EP_CHSM, 200,
            jd({"requestId": "uuid", "oprType": "getinfo"}),
            assert_rules=ok_with(
                {"type": "json_contains", "key": "result.ntpSyncPeriod", "value": 30},
            ),
            scenario_id="SCN_VERIFY_NTP", step=4, step_type="verify", section="7.1.1"),
        row("VERIFY_NTP_005", "配置NTP syncPeriod=1440(一天)", "/api/1.0/chsm/ntp", 200,
            jd({"requestId": "uuid", "addr": "ntp.aliyun.com", "syncPeriod": 1440}),
            scenario_id="SCN_VERIFY_NTP", step=5, step_type="test", section="7.1.5"),
        row("VERIFY_NTP_006", "查询验证syncPeriod=1440", EP_CHSM, 200,
            jd({"requestId": "uuid", "oprType": "getinfo"}),
            assert_rules=ok_with(
                {"type": "json_contains", "key": "result.ntpSyncPeriod", "value": 1440},
            ),
            scenario_id="SCN_VERIFY_NTP", step=6, step_type="verify", section="7.1.1"),
        row("VERIFY_NTP_007", "配置NTP syncPeriod=60(恢复默认)", "/api/1.0/chsm/ntp", 200,
            jd({"requestId": "uuid", "addr": "ntp.aliyun.com", "syncPeriod": 60}),
            scenario_id="SCN_VERIFY_NTP", step=7, step_type="teardown", section="7.1.5"),
    ])

    # --- SCN_VERIFY_UPLOAD: 配上传地址→查验证 ---
    VERIFY_UPLOAD = "http://127.0.0.1:9443/images"
    rows.extend([
        row("VERIFY_UPLOAD_001", "配置影像上传地址用特定值", "/api/1.0/chsm/imageuploader", 200,
            jd({"requestId": "uuid", "url": VERIFY_UPLOAD}),
            scenario_id="SCN_VERIFY_UPLOAD", step=1, step_type="setup", section="7.1.6"),
        row("VERIFY_UPLOAD_002", "getCHSMInfo查询验证上传地址生效", EP_CHSM, 200,
            jd({"requestId": "uuid", "oprType": "getinfo"}),
            assert_rules=ok_with(
                {"type": "json_contains", "key": "result.imageUploaderUrl", "value": VERIFY_UPLOAD},
            ),
            scenario_id="SCN_VERIFY_UPLOAD", step=2, step_type="verify", section="7.1.1"),
    ])

    # --- SCN_VERIFY_ALARM: 配告警地址→查验证 ---
    VERIFY_ALARM = "http://10.99.99.99:9999/verify_alarm"
    VERIFY_MONITOR = "http://10.99.99.99:9999/verify_monitor"
    rows.extend([
        row("VERIFY_ALARM_001", "配置告警地址用特定值", "/api/1.0/chsm/alarmaddress", 200,
            jd({"requestId": "uuid", "url": VERIFY_ALARM, "monitoringUrl": VERIFY_MONITOR}),
            scenario_id="SCN_VERIFY_ALARM", step=1, step_type="setup", section="7.1.14"),
        row("VERIFY_ALARM_002", "getCHSMInfo查询验证告警地址生效", EP_CHSM, 200,
            jd({"requestId": "uuid", "oprType": "getinfo"}),
            assert_rules=ok_with(
                {"type": "json_contains", "key": "result.alarmAddress", "value": VERIFY_ALARM},
            ),
            scenario_id="SCN_VERIFY_ALARM", step=2, step_type="verify", section="7.1.1"),
    ])

    # --- SCN_VERIFY_TOKEN: 配Token→查验证 ---
    VERIFY_TOKEN = "VERIFY_TOKEN_VALUE_20260601"
    rows.extend([
        row("VERIFY_TOKEN_001", "配置cloudToken用特定值", "/api/1.0/chsm/cloudtoken", 200,
            jd({"requestId": "uuid", "cloudToken": VERIFY_TOKEN}),
            scenario_id="SCN_VERIFY_TOKEN", step=1, step_type="setup", section="7.1.15"),
        row("VERIFY_TOKEN_002", "getCHSMInfo查询验证cloudToken生效", EP_CHSM, 200,
            jd({"requestId": "uuid", "oprType": "getinfo"}),
            assert_rules=ok_with(
                {"type": "json_contains", "key": "result.cloudToken", "value": VERIFY_TOKEN},
            ),
            scenario_id="SCN_VERIFY_TOKEN", step=2, step_type="verify", section="7.1.1"),
    ])

    # --- SCN_VERIFY_SYSLOG: 配syslog→查看效果→配logserver→验证URL ---
    VERIFY_LOG_URL = "http://10.99.99.99:9999/verify_log"
    rows.extend([
        row("VERIFY_SYSLOG_001", "配置syslog类型(观察查询效果)", "/api/1.0/chsm/loguploader", 200,
            jd({"requestId": "uuid", "logServerType": "syslog", "logServerAddress": "10.99.99.99:5140"}),
            scenario_id="SCN_VERIFY_SYSLOG", step=1, step_type="setup", section="7.1.7"),
        row("VERIFY_SYSLOG_002", "getCHSMInfo查询syslog配置效果", EP_CHSM, 200,
            jd({"requestId": "uuid", "oprType": "getinfo"}),
            assert_rules=ok_with(
                {"type": "json_contains", "key": "result.sysLogUrl"},
            ),
            scenario_id="SCN_VERIFY_SYSLOG", step=2, step_type="verify", section="7.1.1"),
        row("VERIFY_SYSLOG_003", "配置logserver类型用URL", "/api/1.0/chsm/loguploader", 200,
            jd({"requestId": "uuid", "logServerType": "logserver", "logServerAddress": VERIFY_LOG_URL}),
            scenario_id="SCN_VERIFY_SYSLOG", step=3, step_type="test", section="7.1.7"),
        row("VERIFY_SYSLOG_004", "getCHSMInfo查询验证logserver地址生效", EP_CHSM, 200,
            jd({"requestId": "uuid", "oprType": "getinfo"}),
            assert_rules=ok_with(
                {"type": "json_contains", "key": "result.sysLogUrl", "value": VERIFY_LOG_URL},
            ),
            scenario_id="SCN_VERIFY_SYSLOG", step=4, step_type="verify", section="7.1.1"),
    ])

    # --- SCN_VERIFY_MOOC: 配MOOC地址→查验证 ---
    rows.extend([
        row("VERIFY_MOOC_001", "配置MOOC对接地址用特定值", "/api/1.0/chsm/moocaddress", 200,
            jd({"requestId": "uuid", "ocIp": "10.77.77.77", "ocProtocol": "https",
                "ocPort": "26635", "resourceUrl": "/rest/cloudInfra/v1/resource",
                "performanceUrl": "/rest/cloudInfra/v1/performance",
                "bizRegionNativeId": "region-verify-v2", "cloudInfraType": "FusionSphere",
                "dewServiceHost": "kms.verify.myhuaweicloud.com"}),
            scenario_id="SCN_VERIFY_MOOC", step=1, step_type="setup", section="7.1.16"),
        row("VERIFY_MOOC_002", "getCHSMInfo查询验证MOOC配置生效", EP_CHSM, 200,
            jd({"requestId": "uuid", "oprType": "getinfo"}),
            assert_rules=ok_with(
                {"type": "json_contains", "key": "result.moOcAddress.ocIp", "value": "10.77.77.77"},
                {"type": "json_contains", "key": "result.moOcAddress.bizRegionNativeId", "value": "region-verify-v2"},
            ),
            scenario_id="SCN_VERIFY_MOOC", step=2, step_type="verify", section="7.1.1"),
    ])

    return rows


# ================================================================
# 认证异常测试
# ================================================================

def gen_auth_test():
    ERR401 = json.dumps([{"type": "status_code", "expected_code": 401}], ensure_ascii=False)
    return [
        row("AUTH_401_001", "getCHSMInfo 不带认证头(trusted接口)", "/api/1.0/chsm", 401,
            json_data=jd({"requestId": "uuid", "oprType": "getinfo"}),
            assert_rules=ERR401, ref_case_id="CHSM_INFO_005", section="7.1.1", auth="no"),
    ]


# ================================================================
# 主流程
# ================================================================

def write_sheet(ws, data_rows):
    for ci, h in enumerate(HEADERS, 1):
        ws.cell(1, ci, h)
    for ri, data in enumerate(data_rows, 2):
        for ci, h in enumerate(HEADERS, 1):
            val = data.get(h)
            if val is not None:
                ws.cell(ri, ci, val)


def main():
    xlsx = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "test_data.xlsx")
    wb = openpyxl.load_workbook(xlsx)

    # 删除旧 sheet
    for name in ["trusted", "guest", "testdata", "chsm", "vsm"]:
        if name in wb.sheetnames:
            del wb[name]

    # ================================================================
    # chsm: 7.1 CHSM 管理接口
    # ================================================================
    td_data = []
    # --- 前置: 配置公钥 ---
    td_data.extend([
        row("INIT_RSA_POST_001", "首次配置1个RSA公钥", "/api/1.0/chsm/authpk", 200,
            jd({"requestId": "uuid", "algorithm": "RSAWithSHA256", "pks": ["${keys.rsa.key1.public_key_pem}"]}),
            ref_case_id="AUTH_PK_001", section="7.3.2", auth="yes"),
        row("INIT_RSA_GET_002", "查询验证1个指纹", "/api/1.0/chsm/authpk", 200,
            params="requestId=uuid", method="GET",
            assert_rules=json.dumps([
                {"type": "status_code", "expected_code": 200},
                {"type": "json_contains", "key": "result.algorithm", "value": "sha256"},
                {"type": "json_contains", "key": "result.fingerprints[0]"},
                {"type": "json_not_contains", "key": "result.fingerprints[1]"},
            ], ensure_ascii=False),
            ref_case_id="AUTH_PK_002", section="7.3.1", auth="no"),
    ])
    for gen_fn in [
        gen_getCHSMInfo,              # 7.1.1
        gen_getCHSMStatus,            # 7.1.2 (guest)
        gen_getCHSMAllStatus,         # 7.1.3 (guest)
        gen_configCHSMNet,            # 7.1.4
        gen_configCHSMNtp,            # 7.1.5
        gen_configCHSMUploadAddress,  # 7.1.6
        gen_configSyslogAddr,         # 7.1.7
        gen_exportCHSM,               # 7.1.8
        gen_importCHSM,               # 7.1.9
        gen_upgradeCHSM,              # 7.1.10
        gen_restartCHSM,              # 7.1.11
        gen_backupCHSM,               # 7.1.12
        gen_restoreCHSM,              # 7.1.13
        gen_configCHSMAlarmAddress,   # 7.1.14
        gen_configCHSMToken,          # 7.1.15
        gen_configCHSMMOOCAddress,    # 7.1.16
        gen_getCHSMDebugInfo,         # 7.1.17 (guest)
        gen_getCHSMDeviceInfo,        # 7.1.18
    ]:
        td_data.extend(gen_fn())
    # --- 配置→查询验证场景 ---
    td_data.extend(gen_config_verify_scenarios())
    # --- 认证异常 ---
    td_data.extend(gen_auth_test())
    # --- 尾部: 清空公钥 ---
    td_data.extend([
        row("PK_CLEAR_H001", "清空所有公钥", "/api/1.0/chsm/authpk", 200,
            jd({"requestId": "uuid"}), method="DELETE",
            section="7.3.3", auth="yes"),
        row("PK_GET_H002", "清空后查询验证为空", "/api/1.0/chsm/authpk", 200,
            params="requestId=uuid", method="GET",
            assert_rules=json.dumps([
                {"type": "status_code", "expected_code": 200},
                {"type": "json_not_contains", "key": "result.fingerprints[0]"},
            ], ensure_ascii=False),
            section="7.3.1", auth="no"),
    ])

    TESTDATA_ENABLED = {
        "7.3.1", "7.3.2", "7.3.3",
        "7.1.1", "7.1.2", "7.1.3", "7.1.4", "7.1.5", "7.1.6", "7.1.7",
        "7.1.8", "7.1.9", "7.1.10", "7.1.11", "7.1.12", "7.1.13",
        "7.1.14", "7.1.15", "7.1.16", "7.1.17", "7.1.18",
    }
    TESTDATA_APPROVED = {
        "7.3.1", "7.3.2", "7.3.3",
        "7.1.1", "7.1.2", "7.1.3", "7.1.4", "7.1.5", "7.1.6", "7.1.7",
        "7.1.10", "7.1.11",
        "7.1.14", "7.1.15", "7.1.16", "7.1.17", "7.1.18",
    }
    for r in td_data:
        sec = r.get("section", "")
        if sec not in TESTDATA_ENABLED:
            r["enabled"] = "no"
        if sec in TESTDATA_APPROVED:
            r["approved"] = "yes"

    ws_td = wb.create_sheet("chsm")
    write_sheet(ws_td, td_data)

    td_enabled = sum(1 for r in td_data if r.get("enabled", "yes") == "yes")
    print(f"chsm: {len(td_data)} 行 ({td_enabled} enabled, {len(td_data) - td_enabled} disabled)")

    # ================================================================
    # vsm: 7.2 VSM 管理接口
    # ================================================================
    vsm_data = []
    # --- 前置: 配置公钥 ---
    vsm_data.extend([
        row("INIT_RSA_POST_001", "首次配置1个RSA公钥", "/api/1.0/chsm/authpk", 200,
            jd({"requestId": "uuid", "algorithm": "RSAWithSHA256", "pks": ["${keys.rsa.key1.public_key_pem}"]}),
            ref_case_id="AUTH_PK_001", section="7.3.2", auth="yes"),
        row("INIT_RSA_GET_002", "查询验证1个指纹", "/api/1.0/chsm/authpk", 200,
            params="requestId=uuid", method="GET",
            assert_rules=json.dumps([
                {"type": "status_code", "expected_code": 200},
                {"type": "json_contains", "key": "result.algorithm", "value": "sha256"},
                {"type": "json_contains", "key": "result.fingerprints[0]"},
                {"type": "json_not_contains", "key": "result.fingerprints[1]"},
            ], ensure_ascii=False),
            ref_case_id="AUTH_PK_002", section="7.3.1", auth="no"),
    ])
    for vsm_fn in [
        gen_vsm_info_scenario,        # 7.2.1
        gen_getVSMStatus,             # 7.2.2 (guest)
        gen_vsm_net_scenario,         # 7.2.3
        gen_vsm_token_scenario,       # 7.2.4
        gen_vsm_verify_token,         # 7.2.4 配token→查验证
        gen_vsm_export_scenario,      # 7.2.5
        gen_vsm_import_scenario,      # 7.2.6
        gen_vsm_start_scenario,       # 7.2.7
        gen_vsm_stop_scenario,        # 7.2.8
        gen_vsm_restart_scenario,     # 7.2.9
        gen_vsm_reset_scenario,       # 7.2.10
        gen_vsm_upgrade_scenario,     # 7.2.11
        gen_vsm_create,               # 7.2.12 (预留)
        gen_vsm_destroy,              # 7.2.13 (预留)
        gen_vsm_mac,                  # 7.2.14 (预留)
        gen_vsm_vlan,                 # 7.2.15 (预留)
        gen_vsm_del_network,          # 7.2.16 (预留)
        gen_vsm_device_info,          # 7.2.17
    ]:
        vsm_data.extend(vsm_fn())
    vsm_data.extend(gen_vsm_standalone())
    # --- 尾部: 清空公钥 ---
    vsm_data.extend([
        row("PK_CLEAR_H001", "清空所有公钥", "/api/1.0/chsm/authpk", 200,
            jd({"requestId": "uuid"}), method="DELETE",
            section="7.3.3", auth="yes"),
        row("PK_GET_H002", "清空后查询验证为空", "/api/1.0/chsm/authpk", 200,
            params="requestId=uuid", method="GET",
            assert_rules=json.dumps([
                {"type": "status_code", "expected_code": 200},
                {"type": "json_not_contains", "key": "result.fingerprints[0]"},
            ], ensure_ascii=False),
            section="7.3.1", auth="no"),
    ])

    VSM_DISABLED = {"7.2.3", "7.2.8"}
    VSM_APPROVED = {"7.3.1", "7.3.2", "7.3.3"}
    for r in vsm_data:
        sec = r.get("section", "")
        if sec in VSM_DISABLED:
            r["enabled"] = "no"
        if sec in VSM_APPROVED:
            r["approved"] = "yes"

    ws_vsm = wb.create_sheet("vsm")
    write_sheet(ws_vsm, vsm_data)

    vsm_enabled = sum(1 for r in vsm_data if r.get("enabled", "yes") == "yes")
    print(f"vsm: {len(vsm_data)} 行 ({vsm_enabled} enabled, {len(vsm_data) - vsm_enabled} disabled)")

    wb.save(xlsx)
    print(f"\n已保存: {xlsx}")


if __name__ == "__main__":
    main()
