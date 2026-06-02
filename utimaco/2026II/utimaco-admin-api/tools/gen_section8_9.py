#!/usr/bin/env python3
"""
生成 section8_9 sheet — 覆盖第8章(FileServer)和第9章(租户密评)全部接口。

设计思路:
  - Section 8: FileServer，host = fileServerIp:port
    - 上传(multipart→enabled=no) / 获取描述 / 下载(binary→enabled=no) / 删除
    - 独立用例，无场景
  - Section 9: 租户密评，host = vsmIp:port (不同于管理接口)
    - 认证机制同CHSM（CHSM-AuthPK/SignatureAlg/Signature header）
    - 配置公钥指纹/清除公钥指纹 无需认证(guest)
    - 其他接口需要认证(trusted)
    - 用 SCN_TENANT 场景链:
      step 1: getStatus 获取基线状态
      step 2: authPK 配置第1个RSA公钥
      step 3: getAuthPKFingerprints 验证1个指纹
      step 4: authPK 配置第2个SM2公钥(混合算法)
      step 5: getAuthPKFingerprints 验证2组指纹
      step 6: cleanPK 清除全部
      step 7: getAuthPKFingerprints 验证已清空
      step 8: authPK 恢复公钥(后续接口需要认证)
      step 9-N: 各接口异常测试
    - 高危操作 initKey / doVsmInit → enabled=no
    - 文件操作 exportBackupKeys / importBackupKeys → enabled=no(binary/multipart)

注意:
  - host 列填写占位符 ${fileServerHost} 和 ${vsmHost}，
    运行时需在 config 或手动填入实际地址
  - Section 9 的 servlet URL 通过 params 列传 method 参数
"""

import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

HEADERS = [
    "case_id", "description", "host", "endpoint", "method",
    "params", "json_data", "expected_status", "assert_rules",
    "scenario_id", "step", "step_type", "save_vars",
    "enabled", "section", "ref_case_id", "auth", "file_path", "approved",
]

OK = json.dumps([
    {"type": "status_code", "expected_code": 200},
    {"type": "json_contains", "key": "message", "value": "success"},
], ensure_ascii=False)

ERR400 = json.dumps([
    {"type": "status_code", "expected_code": 400},
], ensure_ascii=False)


def ok_with(*extra_rules):
    base = [
        {"type": "status_code", "expected_code": 200},
        {"type": "json_contains", "key": "message", "value": "success"},
    ]
    base.extend(extra_rules)
    return json.dumps(base, ensure_ascii=False)


def ok_status_only():
    return json.dumps([
        {"type": "status_code", "expected_code": 200},
    ], ensure_ascii=False)


def ok_data(*extra_rules):
    """Section 9 返回 data 而非 result"""
    base = [
        {"type": "status_code", "expected_code": 200},
    ]
    base.extend(extra_rules)
    return json.dumps(base, ensure_ascii=False)


def jd(d):
    return json.dumps(d, ensure_ascii=False)


def row(case_id, desc, endpoint, expected_status, json_data=None, params=None,
        assert_rules=None, ref_case_id=None, enabled="yes", section="",
        scenario_id=None, step=None, step_type=None, save_vars=None,
        method="POST", host=None, auth="yes", file_path=None, approved=""):
    if assert_rules is None:
        assert_rules = OK if expected_status == 200 else ERR400
    r = {
        "case_id": case_id,
        "description": desc,
        "host": host,
        "endpoint": endpoint,
        "method": method,
        "params": params,
        "json_data": json_data,
        "expected_status": expected_status,
        "assert_rules": assert_rules,
        "scenario_id": scenario_id,
        "step": step,
        "step_type": step_type,
        "save_vars": save_vars,
        "enabled": enabled,
        "section": section,
        "ref_case_id": ref_case_id or case_id,
        "auth": auth,
        "approved": approved,
    }
    if file_path:
        r["file_path"] = file_path
    return r


# FileServer host 占位符（${env.xxx} 语法，由 excel_handler 从 config sheet 替换）
FS_HOST = "${env.fileServerHost}"
# 租户密评 VSM host 占位符
VSM_HOST = "${env.vsmHost}"


# ================================================================
# Section 8: FileServer
# ================================================================

def gen_fileserver():
    """FileServer 接口，host = fileServerIp:port"""
    rows = []

    # --- 8.1.1 上传镜像文件 POST /images (multipart) ---
    EP_UPLOAD = "/images"
    rows.extend([
        row("FS_UPLOAD_001", "上传镜像文件 正常上传", EP_UPLOAD, 200,
            json_data=jd({"requestId": "uuid", "info": {"token": "/chsm/image_v1.bin", "type": "hsm"}}),
            host=FS_HOST, section="8.1.1", enabled="no"),
        row("FS_UPLOAD_002", "上传镜像文件 type=vsm", EP_UPLOAD, 200,
            json_data=jd({"requestId": "uuid", "info": {"token": "/vsm/image_v1.bin", "type": "vsm"}}),
            host=FS_HOST, section="8.1.1", enabled="no"),
        row("FS_UPLOAD_003", "上传镜像文件 type无效值", EP_UPLOAD, 400,
            json_data=jd({"requestId": "uuid", "info": {"token": "/chsm/image.bin", "type": "invalid"}}),
            host=FS_HOST, section="8.1.1", enabled="no"),
        row("FS_UPLOAD_004", "上传镜像文件 缺token", EP_UPLOAD, 400,
            json_data=jd({"requestId": "uuid", "info": {"type": "hsm"}}),
            host=FS_HOST, section="8.1.1", enabled="no"),
        row("FS_UPLOAD_005", "上传镜像文件 缺requestId", EP_UPLOAD, 400,
            json_data=jd({"info": {"token": "/chsm/image.bin", "type": "hsm"}}),
            host=FS_HOST, section="8.1.1", enabled="no"),
    ])

    # --- 8.1.2 获取镜像文件描述信息 GET /info?requestId=xxx ---
    EP_INFO = "/info"
    rows.extend([
        row("FS_GETINFO_001", "获取镜像描述 正常获取", EP_INFO, 200,
            params="requestId=uuid",
            assert_rules=ok_with(
                {"type": "json_contains", "key": "result.info.token"},
                {"type": "json_contains", "key": "result.info.type"},
            ),
            host=FS_HOST, method="GET", section="8.1.2", enabled="no"),
        row("FS_GETINFO_002", "获取镜像描述 缺requestId", EP_INFO, 400,
            host=FS_HOST, method="GET", section="8.1.2", enabled="no"),
        row("FS_GETINFO_003", "获取镜像描述 requestId空字符串", EP_INFO, 400,
            params="requestId=",
            host=FS_HOST, method="GET", section="8.1.2", enabled="no"),
    ])

    # --- 8.1.3 下载文件 POST /images?file=xxx ---
    EP_DOWNLOAD = "/images"
    rows.extend([
        row("FS_DOWNLOAD_001", "下载文件 正常下载", EP_DOWNLOAD, 200,
            params="file=/chsm/image_v1.bin",
            assert_rules=ok_status_only(),
            host=FS_HOST, method="POST", section="8.1.3", enabled="no"),
        row("FS_DOWNLOAD_002", "下载文件 file不存在", EP_DOWNLOAD, 400,
            params="file=/nonexistent/file.bin",
            host=FS_HOST, method="POST", section="8.1.3", enabled="no"),
        row("FS_DOWNLOAD_003", "下载文件 缺file参数", EP_DOWNLOAD, 400,
            host=FS_HOST, method="POST", section="8.1.3", enabled="no"),
    ])

    # --- 8.1.4 删除镜像文件 POST /images/{requestId}/{token} (路径参数) ---
    rows.extend([
        row("FS_DELETE_001", "删除镜像文件 正常删除",
            "/images/uuid/chsm/image_v1.bin", 200,
            host=FS_HOST, section="8.1.4", enabled="no"),
        row("FS_DELETE_002", "删除镜像文件 token不存在",
            "/images/uuid/nonexistent/file.bin", 400,
            host=FS_HOST, section="8.1.4", enabled="no"),
        row("FS_DELETE_003", "删除镜像文件 缺requestId(路径为空)",
            "/images//chsm/image_v1.bin", 400,
            host=FS_HOST, section="8.1.4", enabled="no"),
    ])

    return rows


# ================================================================
# Section 9: 租户密评
# ================================================================

def gen_tenant_authpk_scenario():
    """SCN_TENANT_AUTHPK: 公钥配置生命周期（最多2个，支持混合算法）+ 指定删除 + 全部清除"""
    rows = []
    step = 0

    def add(case_id, desc, endpoint, expected_status, json_data=None, params=None,
            ref_case_id=None, enabled="yes", section="", assert_rules=None,
            save_vars=None, step_type="test", method="POST", auth="no"):
        nonlocal step
        step += 1
        rows.append(row(
            case_id, desc, endpoint, expected_status,
            json_data=json_data, params=params,
            assert_rules=assert_rules,
            ref_case_id=ref_case_id, enabled=enabled, section=section,
            scenario_id="SCN_TENANT_AUTHPK", step=step, step_type=step_type,
            save_vars=save_vars, method=method, host=VSM_HOST, auth=auth,
        ))

    # 断言辅助 — 验证 alg 分组 + 数量
    def assert_1_rsa():
        """1组sha256，1个指纹"""
        return ok_data(
            {"type": "json_contains", "key": "data[0].alg", "value": "sha256"},
            {"type": "json_contains", "key": "data[0].fingerprints[0]"},
            {"type": "json_not_contains", "key": "data[0].fingerprints[1]"},
            {"type": "json_not_contains", "key": "data[1]"},
        )

    def assert_rsa_and_sm2():
        """2组(sha256+sm3)，各1个指纹"""
        return ok_data(
            {"type": "json_contains", "key": "data[0]"},
            {"type": "json_contains", "key": "data[1]"},
            {"type": "json_not_contains", "key": "data[2]"},
        )

    def assert_1_sm2():
        """1组sm3，1个指纹，sha256组不存在"""
        return ok_data(
            {"type": "json_contains", "key": "data[0].alg", "value": "sm3"},
            {"type": "json_contains", "key": "data[0].fingerprints[0]"},
            {"type": "json_not_contains", "key": "data[0].fingerprints[1]"},
            {"type": "json_not_contains", "key": "data[1]"},
        )

    def assert_data_empty():
        return ok_data({"type": "json_not_contains", "key": "data[0]"})

    # step 1: 清空残留公钥（确保干净初始状态）
    add("TENANT_CLEAN_INIT", "cleanPK 清空残留公钥(确保初始状态干净)",
        "/authServlet", 200,
        params="method=cleanPK",
        json_data=jd({"requestId": "uuid"}),
        step_type="setup", section="9.2.3")

    # step 2: 配置第1个RSA公钥
    add("TENANT_AUTHPK_001", "authPK 配置第1个RSA公钥",
        "/authServlet", 200,
        params="method=authPK",
        json_data=jd({"requestId": "uuid", "alg": "RSAWithSHA256", "pks": ["${keys.rsa.key1.public_key_pem}"]}),
        step_type="setup", section="9.2.2")

    # step 2: 查询验证1个RSA指纹（1组sha256，1个指纹）
    add("TENANT_GETPK_001", "getAuthPKFingerprints 验证1组sha256+1个指纹",
        "/authServlet", 200,
        params="method=getAuthPKFingerprints&requestId=uuid",
        assert_rules=assert_1_rsa(),
        method="GET", step_type="verify", section="9.2.1",
        save_vars="rsa_fp=data[0].fingerprints[0]")

    # step 3: 配置第2个SM2公钥（混合算法，总数2不超限）
    add("TENANT_AUTHPK_002", "authPK 配置第2个SM2公钥(混合算法,总数2)",
        "/authServlet", 200,
        params="method=authPK",
        json_data=jd({"requestId": "uuid", "alg": "sm2", "pks": ["${keys.sm2.key1.public_key_pem}"]}),
        section="9.2.2")

    # step 4: 查询验证RSA+SM2两组指纹（2组，各1个）
    add("TENANT_GETPK_002", "getAuthPKFingerprints 验证2组(sha256+sm3)",
        "/authServlet", 200,
        params="method=getAuthPKFingerprints&requestId=uuid",
        assert_rules=assert_rsa_and_sm2(),
        method="GET", step_type="verify", section="9.2.1")

    # step 5: 重复配置相同RSA公钥（幂等）
    add("TENANT_AUTHPK_003", "authPK 重复配置相同RSA公钥(幂等)",
        "/authServlet", 200,
        params="method=authPK",
        json_data=jd({"requestId": "uuid", "alg": "RSAWithSHA256", "pks": ["${keys.rsa.key1.public_key_pem}"]}),
        section="9.2.2")

    # step 6: 幂等后查询仍为2组
    add("TENANT_GETPK_003", "getAuthPKFingerprints 幂等后仍为2组",
        "/authServlet", 200,
        params="method=getAuthPKFingerprints&requestId=uuid",
        assert_rules=assert_rsa_and_sm2(),
        method="GET", step_type="verify", section="9.2.1")

    # step 7: 指定删除RSA公钥
    add("TENANT_CLEANPK_SPEC", "cleanPK 指定fingerprint删除RSA公钥",
        "/authServlet", 200,
        params="method=cleanPK",
        json_data=jd({"requestId": "uuid", "fingerprint": "${rsa_fp}", "action": "remove"}),
        section="9.2.3")

    # step 8: 验证仅剩SM2（1组sm3，sha256不存在）
    add("TENANT_GETPK_004", "getAuthPKFingerprints 验证仅剩1组sm3",
        "/authServlet", 200,
        params="method=getAuthPKFingerprints&requestId=uuid",
        assert_rules=assert_1_sm2(),
        method="GET", step_type="verify", section="9.2.1")

    # step 9: 全部清除
    add("TENANT_CLEANPK_ALL", "cleanPK 不传fingerprint清除全部",
        "/authServlet", 200,
        params="method=cleanPK",
        json_data=jd({"requestId": "uuid"}),
        section="9.2.3")

    # step 10: 验证全部清空
    add("TENANT_GETPK_005", "getAuthPKFingerprints 验证全部清空",
        "/authServlet", 200,
        params="method=getAuthPKFingerprints&requestId=uuid",
        assert_rules=assert_data_empty(),
        method="GET", step_type="verify", section="9.2.1")

    # step 11: 恢复RSA公钥
    add("TENANT_AUTHPK_004", "authPK 恢复RSA公钥(为后续接口准备)",
        "/authServlet", 200,
        params="method=authPK",
        json_data=jd({"requestId": "uuid", "alg": "RSAWithSHA256", "pks": ["${keys.rsa.key1.public_key_pem}"]}),
        step_type="teardown", section="9.2.2")

    # step 12: 验证恢复成功（1组sha256，1个指纹）
    add("TENANT_GETPK_006", "getAuthPKFingerprints 验证恢复1组sha256",
        "/authServlet", 200,
        params="method=getAuthPKFingerprints&requestId=uuid",
        assert_rules=assert_1_rsa(),
        method="GET", step_type="verify", section="9.2.1")

    return rows


def gen_tenant_limit_scenario():
    """SCN_TENANT_LIMIT: 超限测试 (已有2个再加1个→400)"""
    return [
        row("TENANT_LIMIT_CLEAN", "cleanPK 清空残留公钥(确保初始状态)",
            "/authServlet", 200,
            params="method=cleanPK",
            json_data=jd({"requestId": "uuid"}),
            scenario_id="SCN_TENANT_LIMIT", step=1, step_type="setup",
            section="9.2.3", host=VSM_HOST, auth="no"),
        row("TENANT_LIMIT_S01", "authPK 配置第1个RSA公钥(前置)",
            "/authServlet", 200,
            params="method=authPK",
            json_data=jd({"requestId": "uuid", "alg": "RSAWithSHA256", "pks": ["${keys.rsa.key1.public_key_pem}"]}),
            scenario_id="SCN_TENANT_LIMIT", step=2, step_type="setup",
            section="9.2.2", host=VSM_HOST, auth="no"),
        row("TENANT_LIMIT_S02", "authPK 配置第2个SM2公钥(前置,总数2)",
            "/authServlet", 200,
            params="method=authPK",
            json_data=jd({"requestId": "uuid", "alg": "sm2", "pks": ["${keys.sm2.key1.public_key_pem}"]}),
            scenario_id="SCN_TENANT_LIMIT", step=3, step_type="setup",
            section="9.2.2", host=VSM_HOST, auth="no"),
        row("TENANT_LIMIT_T01", "authPK 超限-已有2个再传1个第3个",
            "/authServlet", 409,
            params="method=authPK",
            json_data=jd({"requestId": "uuid", "alg": "RSAWithSHA256", "pks": ["${keys.rsa.key2.public_key_pem}"]}),
            assert_rules=json.dumps([{"type": "status_code", "expected_code": 409}], ensure_ascii=False),
            scenario_id="SCN_TENANT_LIMIT", step=4, step_type="test",
            section="9.2.2", host=VSM_HOST, auth="no"),
        row("TENANT_LIMIT_T02", "authPK 超限-一次传3个公钥",
            "/authServlet", 409,
            params="method=authPK",
            json_data=jd({"requestId": "uuid", "alg": "RSAWithSHA256", "pks": ["${keys.rsa.key1.public_key_pem}", "${keys.rsa.key2.public_key_pem}", "${keys.rsa.key3.public_key_pem}"]}),
            assert_rules=json.dumps([{"type": "status_code", "expected_code": 409}], ensure_ascii=False),
            scenario_id="SCN_TENANT_LIMIT", step=5, step_type="test",
            section="9.2.2", host=VSM_HOST, auth="no"),
        row("TENANT_LIMIT_TEARDOWN", "cleanPK 超限测试后清理",
            "/authServlet", 200,
            params="method=cleanPK",
            json_data=jd({"requestId": "uuid"}),
            scenario_id="SCN_TENANT_LIMIT", step=6, step_type="teardown",
            section="9.2.3", host=VSM_HOST, auth="no"),
        row("TENANT_LIMIT_RESTORE", "authPK 恢复RSA公钥(为后续场景准备)",
            "/authServlet", 200,
            params="method=authPK",
            json_data=jd({"requestId": "uuid", "alg": "RSAWithSHA256", "pks": ["${keys.rsa.key1.public_key_pem}"]}),
            scenario_id="SCN_TENANT_LIMIT", step=7, step_type="teardown",
            section="9.2.2", host=VSM_HOST, auth="no"),
    ]


def gen_tenant_standalone():
    """租户密评独立用例 (纯参数校验，无状态依赖)"""
    rows = []

    # --- 9.2.2 authPK 参数异常 ---
    rows.extend([
        row("TENANT_AUTHPK_E01", "authPK alg无效值",
            "/authServlet", 400,
            params="method=authPK",
            json_data=jd({"requestId": "uuid", "alg": "invalid_alg", "pks": ["${keys.rsa.key1.public_key_pem}"]}),
            section="9.2.2", host=VSM_HOST, auth="no"),
        row("TENANT_AUTHPK_E02", "authPK pks为空列表",
            "/authServlet", 400,
            params="method=authPK",
            json_data=jd({"requestId": "uuid", "alg": "RSAWithSHA256", "pks": []}),
            section="9.2.2", host=VSM_HOST, auth="no"),
        row("TENANT_AUTHPK_E03", "authPK 缺alg字段",
            "/authServlet", 400,
            params="method=authPK",
            json_data=jd({"requestId": "uuid", "pks": ["${keys.rsa.key1.public_key_pem}"]}),
            section="9.2.2", host=VSM_HOST, auth="no"),
        row("TENANT_AUTHPK_E04", "authPK 缺pks字段",
            "/authServlet", 400,
            params="method=authPK",
            json_data=jd({"requestId": "uuid", "alg": "RSAWithSHA256"}),
            section="9.2.2", host=VSM_HOST, auth="no"),
        row("TENANT_AUTHPK_E05", "authPK 缺requestId",
            "/authServlet", 400,
            params="method=authPK",
            json_data=jd({"alg": "RSAWithSHA256", "pks": ["${keys.rsa.key1.public_key_pem}"]}),
            section="9.2.2", host=VSM_HOST, auth="no"),
        row("TENANT_AUTHPK_E06", "authPK requestId空字符串",
            "/authServlet", 400,
            params="method=authPK",
            json_data=jd({"requestId": "", "alg": "RSAWithSHA256", "pks": ["${keys.rsa.key1.public_key_pem}"]}),
            section="9.2.2", host=VSM_HOST, auth="no"),
        row("TENANT_AUTHPK_E07", "authPK pks含非法格式公钥",
            "/authServlet", 400,
            params="method=authPK",
            json_data=jd({"requestId": "uuid", "alg": "RSAWithSHA256", "pks": ["NOT_A_VALID_PUBLIC_KEY!!!"]}),
            section="9.2.2", host=VSM_HOST, auth="no"),
    ])

    # --- 9.2.1 getAuthPKFingerprints 参数异常 ---
    rows.extend([
        row("TENANT_GETPK_E01", "getAuthPKFingerprints 缺requestId",
            "/authServlet", 400,
            params="method=getAuthPKFingerprints",
            method="GET", section="9.2.1", host=VSM_HOST, auth="no"),
        row("TENANT_GETPK_E02", "getAuthPKFingerprints requestId空字符串",
            "/authServlet", 400,
            params="method=getAuthPKFingerprints&requestId=",
            method="GET", section="9.2.1", host=VSM_HOST, auth="no"),
    ])

    # --- 9.2.3 cleanPK 参数异常 ---
    rows.extend([
        row("TENANT_CLEANPK_E01", "cleanPK 缺requestId",
            "/authServlet", 400,
            params="method=cleanPK",
            json_data=jd({}),
            section="9.2.3", host=VSM_HOST, auth="no"),
        row("TENANT_CLEANPK_E02", "cleanPK requestId空字符串",
            "/authServlet", 400,
            params="method=cleanPK",
            json_data=jd({"requestId": ""}),
            section="9.2.3", host=VSM_HOST, auth="no"),
        row("TENANT_CLEANPK_E03", "cleanPK 指定不存在的fingerprint",
            "/authServlet", 400,
            params="method=cleanPK",
            json_data=jd({"requestId": "uuid", "fingerprint": "nonexistent_fp_value"}),
            section="9.2.3", host=VSM_HOST, auth="no"),
        row("TENANT_CLEANPK_E04", "cleanPK fingerprint空字符串(等同全部清空)",
            "/authServlet", 200,
            params="method=cleanPK",
            json_data=jd({"requestId": "uuid", "fingerprint": ""}),
            section="9.2.3", host=VSM_HOST, auth="no"),
        row("TENANT_CLEANPK_E05", "cleanPK action无效值",
            "/authServlet", 400,
            params="method=cleanPK",
            json_data=jd({"requestId": "uuid", "action": "invalid_action"}),
            section="9.2.3", host=VSM_HOST, auth="no"),
    ])

    # --- 恢复租户公钥（cleanPK_E04 fingerprint空字符串会清空，需恢复后续认证）---
    rows.extend([
        row("TENANT_RESTORE_PK", "恢复RSA租户公钥(为9.2.4-9.2.8认证准备)",
            "/authServlet", 200,
            params="method=authPK",
            json_data=jd({"requestId": "uuid", "alg": "RSAWithSHA256", "pks": ["${keys.rsa.key1.public_key_pem}"]}),
            section="9.2.2", host=VSM_HOST, auth="no"),
    ])

    # getStatus (POST)
    rows.extend([
        row("TENANT_STATUS_001", "getStatus 正常获取VSM运行状态",
            "/platformServlet", 200,
            params="method=getStatus",
            json_data=jd({"requestId": "uuid"}),
            assert_rules=ok_data(
                {"type": "json_contains", "key": "data.status", "value": "success"},
                {"type": "json_contains", "key": "data.initFlag"},
                {"type": "json_contains", "key": "data.serviceStatus"},
                {"type": "json_contains", "key": "data.keyDigest"},
                {"type": "json_contains", "key": "data.certDigest"},
                {"type": "json_contains", "key": "data.configDigest"},
            ),
            section="9.2.8", host=VSM_HOST),
        row("TENANT_STATUS_T01", "getStatus 缺requestId",
            "/platformServlet", 400,
            params="method=getStatus",
            json_data=jd({}),
            section="9.2.8", host=VSM_HOST),
        row("TENANT_STATUS_T02", "getStatus requestId空字符串",
            "/platformServlet", 400,
            params="method=getStatus",
            json_data=jd({"requestId": ""}),
            section="9.2.8", host=VSM_HOST),
    ])

    # initKey (高危)
    rows.extend([
        row("TENANT_INITKEY_001", "initKey 正常初始化(高危操作)",
            "/platformServlet", 200,
            params="method=initKey&requestId=uuid",
            method="POST", section="9.2.4", enabled="no", host=VSM_HOST),
        row("TENANT_INITKEY_002", "initKey 缺requestId",
            "/platformServlet", 400,
            params="method=initKey",
            method="POST", section="9.2.4", host=VSM_HOST),
        row("TENANT_INITKEY_003", "initKey requestId空字符串",
            "/platformServlet", 400,
            params="method=initKey&requestId=",
            method="POST", section="9.2.4", host=VSM_HOST),
    ])

    # exportBackupKeys
    rows.extend([
        row("TENANT_EXPORT_001", "exportBackupKeys 正常导出(返回binary)",
            "/platformServlet", 200,
            params="method=exportBackupKeys&requestId=uuid",
            json_data=jd({"requestId": "uuid", "backupKey": "test_backup_key_123"}),
            assert_rules=ok_status_only(),
            method="POST", section="9.2.5", host=VSM_HOST),
        row("TENANT_EXPORT_002", "exportBackupKeys 缺backupKey",
            "/platformServlet", 400,
            params="method=exportBackupKeys&requestId=uuid",
            json_data=jd({"requestId": "uuid"}),
            method="POST", section="9.2.5", host=VSM_HOST),
        row("TENANT_EXPORT_003", "exportBackupKeys backupKey空字符串",
            "/platformServlet", 400,
            params="method=exportBackupKeys&requestId=uuid",
            json_data=jd({"requestId": "uuid", "backupKey": ""}),
            method="POST", section="9.2.5", host=VSM_HOST),
        row("TENANT_EXPORT_004", "exportBackupKeys 缺requestId",
            "/platformServlet", 400,
            params="method=exportBackupKeys",
            json_data=jd({"backupKey": "test_backup_key_123"}),
            method="POST", section="9.2.5", host=VSM_HOST),
    ])

    # importBackupKeys (multipart文件上传)
    rows.extend([
        row("TENANT_IMPORT_001", "importBackupKeys 正常导入",
            "/platformServlet", 200,
            params="method=importBackupKeys",
            json_data=jd({"requestId": "uuid", "backupKey": "test_backup_key_123"}),
            method="POST", section="9.2.6", host=VSM_HOST,
            file_path="data/test_backup.dat"),
        row("TENANT_IMPORT_002", "importBackupKeys 缺backupKey",
            "/platformServlet", 400,
            params="method=importBackupKeys",
            json_data=jd({"requestId": "uuid"}),
            method="POST", section="9.2.6", host=VSM_HOST,
            file_path="data/test_backup.dat"),
        row("TENANT_IMPORT_003", "importBackupKeys 缺requestId",
            "/platformServlet", 400,
            params="method=importBackupKeys",
            json_data=jd({"backupKey": "test_backup_key_123"}),
            method="POST", section="9.2.6", host=VSM_HOST,
            file_path="data/test_backup.dat"),
        row("TENANT_IMPORT_004", "importBackupKeys 缺文件",
            "/platformServlet", 400,
            params="method=importBackupKeys",
            json_data=jd({"requestId": "uuid", "backupKey": "test_backup_key_123"}),
            method="POST", section="9.2.6", host=VSM_HOST),
    ])

    # doVsmInit (高危，场景验证 clearPK 效果)
    # SCN_VSMINIT: 配公钥→clearPK=false→查验证公钥还在→clearPK=true→查验证公钥已清→恢复
    rows.extend([
        row("TENANT_VSMINIT_S01", "配置RSA公钥(doVsmInit前置)",
            "/authServlet", 200,
            params="method=authPK",
            json_data=jd({"requestId": "uuid", "alg": "RSAWithSHA256", "pks": ["${keys.rsa.key1.public_key_pem}"]}),
            scenario_id="SCN_VSMINIT", step=1, step_type="setup",
            section="9.2.2", host=VSM_HOST, auth="no"),
        row("TENANT_VSMINIT_001", "doVsmInit clearPK=false(不清公钥)",
            "/platformServlet", 200,
            params="method=doVsmInit",
            json_data=jd({"requestId": "uuid", "clearPK": "false"}),
            scenario_id="SCN_VSMINIT", step=2, step_type="test",
            section="9.2.7", host=VSM_HOST),
        row("TENANT_VSMINIT_V01", "查询验证公钥仍存在(clearPK=false)",
            "/authServlet", 200,
            params="method=getAuthPKFingerprints&requestId=uuid",
            assert_rules=ok_data({"type": "json_contains", "key": "data[0]"}),
            method="GET",
            scenario_id="SCN_VSMINIT", step=3, step_type="verify",
            section="9.2.1", host=VSM_HOST, auth="no"),
        row("TENANT_VSMINIT_002", "doVsmInit clearPK=true(清除公钥)",
            "/platformServlet", 200,
            params="method=doVsmInit",
            json_data=jd({"requestId": "uuid", "clearPK": "true"}),
            scenario_id="SCN_VSMINIT", step=4, step_type="test",
            section="9.2.7", host=VSM_HOST),
        row("TENANT_VSMINIT_V03", "查询验证公钥已清空(clearPK=true)",
            "/authServlet", 200,
            params="method=getAuthPKFingerprints&requestId=uuid",
            assert_rules=ok_data({"type": "json_not_contains", "key": "data[0]"}),
            method="GET",
            scenario_id="SCN_VSMINIT", step=5, step_type="verify",
            section="9.2.1", host=VSM_HOST, auth="no"),
        row("TENANT_VSMINIT_S02", "恢复RSA公钥(doVsmInit后恢复)",
            "/authServlet", 200,
            params="method=authPK",
            json_data=jd({"requestId": "uuid", "alg": "RSAWithSHA256", "pks": ["${keys.rsa.key1.public_key_pem}"]}),
            scenario_id="SCN_VSMINIT", step=6, step_type="teardown",
            section="9.2.2", host=VSM_HOST, auth="no"),
    ])
    # doVsmInit 独立错误用例
    rows.extend([
        row("TENANT_VSMINIT_003", "doVsmInit 缺requestId",
            "/platformServlet", 400,
            params="method=doVsmInit",
            json_data=jd({"clearPK": "false"}),
            method="POST", section="9.2.7", host=VSM_HOST),
        row("TENANT_VSMINIT_004", "doVsmInit 缺clearPK",
            "/platformServlet", 400,
            params="method=doVsmInit",
            json_data=jd({"requestId": "uuid"}),
            method="POST", section="9.2.7", host=VSM_HOST),
        row("TENANT_VSMINIT_005", "doVsmInit clearPK无效值",
            "/platformServlet", 400,
            params="method=doVsmInit",
            json_data=jd({"requestId": "uuid", "clearPK": "maybe"}),
            method="POST", section="9.2.7", host=VSM_HOST),
    ])

    return rows


# ================================================================
# 主流程
# ================================================================

def write_sheet(ws, data_rows):
    for ci, h in enumerate(HEADERS, 1):
        ws.cell(1, ci, h)
    for ri, data in enumerate(data_rows, 2):
        for ci, h in enumerate(HEADERS, 1):
            val = data.get(h)
            if val is not None:
                ws.cell(ri, ci, val)


def main():
    xlsx = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "test_data.xlsx")
    wb = openpyxl.load_workbook(xlsx)

    # 删除旧 sheet
    for name in ["section8_9", "section8", "section9", "section9_pk"]:
        if name in wb.sheetnames:
            del wb[name]

    # ================================================================
    # Section 8: FileServer
    # ================================================================
    s8_data = []
    # --- 前置: 配置公钥 ---
    s8_data.extend([
        row("INIT_RSA_POST_001", "首次配置1个RSA公钥",
            "/api/1.0/chsm/authpk", 200,
            json_data=jd({"requestId": "uuid", "algorithm": "RSAWithSHA256", "pks": ["${keys.rsa.key1.public_key_pem}"]}),
            ref_case_id="AUTH_PK_001", section="7.3.2", auth="yes"),
        row("INIT_RSA_GET_002", "查询验证1个指纹",
            "/api/1.0/chsm/authpk", 200,
            params="requestId=uuid",
            assert_rules=json.dumps([
                {"type": "status_code", "expected_code": 200},
                {"type": "json_contains", "key": "result.algorithm", "value": "sha256"},
                {"type": "json_contains", "key": "result.fingerprints[0]"},
                {"type": "json_not_contains", "key": "result.fingerprints[1]"},
            ], ensure_ascii=False),
            ref_case_id="AUTH_PK_002", section="7.3.1", method="GET", auth="no"),
    ])
    s8_data.extend(gen_fileserver())
    # --- 尾部: 清空公钥 ---
    s8_data.extend([
        row("PK_CLEAR_H001", "清空CHSM所有公钥", "/api/1.0/chsm/authpk", 200,
            json_data=jd({"requestId": "uuid"}), method="DELETE",
            section="7.3.3", auth="yes"),
        row("PK_GET_H002", "清空后查询CHSM验证为空", "/api/1.0/chsm/authpk", 200,
            params="requestId=uuid", method="GET",
            assert_rules=json.dumps([
                {"type": "status_code", "expected_code": 200},
                {"type": "json_not_contains", "key": "result.fingerprints[0]"},
            ], ensure_ascii=False),
            section="7.3.1", auth="no"),
    ])
    ws8 = wb.create_sheet("section8")
    write_sheet(ws8, s8_data)
    print(f"section8: {len(s8_data)} 行 (FileServer全部enabled=no)")

    # ================================================================
    # section9_pk: 9.2.1-3 公钥接口（不需要认证）
    # ================================================================
    s9pk_data = []
    authpk_data = gen_tenant_authpk_scenario()
    limit_data = gen_tenant_limit_scenario()
    all_standalone = gen_tenant_standalone()
    standalone_pk = [r for r in all_standalone if r.get("section", "") in ("9.2.1", "9.2.2", "9.2.3")
                     and r.get("scenario_id", "") != "SCN_VSMINIT"]
    s9pk_data.extend(authpk_data)
    s9pk_data.extend(limit_data)
    s9pk_data.extend(standalone_pk)
    for r in s9pk_data:
        sec = r.get("section", "")
        if sec in ("9.2.1", "9.2.2", "9.2.3"):
            r["approved"] = "yes"

    ws9pk = wb.create_sheet("section9_pk")
    write_sheet(ws9pk, s9pk_data)
    s9pk_en = sum(1 for r in s9pk_data if r.get("enabled", "yes") == "yes")
    print(f"section9_pk: {len(s9pk_data)} 行 ({s9pk_en} enabled) — 9.2.1-3")
    print(f"    SCN_TENANT_AUTHPK: {len(authpk_data)} 步")
    print(f"    SCN_TENANT_LIMIT:  {len(limit_data)} 步")
    print(f"    独立用例:          {len(standalone_pk)} 条")

    # ================================================================
    # section9: 9.2.4-8 其他接口（需要认证）
    # ================================================================
    s9_data = []
    s9_data.extend([
        row("INIT_TENANT_PK_001", "配置VSM租户RSA公钥(认证用)",
            "/authServlet", 200,
            params="method=authPK",
            json_data=jd({"requestId": "uuid", "alg": "RSAWithSHA256", "pks": ["${keys.rsa.key1.public_key_pem}"]}),
            section="9.2.2", host=VSM_HOST, auth="no"),
        row("INIT_TENANT_PK_002", "验证VSM租户公钥已配置",
            "/authServlet", 200,
            params="method=getAuthPKFingerprints&requestId=uuid",
            assert_rules=ok_data({"type": "json_contains", "key": "data[0]"}),
            method="GET", section="9.2.1", host=VSM_HOST, auth="no"),
    ])
    standalone_other = [r for r in all_standalone
                        if r.get("section", "") not in ("9.2.1", "9.2.2", "9.2.3")
                        and r.get("scenario_id", "") != "SCN_VSMINIT"]
    # SCN_VSMINIT 场景完整放这里（含 authServlet + platformServlet 步骤）
    vsminit_rows = [r for r in all_standalone if r.get("scenario_id", "") == "SCN_VSMINIT"]
    s9_data.extend(vsminit_rows)
    s9_data.extend(standalone_other)
    s9_data.extend([
        row("CLEAN_TENANT_PK", "清空VSM租户公钥",
            "/authServlet", 200,
            params="method=cleanPK",
            json_data=jd({"requestId": "uuid"}),
            section="9.2.3", host=VSM_HOST, auth="no"),
    ])
    S9_ENABLED = {"9.2.1", "9.2.2", "9.2.3", "9.2.5", "9.2.6", "9.2.7", "9.2.8"}
    S9_APPROVED = {"9.2.7", "9.2.8"}
    for r in s9_data:
        sec = r.get("section", "")
        sid = r.get("scenario_id", "")
        if sec in S9_ENABLED or sid == "SCN_VSMINIT":
            pass
        else:
            r["enabled"] = "no"
        if sec in S9_APPROVED or sid == "SCN_VSMINIT":
            r["approved"] = "yes"

    ws9 = wb.create_sheet("section9")
    write_sheet(ws9, s9_data)
    s9_en = sum(1 for r in s9_data if r.get("enabled", "yes") == "yes")
    print(f"section9: {len(s9_data)} 行 ({s9_en} enabled) — 9.2.4-8")

    wb.save(xlsx)
    print(f"\n已保存: {xlsx}")


if __name__ == "__main__":
    main()
