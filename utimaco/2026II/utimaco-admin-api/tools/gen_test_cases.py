"""
补齐测试用例 Excel — 根据 test_data.xlsx 中的测试数据行，
生成缺失的测试用例到 华为云密码机二期-接口测试用例.xlsx

输出按需求章节编号排序 (7.1.1 → 7.1.2 → ... → 7.3 → 8 → 9)，
使用 openpyxl 写入带格式的 Excel。
"""

import json
import os
import re
from typing import Any, Dict, List, Optional

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd

# ── 文件路径 ──────────────────────────────────────────
_PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TC_FILE = os.path.join(_PROJECT_DIR, "data", "华为云密码机二期-接口测试用例.xlsx")
TD_FILE = os.path.join(_PROJECT_DIR, "data", "test_data.xlsx")
TC_SHEET = "管理接口"

# ── 测试用例列 ────────────────────────────────────────
TC_COLS = [
    "用例ID", "需求章节", "测试模块", "接口名称", "接口路径", "请求方法",
    "测试场景", "测试类型", "优先级", "前置条件",
    "请求头(Headers)", "请求参数(Query/Body)/步骤",
    "预期响应状态码", "预期响应", "预期结果描述",
    "测试状态", "执行人", "执行时间", "缺陷ID", "备注",
]

COL_WIDTHS = {
    "用例ID": 22, "需求章节": 10, "测试模块": 16, "接口名称": 38,
    "接口路径": 28, "请求方法": 9,
    "测试场景": 42, "测试类型": 10, "优先级": 7, "前置条件": 32,
    "请求头(Headers)": 32, "请求参数(Query/Body)/步骤": 48,
    "预期响应状态码": 14, "预期响应": 42, "预期结果描述": 26,
    "测试状态": 9, "执行人": 9, "执行时间": 12, "缺陷ID": 10, "备注": 38,
}

# ── 认证头模板 ────────────────────────────────────────
AUTH_HEADER_TRUSTED = """{
  "CHSM-AuthPK": "<公钥指纹>",
  "CHSM-SignatureAlg": "SHA256withRSA",
  "CHSM-Signature": "<签名值>"
}"""
AUTH_HEADER_GUEST = "无需认证头（guest接口）"
AUTH_HEADER_NONE = "无认证头"
AUTH_HEADER_TENANT = """{
  "CHSM-AuthPK": "<公钥指纹>",
  "CHSM-SignatureAlg": "SHA256withRSA",
  "CHSM-Signature": "<签名值>"
}"""

# ── 模块映射 ──────────────────────────────────────────
MODULE_MAP = {
    "CHSM": "CHSM配置管理",
    "VSM": "VSM配置管理",
    "GUEST": "Guest接口(无认证)",
    "PK_RSA": "授权配置(RSA)",
    "PK_SM2": "授权配置(SM2)",
    "FS": "FileServer文件服务",
    "TENANT": "租户密评",
}

# ── 接口名称映射 ──────────────────────────────────────
INTERFACE_NAME_MAP = {
    "getCHSMInfo": "getCHSMInfo-获取CHSM详细信息",
    "getCHSMStatus": "getCHSMStatus-获取CHSM运行状态",
    "getCHSMAllStatus": "getCHSMAllStatus-获取CHSM所有状态",
    "configCHSMNet": "configCHSMNet-配置CHSM网络",
    "configCHSMNtp": "configCHSMNtp-配置NTP服务器",
    "configCHSMUploadAddress": "configCHSMUploadAddress-配置影像上传地址",
    "configSyslogAddr": "configSyslogAddr-配置日志上传地址",
    "configCHSMAlarmAddress": "configCHSMAlarmAddress-配置告警上传地址",
    "configCHSMToken": "configCHSMToken-配置云平台Token",
    "configCHSMMOOCAddress": "configCHSMMOOCAddress-配置MO OC对接信息",
    "exportCHSM": "exportCHSM-导出CHSM影像",
    "importCHSM": "importCHSM-导入CHSM影像",
    "upgradeCHSM": "upgradeCHSM-升级CHSM",
    "restartCHSM": "restartCHSM-重启CHSM",
    "backupCHSM": "backupCHSM-备份CHSM",
    "restoreCHSM": "restoreCHSM-恢复CHSM",
    "getCHSMDebugInfo": "getCHSMDebugInfo-获取调试信息",
    "getCHSMDeviceInfo": "getCHSMDeviceInfo-获取设备信息",
    "getVSMInfo": "getVSMInfo-获取VSM详细信息",
    "getVSMStatus": "getVSMStatus-获取VSM运行状态",
    "configVSMNet": "configVSMNet-配置VSM网络",
    "configVSMToken": "configVSMToken-配置VSM Token",
    "exportVSM": "exportVSM-导出VSM影像",
    "importVSM": "importVSM-导入VSM影像",
    "startVSM": "startVSM-启动VSM",
    "stopVSM": "stopVSM-停止VSM",
    "restartVSM": "restartVSM-重启VSM",
    "resetVSM": "resetVSM-重置VSM",
    "upgradeVSM": "upgradeVSM-升级VSM",
    "configCHSMPk": "configCHSMPk-配置认证公钥",
    "getCHSMPk": "getCHSMPk-获取认证公钥指纹",
    "clearCHSMPk": "clearCHSMPk-清空认证公钥",
    "authPK": "authPK-配置公钥指纹",
    "getAuthPKFingerprints": "getAuthPKFingerprints-获取公钥指纹",
    "cleanPK": "cleanPK-清除公钥指纹",
    "initKey": "initKey-初始化本地主密钥",
    "exportBackupKeys": "exportBackupKeys-导出数据影像",
    "importBackupKeys": "importBackupKeys-导入数据影像",
    "doVsmInit": "doVsmInit-初始化VSM",
    "getStatus": "getStatus-获取VSM运行状态",
    "上传镜像文件": "uploadImage-上传镜像文件",
    "获取镜像描述": "getImageInfo-获取镜像文件描述信息",
    "下载文件": "downloadFile-下载文件",
    "删除镜像文件": "deleteImage-删除镜像文件",
}

# ── 场景 setup/teardown 接口名覆盖 ───────────────────
SETUP_CASE_MAP = {
    "VSM_SETUP_INFO": "getCHSMInfo-获取CHSM详细信息",
    "VSM_SETUP_NET": "getCHSMInfo-获取CHSM详细信息",
    "VSM_SETUP_TOKEN": "getCHSMInfo-获取CHSM详细信息",
    "VSM_SETUP_EXPORT": "getCHSMInfo-获取CHSM详细信息",
    "VSM_SETUP_IMPORT": "getCHSMInfo-获取CHSM详细信息",
    "VSM_SETUP_START": "getCHSMInfo-获取CHSM详细信息",
    "VSM_SETUP_STOP": "getCHSMInfo-获取CHSM详细信息",
    "VSM_SETUP_RESTART": "getCHSMInfo-获取CHSM详细信息",
    "VSM_SETUP_RESET": "getCHSMInfo-获取CHSM详细信息",
    "VSM_SETUP_UPGRADE": "getCHSMInfo-获取CHSM详细信息",
    "PK_RSA_S01_CLEAR": "clearCHSMPk-清空认证公钥",
    "PK_RSA_S02_SETUP": "configCHSMPk-配置认证公钥",
    "PK_SM2_S01_CLEAR": "clearCHSMPk-清空认证公钥",
    "PK_SM2_S02_SETUP": "configCHSMPk-配置认证公钥",
    "TENANT_LIMIT_SETUP": "authPK-配置公钥指纹",
    "TENANT_LIMIT_TEARDOWN": "cleanPK-清除公钥指纹",
    "PK_RSA_001": "configCHSMPk-配置认证公钥",
    "PK_RSA_012": "configCHSMPk-配置认证公钥",
    "PK_RSA_021": "configCHSMPk-配置认证公钥",
    "PK_SM2_001": "configCHSMPk-配置认证公钥",
    "PK_SM2_012": "configCHSMPk-配置认证公钥",
    "PK_SM2_021": "configCHSMPk-配置认证公钥",
    "TENANT_STATUS_001": "getStatus-获取VSM运行状态",
    "TENANT_CLEANUP_001": "cleanPK-清除公钥指纹",
    "TENANT_CLEANUP_002": "authPK-配置公钥指纹",
}

# ── case_id 前缀 → 章节号 (兜底映射) ─────────────────
SECTION_FALLBACK = {
    "AUTH_PK": "7.3", "CHSM_INFO": "7.1.1", "CHSM_STATUS": "7.1.2",
    "CHSM_ALLSTATUS": "7.1.3", "CHSM_NET": "7.1.4", "CHSM_NTP": "7.1.5",
    "CHSM_UPLOAD": "7.1.6", "CHSM_SYSLOG": "7.1.7", "CHSM_EXPORT": "7.1.8",
    "CHSM_IMPORT": "7.1.9", "CHSM_UPGRADE": "7.1.10", "CHSM_RESTART": "7.1.11",
    "CHSM_BACKUP": "7.1.12", "CHSM_RESTORE": "7.1.13", "CHSM_ALARM": "7.1.14",
    "CHSM_TOKEN": "7.1.15", "CHSM_MOOC": "7.1.16", "CHSM_DEBUG": "7.1.17",
    "CHSM_DEVICE": "7.1.18",
    "VSM_SETUP": "7.2.1", "VSM_INFO": "7.2.1", "VSM_STATUS": "7.2.2",
    "VSM_NET": "7.2.3", "VSM_TOKEN": "7.2.4", "VSM_EXPORT": "7.2.5",
    "VSM_IMPORT": "7.2.6", "VSM_START": "7.2.7", "VSM_STOP": "7.2.8",
    "VSM_RESTART": "7.2.9", "VSM_RESET": "7.2.10", "VSM_UPGRADE": "7.2.11",
    "PK_RSA": "7.3", "PK_SM2": "7.3",
    "GUEST_STATUS": "7.1.2", "GUEST_ALLSTATUS": "7.1.3",
    "GUEST_VSM_STATUS": "7.2.2", "GUEST_AUTH": "7.3",
    "FS_UPLOAD": "8.1.1", "FS_GETINFO": "8.1.2",
    "FS_DOWNLOAD": "8.1.3", "FS_DELETE": "8.1.4",
    "TENANT_STATUS": "9.2.8", "TENANT_AUTHPK": "9.2.2",
    "TENANT_GETPK": "9.2.1", "TENANT_CLEANPK": "9.2.3",
    "TENANT_CLEANUP": "9.2.3", "TENANT_LIMIT": "9.2.2",
    "TENANT_INITKEY": "9.2.4", "TENANT_EXPORT": "9.2.5",
    "TENANT_IMPORT": "9.2.6", "TENANT_VSMINIT": "9.2.7",
}


def _parse_section(sec) -> tuple:
    """'7.1.3' → (7, 1, 3) for numeric sorting"""
    if not sec or not isinstance(sec, str):
        return (99,)
    parts = []
    for p in sec.split("."):
        try:
            parts.append(int(p))
        except ValueError:
            parts.append(999)
    return tuple(parts)


def _valid_str(val) -> bool:
    return isinstance(val, str) and val.strip() != ""


def _build_section_map() -> dict:
    """从 test_data 的 section 列 + ref_case_id 建立 case_id → section 映射"""
    section_map = {}
    for sheet in ["chsm", "vsm", "pk_rsa", "pk_sm2", "section8", "section9", "section9_pk"]:
        df = pd.read_excel(TD_FILE, sheet_name=sheet, dtype=str)
        df = df.replace({np.nan: None})
        for _, r in df.iterrows():
            cid = r.get("case_id")
            sec = r.get("section")
            ref = _clean(r.get("ref_case_id"))
            if _valid_str(cid) and _valid_str(sec):
                section_map[cid] = sec
            if _valid_str(ref) and _valid_str(sec) and ref not in section_map:
                section_map[ref] = sec
    return section_map


def _get_section(case_id: str, section_map: dict) -> str:
    val = section_map.get(case_id)
    if _valid_str(val):
        return val
    for prefix, sec in sorted(SECTION_FALLBACK.items(), key=lambda x: -len(x[0])):
        if case_id.startswith(prefix):
            return sec
    return "99"


def get_module(case_id: str) -> str:
    for prefix, mod in MODULE_MAP.items():
        if case_id.startswith(prefix + "_"):
            return mod
    if case_id.startswith("PK_RSA"):
        return MODULE_MAP["PK_RSA"]
    if case_id.startswith("PK_SM2"):
        return MODULE_MAP["PK_SM2"]
    return "未分类"


def extract_interface_name(desc: str, case_id: str, method: str = "POST") -> str:
    m = re.match(r"^([a-zA-Z]+)", desc)
    if m:
        fn = m.group(1)
        if fn in INTERFACE_NAME_MAP:
            return INTERFACE_NAME_MAP[fn]

    for cn_key in ["上传镜像文件", "获取镜像描述", "下载文件", "删除镜像文件"]:
        if desc.startswith(cn_key):
            return INTERFACE_NAME_MAP[cn_key]

    if case_id in SETUP_CASE_MAP:
        return SETUP_CASE_MAP[case_id]

    if case_id.startswith("PK_RSA") or case_id.startswith("PK_SM2"):
        return _pk_interface_from_desc(desc, method)

    return desc.split(" ")[0] if " " in desc else desc


def _pk_interface_from_desc(desc: str, method: str = "POST") -> str:
    d = desc.lower()
    if "trusted" in d or "getCHSMInfo" in desc or "403" in desc:
        return "getCHSMInfo-获取CHSM详细信息"
    if ("清空" in d or "清除" in d) and "后" not in d:
        return "clearCHSMPk-清空认证公钥"
    if "查询" in d or "验证" in d or "指纹" in d or "确认" in d:
        return "getCHSMPk-获取认证公钥指纹"
    if "配置" in d or "追加" in d or "公钥" in d or "去重" in d or "恢复" in d or "覆盖" in d:
        return "configCHSMPk-配置认证公钥"
    if method == "GET":
        return "getCHSMPk-获取认证公钥指纹"
    if method == "DELETE":
        return "clearCHSMPk-清空认证公钥"
    return "configCHSMPk-配置认证公钥"


def get_auth_type(case_id: str, sheet: str) -> str:
    if sheet in ("chsm", "vsm"):
        if case_id.startswith("GUEST_") or case_id.startswith("CHSM_STATUS") or \
           case_id.startswith("CHSM_ALLSTATUS") or case_id.startswith("CHSM_DEBUG") or \
           case_id.startswith("VSM_STATUS") or case_id.startswith("AUTH_403"):
            return "guest"
        return "trusted"
    if sheet in ("pk_rsa", "pk_sm2"):
        return "pk"
    if sheet == "section8":
        return "fileserver"
    if sheet == "section9", "section9_pk":
        return "tenant"
    return "trusted"


def get_headers(auth_type: str, desc: str, method: str, case_id: str) -> str:
    if auth_type == "guest":
        if "不带认证" in desc or "trusted" in desc.lower():
            return AUTH_HEADER_NONE
        return AUTH_HEADER_GUEST
    if auth_type == "trusted":
        return AUTH_HEADER_TRUSTED
    if auth_type == "pk":
        d = desc.lower()
        if "首次" in d:
            return "首次无需认证头（guest）"
        if "查询" in d or "验证" in d or "确认" in d or "指纹" in d:
            return AUTH_HEADER_GUEST
        if "清空" in d or "清除" in d:
            if method == "DELETE":
                return AUTH_HEADER_TRUSTED
            return AUTH_HEADER_GUEST
        if "trusted" in d or "403" in desc:
            return AUTH_HEADER_NONE
        return AUTH_HEADER_TRUSTED
    if auth_type == "fileserver":
        return "token认证"
    if auth_type == "tenant":
        return AUTH_HEADER_TENANT
    return AUTH_HEADER_TRUSTED


def get_precondition(auth_type: str, desc: str, case_id: str,
                     scenario_id: Optional[str], step: Optional[str]) -> str:
    scenario_id = _clean(scenario_id)
    step = _clean(step)
    if auth_type == "guest":
        if "VSM" in case_id and "vsmId" in desc:
            return "CHSM已运行"
        return "CHSM已正常运行"
    if auth_type == "trusted":
        base = "CHSM已正常运行，已配置公钥"
        if "VSM" in case_id:
            base = "CHSM已运行，至少有1个VSM，已配置公钥"
        if scenario_id:
            base += f"\n场景 {scenario_id} step {step or '?'}"
        return base
    if auth_type == "pk":
        if scenario_id:
            return f"场景 {scenario_id} 按步骤顺序执行"
        return "密码机上已配置公钥"
    if auth_type == "fileserver":
        return "FileServer 服务可达"
    if auth_type == "tenant":
        if scenario_id:
            return f"场景 {scenario_id} step {step or '?'} 按步骤执行"
        return "VSM 已初始化"
    return "CHSM已正常运行"


def format_request_params(json_data, params, method: str) -> str:
    if json_data:
        if isinstance(json_data, str):
            try:
                obj = json.loads(json_data)
                return json.dumps(obj, ensure_ascii=False, indent=2)
            except (json.JSONDecodeError, TypeError):
                return json_data
        elif isinstance(json_data, dict):
            return json.dumps(json_data, ensure_ascii=False, indent=2)
    if params:
        if isinstance(params, str):
            return params
        elif isinstance(params, dict):
            return "&".join(f"{k}={v}" for k, v in params.items())
    return ""


def build_expected_response(status: int) -> str:
    if status == 200:
        return '{\n  "status": 200,\n  "message": "success",\n  "timestamp": "<ISO8601>",\n  "requestId": "<echo-back>",\n  "costMillis": "<int>"\n}'
    elif status == 400:
        return '{\n  "status": 400,\n  "message": "<参数错误描述>",\n  "timestamp": "<ISO8601>",\n  "requestId": "<echo-back>"\n}'
    elif status == 403:
        return '{\n  "status": 403,\n  "message": "认证失败",\n  "timestamp": "<ISO8601>"\n}'
    else:
        return f'{{\n  "status": {status}\n}}'


def build_expected_desc(status: int, desc: str) -> str:
    if status == 200:
        if "查询" in desc or "验证" in desc or "获取" in desc:
            return "返回200，结果符合预期"
        return "返回200，操作成功"
    elif status == 400:
        detail = ""
        if "缺" in desc:
            m = re.search(r"缺少?(\w+)", desc)
            if m:
                detail = f"，缺少{m.group(1)}参数"
        elif "空" in desc:
            detail = "，参数为空"
        elif "非法" in desc or "无效" in desc:
            detail = "，参数格式非法"
        elif "超限" in desc:
            detail = "，超出限制"
        return f"返回400，参数校验失败{detail}"
    elif status == 403:
        return "返回403，认证失败"
    return f"返回{status}"


def get_test_type(status: int, desc: str, step_type: Optional[str]) -> str:
    step_type = _clean(step_type)
    if step_type in ("setup", "teardown", "verify"):
        return "场景步骤"
    if status != 200:
        return "异常测试"
    if any(kw in desc for kw in ["超限", "失败", "非法", "无效", "错误"]):
        return "异常测试"
    return "功能测试"


def get_priority(status: int, step_type: Optional[str]) -> str:
    step_type = _clean(step_type)
    if step_type in ("setup", "teardown", "verify"):
        return "P0"
    if status == 200:
        return "P0"
    return "P1"


def _clean(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "nan", "None"):
        return None
    return s


def get_remark(enabled: str, scenario_id: Optional[str],
               step: Optional[str], step_type: Optional[str],
               ref_case_id: Optional[str]) -> str:
    parts = []
    enabled = _clean(enabled)
    if enabled and enabled.lower() in ("no", "false", "0"):
        parts.append("enabled=no(自动化暂不执行)")
    scenario_id = _clean(scenario_id)
    step = _clean(step)
    step_type = _clean(step_type)
    ref_case_id = _clean(ref_case_id)
    if scenario_id:
        parts.append(f"场景:{scenario_id}")
    if step:
        parts.append(f"step:{step}")
    if step_type:
        parts.append(f"type:{step_type}")
    if ref_case_id:
        parts.append(f"关联:{ref_case_id}")
    return "；".join(parts) if parts else None


# ── openpyxl 样式 ────────────────────────────────────

HEADER_FONT = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_FONT = Font(name="微软雅黑", size=9)
DATA_ALIGN = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def write_styled_excel(filepath: str, sheet_name: str, headers: list, rows: list):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    for ri, row_data in enumerate(rows, 2):
        for ci, h in enumerate(headers, 1):
            val = row_data.get(h) or ""
            cell = ws.cell(ri, ci, val)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = THIN_BORDER

    for ci, h in enumerate(headers, 1):
        letter = get_column_letter(ci)
        ws.column_dimensions[letter].width = COL_WIDTHS.get(h, 13)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(filepath)


def main():
    section_map = _build_section_map()

    # 1. 读取现有测试用例
    tc_df = pd.read_excel(TC_FILE, sheet_name=TC_SHEET, dtype=str)
    tc_df = tc_df.replace({np.nan: None})
    existing_ids = set(tc_df["用例ID"].dropna().tolist())
    print(f"现有测试用例: {len(existing_ids)} 条")

    existing_rows = []
    for _, r in tc_df.iterrows():
        cid = r.get("用例ID")
        if not cid:
            continue
        row_data = {col: r.get(col) for col in TC_COLS}
        sec = _get_section(cid, section_map)
        row_data["需求章节"] = sec
        row_data["_section"] = sec
        row_data["_case_id"] = cid
        existing_rows.append(row_data)

    # 2. 从 test_data 生成缺失用例
    prefix_map = {}
    for row_data in existing_rows:
        cid = row_data["_case_id"]
        prefix = re.sub(r"_\d+[A-Z]?$", "", cid)
        if prefix not in prefix_map:
            prefix_map[prefix] = {
                "接口名称": row_data["接口名称"],
                "接口路径": row_data["接口路径"],
            }

    new_rows = []
    for sheet in ["chsm", "vsm", "pk_rsa", "pk_sm2", "section8", "section9", "section9_pk"]:
        td_df = pd.read_excel(TD_FILE, sheet_name=sheet, dtype=str)
        td_df = td_df.replace({np.nan: None})

        for _, r in td_df.iterrows():
            case_id = r.get("case_id")
            if not case_id or case_id in existing_ids:
                continue

            desc = str(r.get("description", "") or "")
            endpoint = str(r.get("endpoint", "") or "")
            method = str(r.get("method", "") or "")
            json_data = r.get("json_data")
            params = r.get("params")
            expected_status = int(r.get("expected_status", 0) or 0)
            enabled = r.get("enabled")
            scenario_id = r.get("scenario_id")
            step = r.get("step")
            step_type = r.get("step_type")
            ref_case_id = r.get("ref_case_id")
            section = r.get("section") or ""

            auth_type = get_auth_type(case_id, sheet)
            module = get_module(case_id)

            iface_name = extract_interface_name(desc, case_id, method)
            td_prefix = re.sub(r"_T\d+$", "", re.sub(r"_E\d+$", "", re.sub(r"_\d+[A-Z]?$", "", case_id)))
            if td_prefix in prefix_map and iface_name == desc.split(" ")[0]:
                iface_name = prefix_map[td_prefix]["接口名称"]

            headers = get_headers(auth_type, desc, method, case_id)
            precondition = get_precondition(auth_type, desc, case_id, scenario_id, step)
            req_params = format_request_params(json_data, params, method)
            expected_resp = build_expected_response(expected_status)
            expected_desc = build_expected_desc(expected_status, desc)
            test_type = get_test_type(expected_status, desc, step_type)
            priority = get_priority(expected_status, step_type)
            remark = get_remark(enabled, scenario_id, step, step_type, ref_case_id)

            sec = _valid_str(section) and section or _get_section(case_id, section_map)
            row_data = {
                "用例ID": case_id,
                "需求章节": sec,
                "测试模块": module,
                "接口名称": iface_name,
                "接口路径": endpoint,
                "请求方法": method,
                "测试场景": desc,
                "测试类型": test_type,
                "优先级": priority,
                "前置条件": precondition,
                "请求头(Headers)": headers,
                "请求参数(Query/Body)/步骤": req_params,
                "预期响应状态码": str(expected_status),
                "预期响应": expected_resp,
                "预期结果描述": expected_desc,
                "测试状态": None,
                "执行人": None,
                "执行时间": None,
                "缺陷ID": None,
                "备注": remark,
                "_section": sec,
                "_case_id": case_id,
            }
            new_rows.append(row_data)

    print(f"新增测试用例: {len(new_rows)} 条")

    # 3. 合并 + 按章节排序
    all_rows = existing_rows + new_rows
    all_rows.sort(key=lambda r: (
        _parse_section(r.get("_section", "99")),
        r.get("_case_id", ""),
    ))

    for r in all_rows:
        r.pop("_section", None)
        r.pop("_case_id", None)

    # 4. 写入带样式的 Excel
    write_styled_excel(TC_FILE, TC_SHEET, TC_COLS, all_rows)

    total = len(all_rows)
    print(f"\n已保存: {TC_FILE}")
    print(f"总计: {total} 条测试用例 (原 {len(existing_ids)} + 新 {len(new_rows)})")

    # 5. 统计
    from collections import Counter
    mod_counter = Counter(r.get("测试模块", "未分类") for r in all_rows)
    type_counter = Counter(r.get("测试类型", "未分类") for r in all_rows)

    print("\n=== 按模块统计 ===")
    for mod, cnt in sorted(mod_counter.items()):
        print(f"  {mod:25s}: {cnt} 条")

    print("\n=== 按测试类型统计 ===")
    for tt, cnt in sorted(type_counter.items()):
        print(f"  {tt:15s}: {cnt} 条")


if __name__ == "__main__":
    main()
