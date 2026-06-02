#!/usr/bin/env python3
"""生成接口测试结果清单 Excel（含入参/出参/认证/接口介绍/服务地址）"""

import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 服务地址常量
SVC_CHSM = "CHSM管理面\nhttps://<CHSM_IP>:7443"
SVC_FS = "FileServer\nhttp://<FS_IP>:<FS_PORT>"
SVC_VSM = "VSM业务面\nhttps://<VSM_IP>:7443"

# 通用响应格式
RESP_STD = '{"status":int, "message":str,\n "timestamp":str, "requestId":str,\n "costMillis":int}'
RESP_RESULT = RESP_STD.replace('}', ',\n "result":{...}}')

DATA = [
    # (大模块, 小模块, 接口名, 接口介绍, 服务, endpoint, 请求方式, 认证, 入参, 出参, 测试结果)

    # ===== 7.1 CHSM配置管理 =====
    ("7.1 CHSM配置管理", "7.1.1", "getCHSMInfo",
     "获取CHSM详细信息，包括型号、固件版本、序列号、VSM实例列表等",
     SVC_CHSM, "/chsm", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n "oprType":"getinfo"}',
     'result: {\n  model, version, serialNo,\n  vsmIds:["id1","id2",...],\n  ...}',
     ""),
    ("7.1 CHSM配置管理", "7.1.2", "getCHSMStatus",
     "获取CHSM当前运行状态(无需认证)",
     SVC_CHSM, "/chsm/status", "GET",
     "guest",
     "无",
     'result: {status, uptime, ...}', ""),
    ("7.1 CHSM配置管理", "7.1.3", "getCHSMAllStatus",
     "获取CHSM所有状态信息，含各组件及VSM状态",
     SVC_CHSM, "/chsm/allstatus", "GET",
     "guest",
     "无",
     'result: {\n  chsmStatus,\n  vsmStatus:[...], ...}', ""),
    ("7.1 CHSM配置管理", "7.1.4", "configCHSMNet",
     "配置CHSM网络参数：DNS列表或网口IP/掩码/网关",
     SVC_CHSM, "/chsm/network", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n "dnsList":["8.8.8.8","114.114.114.114"]}\n或\n{"requestId":"uuid",\n "name":"eth0", "ip":"x",\n "mask":"x", "gateway":"x"}',
     RESP_STD, ""),
    ("7.1 CHSM配置管理", "7.1.5", "configCHSMNtp",
     "配置NTP时间同步服务器地址和同步周期",
     SVC_CHSM, "/chsm/ntp", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n "addr":"10.10.1.1",\n "syncPeriod":60}',
     RESP_STD, ""),
    ("7.1 CHSM配置管理", "7.1.6", "configCHSMUploadAddress",
     "配置CHSM影像上传目标地址",
     SVC_CHSM, "/chsm/imageuploader", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n "url":"http://..."}',
     RESP_STD, ""),
    ("7.1 CHSM配置管理", "7.1.7", "configSyslogAddr",
     "配置日志上传地址，支持syslog和logserver两种类型",
     SVC_CHSM, "/chsm/loguploader", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n "logServerType":"syslog|logserver",\n "logServerAddress":"ip:port"}',
     RESP_STD, ""),
    ("7.1 CHSM配置管理", "7.1.8", "exportCHSM",
     "导出CHSM影像，异步操作，通过callbackUrl回调通知结果",
     SVC_CHSM, "/chsm/image", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n "oprType":"export",\n "callbackUrl":"http://..."}',
     RESP_STD, ""),
    ("7.1 CHSM配置管理", "7.1.9", "importCHSM",
     "导入CHSM影像，需提供影像URL及数字签名校验",
     SVC_CHSM, "/chsm/image", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n "oprType":"import",\n "imageUrl":"url",\n "alg":"RSAWithSHA256",\n "sign":"base64签名",\n "callbackUrl":"url"}',
     RESP_STD, ""),
    ("7.1 CHSM配置管理", "7.1.10", "upgradeCHSM",
     "升级CHSM固件，需提供升级包URL及数字签名",
     SVC_CHSM, "/chsm", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n "oprType":"upgrade",\n "packVersion":"x.x",\n "packUrl":"url",\n "alg":"...", "sign":"...",\n "callbackUrl":"url"}',
     RESP_STD, ""),
    ("7.1 CHSM配置管理", "7.1.11", "restartCHSM",
     "重启CHSM（高危操作），异步回调通知结果",
     SVC_CHSM, "/chsm", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n "oprType":"restart",\n "callbackUrl":"url"}',
     RESP_STD, ""),
    ("7.1 CHSM配置管理", "7.1.12", "backupCHSM",
     "备份CHSM配置和数据，异步回调通知",
     SVC_CHSM, "/chsm", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n "oprType":"backup",\n "callbackUrl":"url"}',
     RESP_STD, ""),
    ("7.1 CHSM配置管理", "7.1.13", "restoreCHSM",
     "恢复CHSM配置和数据，需签名校验",
     SVC_CHSM, "/chsm", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n "oprType":"restore",\n "backupUrl":"url",\n "alg":"...", "sign":"...",\n "callbackUrl":"url"}',
     RESP_STD, ""),
    ("7.1 CHSM配置管理", "7.1.14", "configCHSMAlarmAddress",
     "配置告警通知地址和监控地址",
     SVC_CHSM, "/chsm/alarmaddress", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n "url":"http://...",\n "monitoringUrl":"http://..."}',
     RESP_STD, ""),
    ("7.1 CHSM配置管理", "7.1.15", "configCHSMToken",
     "配置云平台对接Token(JWT)",
     SVC_CHSM, "/chsm/cloudtoken", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n "cloudToken":"JWT字符串"}',
     RESP_STD, ""),
    ("7.1 CHSM配置管理", "7.1.16", "configCHSMMOOCAddress",
     "配置MO/OC对接地址，含OC和MO的IP、协议、端口、资源路径",
     SVC_CHSM, "/chsm/moocaddress", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n "ocIp":"ip", "ocProtocol":"https",\n "ocPort":"26635",\n "resourceUrl":"...",\n "moIp":"ip", "moProtocol":"https",\n "moPort":"26636"}',
     RESP_STD, ""),
    ("7.1 CHSM配置管理", "7.1.17", "getCHSMDebugInfo",
     "获取CHSM调试诊断信息",
     SVC_CHSM, "/chsm/debuginfo", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid"}',
     RESP_RESULT, ""),
    ("7.1 CHSM配置管理", "7.1.18", "getCHSMDeviceInfo",
     "获取CHSM设备硬件信息",
     SVC_CHSM, "/chsm", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n "oprType":"getdeviceinfo"}',
     RESP_RESULT, ""),

    # ===== 7.2 VSM配置管理 =====
    ("7.2 VSM配置管理", "7.2.1", "getVSMInfo",
     "获取指定VSM实例的详细信息",
     SVC_CHSM, "/vsm", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n   "oprType":"getinfo",\n   "vsmId":"VSM实例ID"}',
     'result: {\n  vsmId, status,\n  version, ip, token, ...}', ""),
    ("7.2 VSM配置管理", "7.2.2", "getVSMStatus",
     "获取指定VSM运行状态，返回ok或fail",
     SVC_CHSM, "/vsm/status", "GET",
     "guest",
     "Query:\n  requestId=uuid\n  vsmId=VSM实例ID",
     'result: {status:"ok|fail"}', ""),
    ("7.2 VSM配置管理", "7.2.3", "configVSMNet",
     "配置VSM网络IP/掩码/网关，支持IPv6前缀长度",
     SVC_CHSM, "/vsm/network", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n   "vsmId":"VSM实例ID",\n   "ip":"ip地址",\n   "mask":"子网掩码(IPv6为前缀长度如64)",\n   "gateway":"网关"}',
     RESP_STD, ""),
    ("7.2 VSM配置管理", "7.2.4", "configVSMToken",
     "配置VSM租户Token和租户ID，用于标识VSM归属租户",
     SVC_CHSM, "/vsm/token", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n   "vsmId":"VSM实例ID",\n   "token":"租户标识符(0=不使用)",\n   "tenantId":"租户ID"}',
     RESP_STD, ""),
    ("7.2 VSM配置管理", "7.2.5", "exportVSM",
     "导出VSM数据影像，异步回调通知结果",
     SVC_CHSM, "/vsm/image", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n   "oprType":"export",\n   "vsmId":"VSM实例ID",\n   "callbackUrl":"回调地址"}',
     RESP_STD, ""),
    ("7.2 VSM配置管理", "7.2.6", "importVSM",
     "导入VSM数据影像，需数字签名校验",
     SVC_CHSM, "/vsm/image", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n   "oprType":"import",\n   "vsmId":"VSM实例ID",\n   "imageUrl":"影像文件URL",\n   "alg":"RSAWithSHA256|SM2WithSM3",\n   "sign":"数字签名BASE64",\n   "callbackUrl":"回调地址"}',
     RESP_STD, ""),
    ("7.2 VSM配置管理", "7.2.7", "startVSM",
     "启动指定VSM实例，异步回调",
     SVC_CHSM, "/vsm", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n   "oprType":"start",\n   "vsmId":"VSM实例ID",\n   "callbackUrl":"回调地址"}',
     RESP_STD, ""),
    ("7.2 VSM配置管理", "7.2.8", "stopVSM",
     "停止指定VSM实例，异步回调",
     SVC_CHSM, "/vsm", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n   "oprType":"stop",\n   "vsmId":"VSM实例ID",\n   "callbackUrl":"回调地址"}',
     RESP_STD, ""),
    ("7.2 VSM配置管理", "7.2.9", "restartVSM",
     "重启指定VSM实例，异步回调",
     SVC_CHSM, "/vsm", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n   "oprType":"restart",\n   "vsmId":"VSM实例ID",\n   "callbackUrl":"回调地址"}',
     RESP_STD, ""),
    ("7.2 VSM配置管理", "7.2.10", "resetVSM",
     "重置VSM到出厂状态，所有密钥/token/ukey信息全部清空（高危操作）",
     SVC_CHSM, "/vsm", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n   "oprType":"reset",\n   "vsmId":"VSM实例ID",\n   "callbackUrl":"回调地址"}',
     RESP_STD, ""),
    ("7.2 VSM配置管理", "7.2.11", "upgradeVSM",
     "升级VSM固件，需提供升级包URL及数字签名，升级前暂停VSM服务",
     SVC_CHSM, "/vsm", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n   "oprType":"upgrade",\n   "vsmId":"VSM实例ID",\n   "packVersion":"版本号",\n   "packUrl":"升级包URL",\n   "alg":"RSAWithSHA256|SM2WithSM3",\n   "sign":"数字签名",\n   "callbackUrl":"回调地址"}',
     RESP_STD, ""),
    ("7.2 VSM配置管理", "7.2.12", "createVSM",
     "创建VSM实例，可指定镜像URL创建，异步回调（预留接口）",
     SVC_CHSM, "/vsm", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n   "oprType":"create",\n   "imageUrl":"镜像文件URL(可选)",\n   "alg":"RSAWithSHA256|SM2WithSM3(imageUrl存在时必填)",\n   "sign":"数字签名BASE64(imageUrl存在时必填)",\n   "flavor":1|2|3,\n   "callbackUrl":"回调地址"}',
     RESP_STD + '\n+ vsmId:新建VSM的实例ID', "预留，不测"),
    ("7.2 VSM配置管理", "7.2.13", "deleteVSM",
     "删除VSM实例，删除数据和影像恢复到初始状态（高危操作，预留接口）",
     SVC_CHSM, "/vsm", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n   "oprType":"destroy",\n   "vsmId":"VSM实例ID",\n   "callbackUrl":"回调地址"}',
     RESP_STD, "预留，不测"),
    ("7.2 VSM配置管理", "7.2.14", "configVSMMac",
     "配置VSM的MAC地址（预留接口）",
     SVC_CHSM, "/vsm/mac", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n   "vsmId":"VSM实例ID",\n   "mac":"ba:ca:c1:d2:e3:e4"}',
     RESP_STD, "预留，不测"),
    ("7.2 VSM配置管理", "7.2.15", "configVSMVlan",
     "为VSM配置或删除VLAN，支持Tag VLAN（预留接口）",
     SVC_CHSM, "/vsm/vlan", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n   "oprType":"modify|delete",\n   "vsmId":"VSM实例ID",\n   "vlanType":"Tag vlan",\n   "vlanId":"2-4094"}',
     RESP_STD, "预留，不测"),
    ("7.2 VSM配置管理", "7.2.16", "deleteVSMNetwork",
     "删除VSM的网络属性配置（预留接口）",
     SVC_CHSM, "/vsm/network", "DELETE",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n   "vsmId":"VSM实例ID"}',
     RESP_STD, "预留，不测"),
    ("7.2 VSM配置管理", "7.2.17", "getVSMDeviceInfo",
     "获取VSM设备信息，含version/ip/sn/mac/hwAuthCode（类似7.1.18）",
     SVC_CHSM, "/vsm", "POST",
     "trusted",
     'Body:\n  {"requestId":"uuid",\n   "oprType":"getDeviceinfo"}',
     'result: {\n  version, ip, sn,\n  mac, hwAuthCode}', ""),

    # ===== 7.3 授权配置 =====
    ("7.3 授权配置", "7.3.1", "getCHSMPk",
     "获取CHSM认证公钥信息，得到CHSM中的公钥杂凑值，算法为SM3或SHA256等，返回值为BASE64编码",
     SVC_CHSM, "/chsm/authpk", "GET",
     "guest",
     "Query:\n  requestId=uuid",
     'result: {\n  algorithm:"sha256|sm3",\n  fingerprints:["fp1","fp2"]}', ""),
    ("7.3 授权配置", "7.3.2", "configCHSMPk",
     "配置CHSM认证公钥，最多2个，按内容去重；首次配置guest权限即可，后续需trusted",
     SVC_CHSM, "/chsm/authpk", "POST",
     "首次guest/后续trusted",
     'Body:\n  {"requestId":"uuid",\n   "algorithm":"RSAWithSHA256|sm2",\n   "pks":["公钥字符串",...]}',
     RESP_STD, ""),
    ("7.3 授权配置", "7.3.3", "clearCHSMPk",
     "清空CHSM所有已配置的认证公钥（清空后trusted接口返回403）",
     SVC_CHSM, "/chsm/authpk", "DELETE",
     "trusted",
     'Body:\n  {"requestId":"uuid"}',
     RESP_STD, ""),

    # ===== 8.1 文件服务 =====
    ("8.1 文件服务", "8.1.1", "uploadImage",
     "上传镜像/固件文件到FileServer（云管平台调用，multipart）",
     SVC_FS, "/images", "POST",
     "trusted",
     'multipart/form-data:\n  file=<二进制文件>\n  requestId=uuid\n  info={"token":"文件路径",\n    "type":"hsm|vsm",\n    "uuid":"设备唯一标识",\n    "version":"版本号",\n    "sign":"签名base64",\n    "alg":"sm2|rsa"}',
     'Body:\n  {"requestId":"uuid",\n "success":true,\n "message":""}', ""),
    ("8.1 文件服务", "8.1.2", "getImageInfo",
     "获取镜像文件描述信息（云管平台调用）",
     SVC_FS, "/info", "GET",
     "trusted",
     "Query:\n  requestId=uuid",
     'result: {\n  info: {token, type,\n    uuid, version,\n    sign, alg}}', ""),
    ("8.1 文件服务", "8.1.3", "downloadFile",
     "下载指定镜像文件，返回二进制文件流",
     SVC_FS, "/images", "POST",
     "trusted",
     "Query:\n  file=文件路径",
     "application/octet-stream\n(二进制文件流)", ""),
    ("8.1 文件服务", "8.1.4", "deleteImage",
     "删除FileServer上的指定镜像文件（云管平台调用）",
     SVC_FS, "/images/{requestId}/{token}", "POST",
     "trusted",
     "Path:\n  requestId=请求ID\n  token=待删除文件路径",
     RESP_STD, ""),

    # ===== 9.2 租户密评 =====
    ("9.2 租户密评", "9.2.1", "getAuthPKFingerprints",
     "获取VSM已配置的认证公钥指纹列表",
     SVC_VSM, "/authServlet", "GET",
     "guest",
     "Query:\n  method=getAuthPKFingerprints\n  requestId=uuid",
     'result: {\n  algorithm:"sha256|sm3",\n  fingerprints:[...]}', ""),
    ("9.2 租户密评", "9.2.2", "authPK",
     "配置VSM认证公钥，最多2个，支持RSA/SM2混合",
     SVC_VSM, "/authServlet", "POST",
     "guest",
     'Query:\n  method=authPK\nBody:\n  {"requestId":"uuid",\n   "alg":"RSAWithSHA256|sm2",\n   "pks":["公钥字符串",...]}',
     RESP_STD, ""),
    ("9.2 租户密评", "9.2.3", "cleanPK",
     "清除VSM认证公钥，可指定fingerprint删除单个或不传清除全部",
     SVC_VSM, "/authServlet", "POST",
     "guest",
     'Query:\n  method=cleanPK\nBody:\n  {"requestId":"uuid"}\n可选:\n  {"fingerprint":"fp值",\n   "action":"remove"}',
     RESP_STD, ""),
    ("9.2 租户密评", "9.2.4", "initKey",
     "初始化VSM本地主密钥（高危操作）",
     SVC_VSM, "/platformServlet", "POST",
     "租户签名",
     "Query:\n  method=initKey\n  requestId=uuid",
     RESP_STD, ""),
    ("9.2 租户密评", "9.2.5", "exportBackupKeys",
     "导出VSM数据影像密钥包，返回二进制流",
     SVC_VSM, "/platformServlet", "POST",
     "租户签名",
     'Query:\n  method=exportBackupKeys\n  requestId=uuid\nBody:\n  {"requestId":"uuid",\n   "backupKey":"密钥串"}',
     "application/octet-stream\n(二进制密钥包)", ""),
    ("9.2 租户密评", "9.2.6", "importBackupKeys",
     "导入VSM数据影像密钥包，multipart文件上传",
     SVC_VSM, "/platformServlet", "POST",
     "租户签名",
     'Query:\n  method=importBackupKeys\nmultipart:\n  file=<密钥包文件>\nBody:\n  {"requestId":"uuid",\n   "backupKey":"密钥串"}',
     RESP_STD, ""),
    ("9.2 租户密评", "9.2.7", "doVsmInit",
     "初始化VSM实例（高危操作），可选是否清除公钥",
     SVC_VSM, "/platformServlet", "POST",
     "租户签名",
     'Query:\n  method=doVsmInit\nBody:\n  {"requestId":"uuid",\n   "clearPK":"false|true"}',
     RESP_STD, ""),
    ("9.2 租户密评", "9.2.8", "getStatus",
     "获取VSM运行状态",
     SVC_VSM, "/platformServlet", "GET",
     "租户签名",
     "Query:\n  method=getStatus",
     'result: {status, ...}', ""),
]

HEADERS = ["序号", "大模块", "小模块", "接口名称", "接口介绍",
           "服务 (IP:Port)", "Endpoint", "请求方式", "认证方式",
           "入参", "出参", "测试结果"]
COL_WIDTHS = [6, 18, 8, 22, 40, 20, 32, 10, 18, 48, 34, 12]

# 样式
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
CELL_FONT = Font(name="微软雅黑", size=9)
CELL_FONT_MONO = Font(name="Consolas", size=8.5)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

STRIPE_FILLS = [
    PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
    PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
]

AUTH_COLORS = {
    "trusted": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "guest": PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"),
    "guest": PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"),
    "首次guest/后续trusted": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),


}


def main():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "接口测试结果清单"

    # 表头
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # 列宽
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 数据
    module_index = {}
    color_idx = 0
    # 需要居中的列 (1-based): 序号、小模块、请求方式、测试结果
    center_cols = {1, 3, 8, 12}
    # 等宽字体列: Endpoint、入参、出参
    mono_cols = {7, 10, 11}

    for row_idx, item in enumerate(DATA, 2):
        big_mod = item[0]
        if big_mod not in module_index:
            module_index[big_mod] = color_idx % len(STRIPE_FILLS)
            color_idx += 1
        fill = STRIPE_FILLS[module_index[big_mod]]

        values = [row_idx - 1] + list(item)
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.fill = fill

            if col_idx in center_cols:
                cell.font = CELL_FONT
                cell.alignment = CENTER
            elif col_idx in mono_cols:
                cell.font = CELL_FONT_MONO
                cell.alignment = LEFT_TOP
            else:
                cell.font = CELL_FONT
                cell.alignment = LEFT_TOP

            # 认证方式列着色 (col 9)
            if col_idx == 9 and val in AUTH_COLORS:
                cell.fill = AUTH_COLORS[val]
                cell.alignment = CENTER

    # 合并大模块列 (col B)
    merge_col = 2
    start_row = 2
    for row_idx in range(3, len(DATA) + 3):
        prev = ws.cell(row=row_idx - 1, column=merge_col).value
        curr = ws.cell(row=row_idx, column=merge_col).value if row_idx <= len(DATA) + 1 else None
        if curr != prev:
            if row_idx - 1 > start_row:
                ws.merge_cells(start_row=start_row, start_column=merge_col,
                               end_row=row_idx - 1, end_column=merge_col)
            start_row = row_idx

    # 合并服务列 (col F) 相同值
    svc_col = 6
    start_row = 2
    for row_idx in range(3, len(DATA) + 3):
        prev = ws.cell(row=row_idx - 1, column=svc_col).value
        curr = ws.cell(row=row_idx, column=svc_col).value if row_idx <= len(DATA) + 1 else None
        if curr != prev:
            if row_idx - 1 > start_row:
                ws.merge_cells(start_row=start_row, start_column=svc_col,
                               end_row=row_idx - 1, end_column=svc_col)
            start_row = row_idx

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30
    for r in range(2, len(DATA) + 2):
        ws.row_dimensions[r].height = 72

    out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data", "接口测试结果清单.xlsx")
    wb.save(out_path)
    print(f"已生成: {out_path}")
    print(f"共 {len(DATA)} 条接口, 列: {' | '.join(HEADERS)}")


if __name__ == "__main__":
    main()
