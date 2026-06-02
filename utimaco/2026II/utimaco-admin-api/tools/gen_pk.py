#!/usr/bin/env python3
"""
生成 pk_rsa / pk_sm2 sheet — 公钥认证接口(7.3)的场景测试和独立错误测试。

7.3 节公钥配置没有数量限制（区别于 Section 9 租户接口的最多 2 个限制）。

每个 sheet 包含 3 个场景 + 10 条独立错误用例:
  SCN_PK_{RSA|SM2}_SINGLE  (5 步)  — 单公钥的配置/查询/幂等/清理
  SCN_PK_{RSA|SM2}_MULTI   (9 步)  — 多公钥追加/跨算法/清空/401验证
  SCN_PK_{RSA|SM2}_BATCH   (11 步) — 批量传入/去重/混合算法/恢复
  独立错误用例 E01-E10             — 参数校验
"""

import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

# ================================================================
# 常量
# ================================================================

EP = "/api/1.0/chsm/authpk"
EP_CHSM = "/api/1.0/chsm"

# configCHSMPk API 请求体中 algorithm 字段的正确取值
API_ALG = {"rsa": "RSAWithSHA256", "sm2": "sm2"}

RSA_KEY1 = "${keys.rsa.key1.public_key_pem}"
RSA_KEY2 = "${keys.rsa.key2.public_key_pem}"
RSA_KEY3 = "${keys.rsa.key3.public_key_pem}"
SM2_KEY1 = "${keys.sm2.key1.public_key_pem}"
SM2_KEY2 = "${keys.sm2.key2.public_key_pem}"
SM2_KEY3 = "${keys.sm2.key3.public_key_pem}"

RSA_FP1 = "/WiF31yZx0q9nF0fQBudD7xSeivFBfdEhG8hl/KQo1c="
RSA_FP2 = "8CPhhUBOk75PaceSD6UjfghIh2V0xEo+1v2nUN4k7uc="
RSA_FP3 = "EZe+Or1pc78hK6RoKjqYPGuZ7h59ohDZqJBa7xBCgb8="
SM2_FP1 = "mdg2R+NkXUI3WidsEba3Dkg/MBx98JaXO23L9XF8xK8="
SM2_FP2 = "RSiUBOLEkAC8AzPRie3UiN8UPHlrLyA0SVpYGEO7tUc="
SM2_FP3 = "gr1fZe+LtAM9IA5QhUWOj7Wm0fCavaBXn5oHLS/9UCM="

# getCHSMPkFingerprints 响应中 result.algorithm 的取值（指纹哈希算法）
RSA_HASH_ALG = "sha256"
SM2_HASH_ALG = "sm3"

HEADERS = [
    "case_id", "description", "host", "endpoint", "method",
    "params", "json_data", "expected_status", "assert_rules",
    "scenario_id", "step", "step_type", "save_vars",
    "enabled", "section", "ref_case_id", "auth", "approved",
]

OK = json.dumps([
    {"type": "status_code", "expected_code": 200},
    {"type": "json_contains", "key": "message", "value": "success"},
], ensure_ascii=False)

ERR400 = json.dumps([
    {"type": "status_code", "expected_code": 400},
], ensure_ascii=False)

ERR401 = json.dumps([
    {"type": "status_code", "expected_code": 401},
], ensure_ascii=False)


# ================================================================
# 工具函数
# ================================================================

def jd(d):
    return json.dumps(d, ensure_ascii=False)


SECTION_BY_METHOD = {"GET": "7.3.1", "POST": "7.3.2", "DELETE": "7.3.3"}


def row(case_id, desc, endpoint, expected_status, json_data=None, params=None,
        assert_rules=None, ref_case_id=None, enabled="yes", section=None,
        scenario_id=None, step=None, step_type=None, save_vars=None,
        method="POST", approved="yes"):
    if assert_rules is None:
        assert_rules = OK if expected_status == 200 else ERR400
    auth = "no" if method == "GET" else "yes"
    if section is None:
        section = SECTION_BY_METHOD.get(method, "7.3")
    return {
        "case_id": case_id,
        "description": desc,
        "host": None,
        "endpoint": endpoint,
        "method": method,
        "params": params,
        "json_data": json_data,
        "expected_status": expected_status,
        "assert_rules": assert_rules,
        "scenario_id": scenario_id,
        "step": step,
        "step_type": step_type,
        "save_vars": save_vars,
        "enabled": enabled,
        "section": section,
        "ref_case_id": ref_case_id or "",
        "auth": auth,
        "approved": approved,
    }


def fp_assert_1(hash_alg, fp):
    return json.dumps([
        {"type": "status_code", "expected_code": 200},
        {"type": "json_contains", "key": "result.algorithm", "value": hash_alg},
        {"type": "json_contains", "key": "result.fingerprints[0]"},
        {"type": "json_not_contains", "key": "result.fingerprints[1]"},
    ], ensure_ascii=False)


def fp_assert_2(hash_alg, fp1, fp2):
    return json.dumps([
        {"type": "status_code", "expected_code": 200},
        {"type": "json_contains", "key": "result.algorithm", "value": hash_alg},
        {"type": "json_contains", "key": "result.fingerprints[0]"},
        {"type": "json_contains", "key": "result.fingerprints[1]"},
        {"type": "json_not_contains", "key": "result.fingerprints[2]"},
    ], ensure_ascii=False)


def fp_assert_3(hash_alg, fp1, fp2, fp3):
    return json.dumps([
        {"type": "status_code", "expected_code": 200},
        {"type": "json_contains", "key": "result.algorithm", "value": hash_alg},
        {"type": "json_contains", "key": "result.fingerprints[0]"},
        {"type": "json_contains", "key": "result.fingerprints[1]"},
        {"type": "json_contains", "key": "result.fingerprints[2]"},
        {"type": "json_not_contains", "key": "result.fingerprints[3]"},
    ], ensure_ascii=False)


def fp_assert_empty():
    return json.dumps([
        {"type": "status_code", "expected_code": 200},
        {"type": "json_not_contains", "key": "result.fingerprints[0]"},
    ], ensure_ascii=False)


def fp_assert_mixed():
    """混合算法指纹 — 检查指纹存在 + algorithm 字段存在"""
    return json.dumps([
        {"type": "status_code", "expected_code": 200},
        {"type": "json_contains", "key": "result.algorithm"},
        {"type": "json_contains", "key": "result.fingerprints[0]"},
        {"type": "json_contains", "key": "result.fingerprints[1]"},
    ], ensure_ascii=False)


# ================================================================
# 场景生成函数
# ================================================================

def gen_pk_single_scenario(alg, prefix, scn_prefix, keys, fps, hash_alg):
    """SCN_PK_{RSA|SM2}_SINGLE: 5 steps — 单公钥配置/查询/幂等/清理"""
    SID = f"{scn_prefix}_SINGLE"
    LABEL = alg.upper()
    ALG = API_ALG[alg]
    K1 = keys[0]
    FP1 = fps[0]
    a1 = fp_assert_1(hash_alg, FP1)

    return [
        row(f"{prefix}_001", f"首次配置1个{LABEL}公钥", EP, 200,
            jd({"requestId": "uuid", "algorithm": ALG, "pks": [K1]}),
            scenario_id=SID, step=1, step_type="setup",
            ref_case_id="AUTH_PK_001"),
        row(f"{prefix}_002", "查询验证1个指纹", EP, 200,
            params="requestId=uuid", method="GET",
            assert_rules=a1,
            scenario_id=SID, step=2, step_type="verify",
            ref_case_id="AUTH_PK_002"),
        row(f"{prefix}_003", f"重复配置相同{LABEL}公钥（幂等）", EP, 200,
            jd({"requestId": "uuid", "algorithm": ALG, "pks": [K1]}),
            scenario_id=SID, step=3, step_type="test"),
        row(f"{prefix}_004", "幂等后查询仍为1个指纹", EP, 200,
            params="requestId=uuid", method="GET",
            assert_rules=a1,
            scenario_id=SID, step=4, step_type="verify"),
        row(f"{prefix}_S01_CLEAR", "清空公钥(场景清理)", EP, 200,
            jd({"requestId": "uuid"}), method="DELETE",
            scenario_id=SID, step=5, step_type="teardown"),
    ]


def gen_pk_multi_scenario(alg, prefix, scn_prefix, keys, fps, hash_alg,
                          cross_alg, cross_keys):
    """SCN_PK_{RSA|SM2}_MULTI: 9 steps — 多公钥追加/跨算法/清空/401验证
    7.3 无数量限制: 3个同算法公钥 + 1个跨算法公钥均可成功配置。
    """
    SID = f"{scn_prefix}_MULTI"
    LABEL = alg.upper()
    ALG = API_ALG[alg]
    K1, K2, K3 = keys
    FP1, FP2, FP3 = fps
    CK1 = cross_keys[0]
    CROSS_ALG = API_ALG[cross_alg]
    CROSS_LABEL = cross_alg.upper()
    a2 = fp_assert_2(hash_alg, FP1, FP2)
    a3 = fp_assert_3(hash_alg, FP1, FP2, FP3)

    return [
        row(f"{prefix}_005", f"配置1个{LABEL}公钥(前置)", EP, 200,
            jd({"requestId": "uuid", "algorithm": ALG, "pks": [K1]}),
            scenario_id=SID, step=1, step_type="setup"),
        row(f"{prefix}_006", f"追加第2个{LABEL}公钥（总数2）", EP, 200,
            jd({"requestId": "uuid", "algorithm": ALG, "pks": [K2]}),
            scenario_id=SID, step=2, step_type="test",
            ref_case_id="AUTH_PK_016"),
        row(f"{prefix}_007", "查询验证2个指纹", EP, 200,
            params="requestId=uuid", method="GET",
            assert_rules=a2,
            scenario_id=SID, step=3, step_type="verify"),
        row(f"{prefix}_008", f"追加第3个{LABEL}公钥（总数3，无数量限制）", EP, 200,
            jd({"requestId": "uuid", "algorithm": ALG, "pks": [K3]}),
            scenario_id=SID, step=4, step_type="test"),
        row(f"{prefix}_009", "查询验证3个指纹", EP, 200,
            params="requestId=uuid", method="GET",
            assert_rules=a3,
            scenario_id=SID, step=5, step_type="verify"),
        row(f"{prefix}_010", f"追加{CROSS_LABEL}跨算法公钥（混合算法）", EP, 200,
            jd({"requestId": "uuid", "algorithm": CROSS_ALG, "pks": [CK1]}),
            scenario_id=SID, step=6, step_type="test"),
        row(f"{prefix}_011", "清空所有公钥", EP, 200,
            jd({"requestId": "uuid"}), method="DELETE",
            scenario_id=SID, step=7, step_type="test",
            ref_case_id="AUTH_PK_017"),
        row(f"{prefix}_012", "清空后查询验证为空", EP, 200,
            params="requestId=uuid", method="GET",
            assert_rules=fp_assert_empty(),
            scenario_id=SID, step=8, step_type="verify",
            ref_case_id="AUTH_PK_018"),
        row(f"{prefix}_013", "清空后调trusted接口(getCHSMInfo)验证返回401",
            EP_CHSM, 401,
            jd({"requestId": "uuid", "oprType": "getinfo"}),
            assert_rules=ERR401,
            scenario_id=SID, step=9, step_type="test",
            ref_case_id="AUTH_PK_019"),
    ]


def gen_pk_batch_scenario(alg, prefix, scn_prefix, keys, fps, hash_alg,
                          cross_alg, cross_keys,
                          cross_ref_case_id):
    """SCN_PK_{RSA|SM2}_BATCH: 11 steps — 批量传入/去重/混合算法/恢复
    7.3 无数量限制: 一次传3个公钥可以成功。
    """
    SID = f"{scn_prefix}_BATCH"
    LABEL = alg.upper()
    ALG = API_ALG[alg]
    K1, K2, K3 = keys
    FP1, FP2, FP3 = fps
    CK1 = cross_keys[0]
    CROSS_ALG = API_ALG[cross_alg]
    CROSS_LABEL = cross_alg.upper()
    a3 = fp_assert_3(hash_alg, FP1, FP2, FP3)
    a1 = fp_assert_1(hash_alg, FP1)

    return [
        row(f"{prefix}_014", f"配置1个{LABEL}公钥", EP, 200,
            jd({"requestId": "uuid", "algorithm": ALG, "pks": [K1]}),
            scenario_id=SID, step=1, step_type="setup"),
        row(f"{prefix}_015", f"追加1个{CROSS_LABEL}公钥（混合算法）", EP, 200,
            jd({"requestId": "uuid", "algorithm": CROSS_ALG, "pks": [CK1]}),
            scenario_id=SID, step=2, step_type="test",
            ref_case_id=cross_ref_case_id),
        row(f"{prefix}_016", "查询验证混合算法指纹", EP, 200,
            params="requestId=uuid", method="GET",
            assert_rules=fp_assert_mixed(),
            scenario_id=SID, step=3, step_type="verify",
            ref_case_id="AUTH_PK_010"),
        row(f"{prefix}_017", "清空所有公钥", EP, 200,
            jd({"requestId": "uuid"}), method="DELETE",
            scenario_id=SID, step=4, step_type="test"),
        row(f"{prefix}_018", f"一次传3个{LABEL}公钥（无数量限制）", EP, 200,
            jd({"requestId": "uuid", "algorithm": ALG, "pks": [K1, K2, K3]}),
            scenario_id=SID, step=5, step_type="test",
            ref_case_id="AUTH_PK_004"),
        row(f"{prefix}_019", "查询验证3个指纹", EP, 200,
            params="requestId=uuid", method="GET",
            assert_rules=a3,
            scenario_id=SID, step=6, step_type="verify"),
        row(f"{prefix}_020",
            "去重测试：传相同3个公钥（幂等成功）", EP, 200,
            jd({"requestId": "uuid", "algorithm": ALG, "pks": [K1, K2, K3]}),
            scenario_id=SID, step=7, step_type="test"),
        row(f"{prefix}_021", "去重后查询仍为3个指纹", EP, 200,
            params="requestId=uuid", method="GET",
            assert_rules=a3,
            scenario_id=SID, step=8, step_type="verify"),
        row(f"{prefix}_022", "清空所有公钥", EP, 200,
            jd({"requestId": "uuid"}), method="DELETE",
            scenario_id=SID, step=9, step_type="test"),
        row(f"{prefix}_023", f"恢复主{LABEL}公钥（为后续测试准备）", EP, 200,
            jd({"requestId": "uuid", "algorithm": ALG, "pks": [K1]}),
            scenario_id=SID, step=10, step_type="setup",
            ref_case_id="AUTH_PK_020"),
        row(f"{prefix}_024", "确认主公钥已就绪", EP, 200,
            params="requestId=uuid", method="GET",
            assert_rules=a1,
            scenario_id=SID, step=11, step_type="verify"),
    ]


def gen_pk_error_standalone(alg, prefix, keys):
    """独立错误用例 E01-E10，无 scenario_id。"""
    ALG = API_ALG[alg]
    LABEL = alg.upper()
    K1 = keys[0]

    return [
        row(f"{prefix}_E01", "algorithm为非法值(aes)", EP, 400,
            jd({"requestId": "uuid", "algorithm": "aes", "pks": [K1]}),
            ref_case_id="AUTH_PK_E01"),
        row(f"{prefix}_E02", "pks为空数组[]", EP, 400,
            jd({"requestId": "uuid", "algorithm": ALG, "pks": []}),
            ref_case_id="AUTH_PK_E02"),
        row(f"{prefix}_E03", "pks中公钥格式错误", EP, 400,
            jd({"requestId": "uuid", "algorithm": ALG, "pks": ["not-valid-key!!!"]}),
            ref_case_id="AUTH_PK_E03"),
        row(f"{prefix}_E04", "缺少algorithm字段", EP, 400,
            jd({"requestId": "uuid", "pks": [K1]}),
            ref_case_id="AUTH_PK_E04"),
        row(f"{prefix}_E05", "缺少pks字段", EP, 400,
            jd({"requestId": "uuid", "algorithm": ALG}),
            ref_case_id="AUTH_PK_E05"),
        row(f"{prefix}_E06", "缺少requestId", EP, 400,
            jd({"algorithm": ALG, "pks": [K1]}),
            ref_case_id="AUTH_PK_E06"),
        row(f"{prefix}_E07", "requestId为空字符串", EP, 400,
            jd({"requestId": "", "algorithm": ALG, "pks": [K1]}),
            ref_case_id="AUTH_PK_E07"),
        row(f"{prefix}_E08", "1个正确+1个错误公钥", EP, 400,
            jd({"requestId": "uuid", "algorithm": ALG, "pks": [K1, "invalid!!!"]})),
        row(f"{prefix}_E09", "clearCHSMPk缺少requestId", EP, 400,
            jd({}), method="DELETE",
            ref_case_id="AUTH_PK_E08"),
        row(f"{prefix}_E10", "getCHSMPk缺少requestId", EP, 400,
            method="GET",
            ref_case_id="AUTH_PK_E09"),
        row(f"{prefix}_E11", "clearCHSMPk requestId空字符串", EP, 400,
            jd({"requestId": ""}), method="DELETE"),
        row(f"{prefix}_E12", "getCHSMPk requestId空字符串", EP, 400,
            params="requestId=", method="GET"),
    ]


def _pk_tail_cleanup():
    return [
        row("PK_CLEAR_H001", "清空所有公钥", EP, 200,
            jd({"requestId": "uuid"}), method="DELETE"),
        row("PK_GET_H002", "清空后查询验证为空", EP, 200,
            params="requestId=uuid", method="GET",
            assert_rules=fp_assert_empty()),
    ]


# ================================================================
# 主流程
# ================================================================

def write_sheet(ws, data_rows):
    for ci, h in enumerate(HEADERS, 1):
        ws.cell(1, ci, h)
    for ri, data in enumerate(data_rows, 2):
        for ci, h in enumerate(HEADERS, 1):
            val = data.get(h)
            if val is not None:
                ws.cell(ri, ci, val)


def main():
    xlsx = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "test_data.xlsx")
    wb = openpyxl.load_workbook(xlsx)

    for name in ["pk_rsa", "pk_sm2"]:
        if name in wb.sheetnames:
            del wb[name]

    # --- pk_rsa ---
    rsa_keys = (RSA_KEY1, RSA_KEY2, RSA_KEY3)
    rsa_fps = (RSA_FP1, RSA_FP2, RSA_FP3)
    sm2_keys = (SM2_KEY1, SM2_KEY2, SM2_KEY3)

    rsa_data = []
    rsa_data.extend(gen_pk_single_scenario(
        "rsa", "PK_RSA", "SCN_PK_RSA", rsa_keys, rsa_fps, RSA_HASH_ALG))
    rsa_data.extend(gen_pk_multi_scenario(
        "rsa", "PK_RSA", "SCN_PK_RSA", rsa_keys, rsa_fps, RSA_HASH_ALG,
        "sm2", sm2_keys))
    rsa_data.extend(gen_pk_batch_scenario(
        "rsa", "PK_RSA", "SCN_PK_RSA", rsa_keys, rsa_fps, RSA_HASH_ALG,
        "sm2", sm2_keys,
        cross_ref_case_id="AUTH_PK_008"))
    rsa_data.extend(gen_pk_error_standalone("rsa", "PK_RSA", rsa_keys))
    rsa_data.extend(_pk_tail_cleanup())

    ws_rsa = wb.create_sheet("pk_rsa")
    write_sheet(ws_rsa, rsa_data)

    # --- pk_sm2 ---
    sm2_fps = (SM2_FP1, SM2_FP2, SM2_FP3)

    sm2_data = []
    sm2_data.extend(gen_pk_single_scenario(
        "sm2", "PK_SM2", "SCN_PK_SM2", sm2_keys, sm2_fps, SM2_HASH_ALG))
    sm2_data.extend(gen_pk_multi_scenario(
        "sm2", "PK_SM2", "SCN_PK_SM2", sm2_keys, sm2_fps, SM2_HASH_ALG,
        "rsa", rsa_keys))
    sm2_data.extend(gen_pk_batch_scenario(
        "sm2", "PK_SM2", "SCN_PK_SM2", sm2_keys, sm2_fps, SM2_HASH_ALG,
        "rsa", rsa_keys,
        cross_ref_case_id="AUTH_PK_009"))
    sm2_data.extend(gen_pk_error_standalone("sm2", "PK_SM2", sm2_keys))
    sm2_data.extend(_pk_tail_cleanup())

    ws_sm2 = wb.create_sheet("pk_sm2")
    write_sheet(ws_sm2, sm2_data)

    wb.save(xlsx)

    for sheet_name, data in [("pk_rsa", rsa_data), ("pk_sm2", sm2_data)]:
        enabled = sum(1 for r in data if r.get("enabled", "yes") == "yes")
        disabled = len(data) - enabled
        print(f"{sheet_name}: {len(data)} 行 ({enabled} enabled, {disabled} disabled)")

        from collections import Counter
        scn_counter = Counter()
        for r in data:
            sid = r.get("scenario_id")
            if sid:
                scn_counter[sid] += 1
        standalone = sum(1 for r in data if not r.get("scenario_id"))
        for sid, cnt in sorted(scn_counter.items()):
            print(f"  {sid}: {cnt} 步")
        if standalone:
            print(f"  独立错误用例: {standalone} 条")

    print(f"\n已保存: {xlsx}")


if __name__ == "__main__":
    main()
