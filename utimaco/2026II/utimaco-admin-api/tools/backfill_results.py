#!/usr/bin/env python3
"""
测试结果回填脚本 — 读取 test_results.json，回填到测试用例 Excel。

用法:
  python tools/backfill_results.py
  python tools/backfill_results.py --results reports/test_results.json --excel ../华为云密码机二期-接口测试用例.xlsx
  python tools/backfill_results.py --dry-run   # 只显示匹配结果，不修改 Excel
"""

import argparse
import json
import os
import shutil
import sys
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("需要 openpyxl: pip install openpyxl")
    sys.exit(1)


ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

DEFAULT_RESULTS = os.path.join(ROOT, "output", "test_results.json")
DEFAULT_EXCEL = os.path.join(ROOT, "..", "华为云密码机二期-接口测试用例.xlsx")

SHEET_NAME = "管理接口"

COL_CASE_ID = "用例ID"
COL_STATUS = "测试状态"
COL_EXECUTOR = "执行人"
COL_EXEC_TIME = "执行时间"
COL_REMARK = "备注"

STATUS_MAP = {
    "passed": "通过",
    "failed": "失败",
    "error": "异常",
}

# 同功能别名: 一条测试结果同时回填多个用例ID
# key=测试数据ref_case_id, value=用例Excel中同功能的别名ID列表
ALIASES = {
    # --- guest接口(同接口,认证/非认证两种用例) ---
    "CHSM_STATUS_001": ["GUEST_STATUS_001"],
    "CHSM_STATUS_T01": ["GUEST_STATUS_T01"],
    "CHSM_STATUS_002": ["GUEST_STATUS_T02"],
    "CHSM_ALLSTATUS_001": ["GUEST_ALLSTATUS_001"],
    "CHSM_ALLSTATUS_T01": ["GUEST_ALLSTATUS_T01"],
    "CHSM_ALLSTATUS_T02": ["GUEST_ALLSTATUS_T02"],
    "VSM_STATUS_001": ["GUEST_VSM_STATUS_001"],
    "VSM_STATUS_002": ["GUEST_VSM_STATUS_002"],
    "VSM_STATUS_T02": ["GUEST_VSM_STATUS_T01"],
    "VSM_STATUS_T03": ["GUEST_VSM_STATUS_T02"],
    "CHSM_INFO_005": ["AUTH_403_001", "GUEST_AUTH_403"],
    "CHSM_NET_T11": ["CHSM_NET_004"],
    "CHSM_NET_003": ["CHSM_NET_006"],

    # --- 7.3 getCHSMInfo验证步骤(场景中调/chsm验证公钥可用) ---
    "CHSM_INFO_001": ["AUTH_PK_003", "AUTH_PK_006", "AUTH_PK_011", "CHSM_INFO_006",
                       "VSM_SETUP_001", "VSM_SETUP_EXPORT", "VSM_SETUP_IMPORT",
                       "VSM_SETUP_INFO", "VSM_SETUP_NET", "VSM_SETUP_RESET",
                       "VSM_SETUP_RESTART", "VSM_SETUP_START", "VSM_SETUP_STOP",
                       "VSM_SETUP_TOKEN", "VSM_SETUP_UPGRADE"],
    "CHSM_INFO_007": ["AUTH_PK_007", "AUTH_PK_012"],

    # --- 7.3 pk_rsa场景步骤 → AUTH_PK_xxx(7.3总用例) ---
    "PK_RSA_001": ["AUTH_PK_001", "AUTH_PK_014"],
    "PK_RSA_005": ["AUTH_PK_016", "PK_RSA_S02_SETUP"],
    "PK_RSA_006": ["AUTH_PK_005"],
    "PK_RSA_008": ["PK_RSA_004A", "AUTH_PK_015"],
    "PK_RSA_009": ["AUTH_PK_016B"],
    "PK_RSA_010": ["AUTH_PK_017"],
    "PK_RSA_011": ["AUTH_PK_018"],
    "PK_RSA_011A": ["AUTH_PK_019"],
    "PK_RSA_013": ["AUTH_PK_008"],
    "PK_RSA_014": ["AUTH_PK_010"],
    "PK_RSA_016": ["AUTH_PK_004", "AUTH_PK_013"],
    "PK_RSA_021": ["AUTH_PK_020"],
    "PK_RSA_004": ["PK_RSA_004B"],

    # --- 7.3 pk_sm2场景步骤 ---
    "PK_SM2_004": ["PK_SM2_004B"],
    "PK_SM2_005": ["PK_SM2_S02_SETUP"],
    "PK_SM2_008": ["PK_SM2_004A"],
    "PK_SM2_013": ["AUTH_PK_009"],

    # --- 9.2 setup/teardown ---
    "TENANT_RESTORE_PK": ["RESTORE_TENANT_PK"],
    "TENANT_LIMIT_S01": ["TENANT_LIMIT_SETUP"],
    "TENANT_CLEANPK_ALL": ["TENANT_CLEANUP_001"],
    "TENANT_AUTHPK_004": ["TENANT_CLEANUP_002"],
    "TENANT_GETPK_006": ["TENANT_CLEANUP_003"],
    "TENANT_STATUS_001": ["TENANT_STATUS_002"],

    # --- SM2跑两遍覆盖(同接口不同签名算法) ---
    "CHSM_NET_001": ["CHSM_NET_005"],
    "CHSM_IMPORT_005": ["CHSM_IMPORT_002"],
    "CHSM_UPGRADE_001": ["CHSM_UPGRADE_002"],
    "CHSM_RESTORE_001": ["CHSM_RESTORE_002"],
    "VSM_INFO_001": ["VSM_INFO_004"],
    "VSM_IMPORT_001": ["VSM_IMPORT_002"],
    "VSM_START_001": ["VSM_START_003"],
    "VSM_STOP_001": ["VSM_STOP_003"],
    "VSM_RESTART_001": ["VSM_RESTART_003"],
    "VSM_RESET_001": ["VSM_RESET_003"],
    "VSM_UPGRADE_001": ["VSM_UPGRADE_002"],
}


def load_results(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    by_ref = {}
    for r in data.get("results", []):
        ref = r.get("ref_case_id")
        if not ref:
            continue
        existing = by_ref.get(ref)
        if existing is None or _status_priority(r["status"]) > _status_priority(existing["status"]):
            by_ref[ref] = r
        for alias in ALIASES.get(ref, []):
            existing = by_ref.get(alias)
            if existing is None or _status_priority(r["status"]) > _status_priority(existing["status"]):
                by_ref[alias] = r
    return by_ref


def _status_priority(status: str) -> int:
    return {"passed": 0, "failed": 2, "error": 1}.get(status, 0)


def find_columns(ws) -> dict:
    """在第 1 行找到目标列的索引"""
    col_map = {}
    for cell in ws[1]:
        val = cell.value
        if val == COL_CASE_ID:
            col_map["case_id"] = cell.column
        elif val == COL_STATUS:
            col_map["status"] = cell.column
        elif val == COL_EXECUTOR:
            col_map["executor"] = cell.column
        elif val == COL_EXEC_TIME:
            col_map["exec_time"] = cell.column
        elif val == COL_REMARK:
            col_map["remark"] = cell.column
    return col_map


def backfill(results_path: str, excel_path: str, dry_run: bool = False):
    if not os.path.isfile(results_path):
        print(f"结果文件不存在: {results_path}")
        sys.exit(1)
    if not os.path.isfile(excel_path):
        print(f"Excel 文件不存在: {excel_path}")
        sys.exit(1)

    by_ref = load_results(results_path)
    if not by_ref:
        print("没有可回填的结果（所有行的 ref_case_id 为空）")
        return

    print(f"加载 {len(by_ref)} 条可回填结果")

    wb = openpyxl.load_workbook(excel_path)
    if SHEET_NAME not in wb.sheetnames:
        print(f"Sheet '{SHEET_NAME}' 不存在，可选: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[SHEET_NAME]
    col_map = find_columns(ws)

    for key in ("case_id", "status", "executor", "exec_time"):
        if key not in col_map:
            print(f"缺少必要列: {key} (期望列名: {COL_CASE_ID}/{COL_STATUS}/{COL_EXECUTOR}/{COL_EXEC_TIME})")
            sys.exit(1)

    matched = 0
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    for row in ws.iter_rows(min_row=2):
        case_id_cell = row[col_map["case_id"] - 1]
        case_id = case_id_cell.value
        if not case_id:
            continue

        result = by_ref.get(str(case_id).strip())
        if not result:
            continue

        matched += 1
        status_cn = STATUS_MAP.get(result["status"], result["status"])

        if dry_run:
            print(f"  [匹配] {case_id} → {status_cn}")
        else:
            row[col_map["status"] - 1].value = status_cn
            row[col_map["executor"] - 1].value = "自动化测试"
            row[col_map["exec_time"] - 1].value = now_str
            if "remark" in col_map:
                if result["status"] in ("failed", "error"):
                    row[col_map["remark"] - 1].value = result.get("error_message", "")
                else:
                    row[col_map["remark"] - 1].value = None

    print(f"\n匹配 {matched}/{len(by_ref)} 条")

    if not dry_run and matched > 0:
        bak_path = excel_path + ".bak"
        if not os.path.exists(bak_path):
            shutil.copy2(excel_path, bak_path)
            print(f"备份: {bak_path}")
        wb.save(excel_path)
        print(f"已保存: {excel_path}")
    elif not dry_run:
        print("无匹配项，未修改 Excel")


def main():
    parser = argparse.ArgumentParser(description="回填测试结果到用例 Excel")
    parser.add_argument("--results", default=DEFAULT_RESULTS, help="test_results.json 路径")
    parser.add_argument("--excel", default=DEFAULT_EXCEL, help="测试用例 Excel 路径")
    parser.add_argument("--dry-run", action="store_true", help="只显示匹配，不修改文件")
    args = parser.parse_args()

    excel_path = os.path.abspath(args.excel)
    backfill(args.results, excel_path, args.dry_run)


if __name__ == "__main__":
    main()
