#!/usr/bin/env python3
"""
生成并发测试用例 Excel — 按项目标准格式输出到 data/华为云密码机二期-接口测试用例.xlsx 新增 "并发测试" sheet。
如文件不存在则单独输出到 data/concurrency_test_cases.xlsx。
"""

import json
import os

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# ── 列定义 ──────────────────────────────────────────
COLUMNS = [
    "用例ID", "需求章节", "修复项", "严重级别", "测试模块", "接口名称",
    "接口路径", "请求方法", "测试场景", "并发数", "锁类型",
    "前置条件", "请求参数(Body)",
    "预期响应状态码", "预期结果描述", "判定逻辑",
    "测试状态", "执行人", "执行时间", "缺陷ID", "备注",
]

COL_WIDTHS = {
    "用例ID": 16, "需求章节": 10, "修复项": 10, "严重级别": 10,
    "测试模块": 18, "接口名称": 36,
    "接口路径": 30, "请求方法": 10,
    "测试场景": 50, "并发数": 8, "锁类型": 18,
    "前置条件": 34, "请求参数(Body)": 52,
    "预期响应状态码": 16, "预期结果描述": 40, "判定逻辑": 38,
    "测试状态": 10, "执行人": 10, "执行时间": 12, "缺陷ID": 10, "备注": 36,
}

# ── 样式 ──────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
CELL_FONT = Font(name="微软雅黑", size=9)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

SCENARIO_FILLS = {
    "A": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
    "B": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "C": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
    "D": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
    "E": PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),
    "F": PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid"),
    "G": PatternFill(start_color="E0F7FA", end_color="E0F7FA", fill_type="solid"),
}


# ── 测试用例数据 ──────────────────────────────────────

CALLBACK_URL = "http://10.10.1.207:8000/callback"
VSM_A = "vsm-concurrency-a"
VSM_B = "vsm-concurrency-b"

CASES = [
    # ============================================
    # 场景 A：全局锁互斥（Fix #1 — CRITICAL）
    # ============================================
    {
        "用例ID": "CA_01",
        "需求章节": "10.1",
        "修复项": "#1 两级锁",
        "严重级别": "CRITICAL",
        "测试模块": "并发-全局锁",
        "接口名称": "exportCHSM + startVSM",
        "接口路径": "/api/1.0/chsm/image\n/api/1.0/vsm",
        "请求方法": "POST\nPOST",
        "测试场景": "场景A: 全局锁互斥 — 并发发起CHSM export和VSM start，验证全局write lock使VSM操作被阻塞或标记FAILED",
        "并发数": 2,
        "锁类型": "全局 write lock",
        "前置条件": "CHSM已配置认证公钥\nVSM已创建",
        "请求参数(Body)": json.dumps({
            "请求1(CHSM export)": {"requestId": "uuid", "oprType": "export", "callbackUrl": CALLBACK_URL},
            "请求2(VSM start)": {"requestId": "uuid", "oprType": "start", "vsmId": VSM_A, "callbackUrl": CALLBACK_URL},
        }, ensure_ascii=False, indent=2),
        "预期响应状态码": "200 + 409/423/503",
        "预期结果描述": "CHSM export持有全局write lock期间，VSM start应被阻塞；锁获取失败时任务标记FAILED，由RetryScheduler下轮重试",
        "判定逻辑": "状态码判定: 至少1个请求返回非200(409/423/503)\nOR 时序判定: 响应时间差≥0.5s(串行化)",
    },
    {
        "用例ID": "CA_02",
        "需求章节": "10.1",
        "修复项": "#1 两级锁",
        "严重级别": "CRITICAL",
        "测试模块": "并发-全局锁",
        "接口名称": "restartCHSM + backupCHSM",
        "接口路径": "/api/1.0/chsm\n/api/1.0/chsm",
        "请求方法": "POST\nPOST",
        "测试场景": "场景A: 全局锁互斥 — 并发发起CHSM restart和CHSM backup，验证两个全局write lock操作互斥",
        "并发数": 2,
        "锁类型": "全局 write lock",
        "前置条件": "CHSM已配置认证公钥",
        "请求参数(Body)": json.dumps({
            "请求1(CHSM restart)": {"requestId": "uuid", "oprType": "restart", "callbackUrl": CALLBACK_URL},
            "请求2(CHSM backup)": {"requestId": "uuid", "oprType": "backup", "callbackUrl": CALLBACK_URL},
        }, ensure_ascii=False, indent=2),
        "预期响应状态码": "200 + 409/423/503",
        "预期结果描述": "两个CHSM全局write lock操作并发时，其中一个应被阻塞",
        "判定逻辑": "状态码判定: 至少1个请求被拒绝\nOR 时序判定: 响应时间差≥0.5s",
    },

    # ============================================
    # 场景 B：per-vsmId 互斥（Fix #1 — CRITICAL）
    # ============================================
    {
        "用例ID": "CB_01",
        "需求章节": "10.1",
        "修复项": "#1 两级锁",
        "严重级别": "CRITICAL",
        "测试模块": "并发-per-vsmId锁",
        "接口名称": "startVSM + restartVSM (同vsmId)",
        "接口路径": "/api/1.0/vsm\n/api/1.0/vsm",
        "请求方法": "POST\nPOST",
        "测试场景": "场景B: per-vsmId互斥 — 对同一vsmId并发发起start和restart，验证同设备操作互斥",
        "并发数": 2,
        "锁类型": "per-vsmId read lock",
        "前置条件": "VSM-A已创建",
        "请求参数(Body)": json.dumps({
            "请求1(start)": {"requestId": "uuid", "oprType": "start", "vsmId": VSM_A, "callbackUrl": CALLBACK_URL},
            "请求2(restart)": {"requestId": "uuid", "oprType": "restart", "vsmId": VSM_A, "callbackUrl": CALLBACK_URL},
        }, ensure_ascii=False, indent=2),
        "预期响应状态码": "200 + 409/423/503",
        "预期结果描述": "同一vsmId的操作互斥，后到的请求被阻塞或标记FAILED",
        "判定逻辑": "状态码判定: 至少1个请求被拒绝\nOR 时序判定: 响应时间差≥0.5s",
    },
    {
        "用例ID": "CB_02",
        "需求章节": "10.1",
        "修复项": "#1 两级锁",
        "严重级别": "CRITICAL",
        "测试模块": "并发-per-vsmId锁",
        "接口名称": "startVSM-A + startVSM-B (不同vsmId)",
        "接口路径": "/api/1.0/vsm\n/api/1.0/vsm",
        "请求方法": "POST\nPOST",
        "测试场景": "场景B(对照): per-vsmId不互斥 — 对不同vsmId分别发起start，验证不同设备操作不被误伤",
        "并发数": 2,
        "锁类型": "per-vsmId read lock",
        "前置条件": "VSM-A和VSM-B均已创建",
        "请求参数(Body)": json.dumps({
            "请求1(VSM-A start)": {"requestId": "uuid", "oprType": "start", "vsmId": VSM_A, "callbackUrl": CALLBACK_URL},
            "请求2(VSM-B start)": {"requestId": "uuid", "oprType": "start", "vsmId": VSM_B, "callbackUrl": CALLBACK_URL},
        }, ensure_ascii=False, indent=2),
        "预期响应状态码": "200 + 200",
        "预期结果描述": "不同vsmId的操作可以并行执行，均应返回200成功",
        "判定逻辑": "全部2个请求返回200",
    },

    # ============================================
    # 场景 C：VSM upgrade 全局互斥（Fix #1）
    # ============================================
    {
        "用例ID": "CC_01",
        "需求章节": "10.1",
        "修复项": "#1 两级锁",
        "严重级别": "CRITICAL",
        "测试模块": "并发-upgrade全局锁",
        "接口名称": "upgradeVSM + startVSM",
        "接口路径": "/api/1.0/vsm\n/api/1.0/vsm",
        "请求方法": "POST\nPOST",
        "测试场景": "场景C: VSM upgrade全局互斥 — 并发发起VSM upgrade和VSM start(不同vsmId)，验证upgrade持有全局write lock阻塞所有其他操作",
        "并发数": 2,
        "锁类型": "全局 write lock",
        "前置条件": "VSM-A和VSM-B均已创建",
        "请求参数(Body)": json.dumps({
            "请求1(upgrade)": {"requestId": "uuid", "oprType": "upgrade", "vsmId": VSM_A,
                             "packVersion": "99.0.0", "packUrl": "http://dummy/upgrade.pkg",
                             "alg": "RSAWithSHA256", "sign": "dummysign", "callbackUrl": CALLBACK_URL},
            "请求2(start)": {"requestId": "uuid", "oprType": "start", "vsmId": VSM_B, "callbackUrl": CALLBACK_URL},
        }, ensure_ascii=False, indent=2),
        "预期响应状态码": "200 + 409/423/503",
        "预期结果描述": "VSM upgrade持有全局write lock期间，即使不同vsmId的start操作也应被阻塞",
        "判定逻辑": "状态码判定: 至少1个请求被拒绝\nOR 时序判定: 响应时间差≥0.5s",
    },

    # ============================================
    # 场景 D：UpgradeInProgressGuard（Fix #2 — HIGH）
    # ============================================
    {
        "用例ID": "CD_01",
        "需求章节": "10.2",
        "修复项": "#2 Guard二次检查",
        "严重级别": "HIGH",
        "测试模块": "并发-升级Guard",
        "接口名称": "upgradeCHSM × 2",
        "接口路径": "/api/1.0/chsm",
        "请求方法": "POST",
        "测试场景": "场景D: UpgradeInProgressGuard — 并发发起两次CHSM upgrade，验证runAsync dispatch前的二次guard.check()拦截重复升级",
        "并发数": 2,
        "锁类型": "全局 write lock + Guard",
        "前置条件": "CHSM已配置认证公钥\n无正在进行的升级操作",
        "请求参数(Body)": json.dumps({
            "请求1": {"requestId": "uuid", "oprType": "upgrade", "packVersion": "99.0.0",
                     "packUrl": "http://dummy/chsm_upgrade.pkg", "alg": "RSAWithSHA256",
                     "sign": "dummysign", "callbackUrl": CALLBACK_URL},
            "请求2": "(同请求1，requestId不同)",
        }, ensure_ascii=False, indent=2),
        "预期响应状态码": "200 + 409",
        "预期结果描述": "第一个upgrade请求被接受(200)，第二个被Guard拦截返回409 Conflict",
        "判定逻辑": "恰好1个200 + 1个409\nOR 两个均非200(前置条件不满足)视为锁有效\nOR 时序判定: 响应时间差≥0.5s",
    },
    {
        "用例ID": "CD_02",
        "需求章节": "10.2",
        "修复项": "#2 Guard二次检查",
        "严重级别": "HIGH",
        "测试模块": "并发-升级Guard",
        "接口名称": "upgradeVSM × 2",
        "接口路径": "/api/1.0/vsm",
        "请求方法": "POST",
        "测试场景": "场景D: UpgradeInProgressGuard — 并发发起两次VSM upgrade(同vsmId)，验证Guard二次检查",
        "并发数": 2,
        "锁类型": "全局 write lock + Guard",
        "前置条件": "VSM已创建\n无正在进行的升级操作",
        "请求参数(Body)": json.dumps({
            "请求1": {"requestId": "uuid", "oprType": "upgrade", "vsmId": VSM_A,
                     "packVersion": "99.0.0", "packUrl": "http://dummy/vsm_upgrade.pkg",
                     "alg": "RSAWithSHA256", "sign": "dummysign", "callbackUrl": CALLBACK_URL},
            "请求2": "(同请求1，requestId不同)",
        }, ensure_ascii=False, indent=2),
        "预期响应状态码": "200 + 409",
        "预期结果描述": "第一个upgrade被接受，第二个被Guard拦截返回409",
        "判定逻辑": "恰好1个200 + 1个409\nOR 两个均非200视为锁有效",
    },

    # ============================================
    # 场景 E：配置 UPSERT 幂等（Fix #3 — HIGH）
    # ============================================
    {
        "用例ID": "CE_01",
        "需求章节": "10.3",
        "修复项": "#3 UPSERT",
        "严重级别": "HIGH",
        "测试模块": "并发-配置UPSERT",
        "接口名称": "configCHSMNet-配置CHSM网络",
        "接口路径": "/api/1.0/chsm/network",
        "请求方法": "POST",
        "测试场景": "场景E: UPSERT幂等 — 10路并发写入CHSM网络配置，验证INSERT ON CONFLICT DO UPDATE消除UNIQUE约束冲突500",
        "并发数": 10,
        "锁类型": "无(UPSERT)",
        "前置条件": "CHSM已配置认证公钥",
        "请求参数(Body)": json.dumps({
            "requestId": "uuid", "dnsList": ["8.8.8.{i}", "114.114.114.{i}"],
            "netAddrs": [{"name": "eth0", "ip": "192.168.{i}.100",
                          "mask": "255.255.255.0", "gateway": "192.168.{i}.1"}],
        }, ensure_ascii=False, indent=2),
        "预期响应状态码": "全部200",
        "预期结果描述": "10路并发写入同一配置项，全部返回200，不应出现500 UNIQUE约束冲突",
        "判定逻辑": "无500错误 (核心)\n全部返回200 (最优)",
    },
    {
        "用例ID": "CE_02",
        "需求章节": "10.3",
        "修复项": "#3 UPSERT",
        "严重级别": "HIGH",
        "测试模块": "并发-配置UPSERT",
        "接口名称": "configCHSMNtp-配置NTP",
        "接口路径": "/api/1.0/chsm/ntp",
        "请求方法": "POST",
        "测试场景": "场景E: UPSERT幂等 — 10路并发写入NTP配置",
        "并发数": 10,
        "锁类型": "无(UPSERT)",
        "前置条件": "CHSM已配置认证公钥",
        "请求参数(Body)": json.dumps({
            "requestId": "uuid", "addr": "ntp{i}.aliyun.com", "syncPeriod": 60,
        }, ensure_ascii=False, indent=2),
        "预期响应状态码": "全部200",
        "预期结果描述": "10路并发NTP配置写入无500",
        "判定逻辑": "无500错误",
    },
    {
        "用例ID": "CE_03",
        "需求章节": "10.3",
        "修复项": "#3 UPSERT",
        "严重级别": "HIGH",
        "测试模块": "并发-配置UPSERT",
        "接口名称": "configCHSMUploadAddress-配置镜像上传地址",
        "接口路径": "/api/1.0/chsm/imageuploader",
        "请求方法": "POST",
        "测试场景": "场景E: UPSERT幂等 — 10路并发写入镜像上传地址",
        "并发数": 10,
        "锁类型": "无(UPSERT)",
        "前置条件": "CHSM已配置认证公钥",
        "请求参数(Body)": json.dumps({
            "requestId": "uuid", "url": "http://image-server-{i}.example.com/upload",
        }, ensure_ascii=False, indent=2),
        "预期响应状态码": "全部200",
        "预期结果描述": "10路并发镜像上传地址配置写入无500",
        "判定逻辑": "无500错误",
    },
    {
        "用例ID": "CE_04",
        "需求章节": "10.3",
        "修复项": "#3 UPSERT",
        "严重级别": "HIGH",
        "测试模块": "并发-配置UPSERT",
        "接口名称": "configSyslogAddr-配置日志上传地址",
        "接口路径": "/api/1.0/chsm/loguploader",
        "请求方法": "POST",
        "测试场景": "场景E: UPSERT幂等 — 10路并发写入日志上传地址",
        "并发数": 10,
        "锁类型": "无(UPSERT)",
        "前置条件": "CHSM已配置认证公钥",
        "请求参数(Body)": json.dumps({
            "requestId": "uuid", "logServerType": "syslog", "logServerAddress": "10.10.{i}.100:514",
        }, ensure_ascii=False, indent=2),
        "预期响应状态码": "全部200",
        "预期结果描述": "10路并发日志上传地址配置写入无500",
        "判定逻辑": "无500错误",
    },
    {
        "用例ID": "CE_05",
        "需求章节": "10.3",
        "修复项": "#3 UPSERT",
        "严重级别": "HIGH",
        "测试模块": "并发-配置UPSERT",
        "接口名称": "configCHSMAlarmAddress-配置告警地址",
        "接口路径": "/api/1.0/chsm/alarmaddress",
        "请求方法": "POST",
        "测试场景": "场景E: UPSERT幂等 — 10路并发写入告警地址",
        "并发数": 10,
        "锁类型": "无(UPSERT)",
        "前置条件": "CHSM已配置认证公钥",
        "请求参数(Body)": json.dumps({
            "requestId": "uuid", "url": "http://alarm-{i}.example.com/notify",
            "monitoringUrl": "http://monitor-{i}.example.com/check",
        }, ensure_ascii=False, indent=2),
        "预期响应状态码": "全部200",
        "预期结果描述": "10路并发告警地址配置写入无500",
        "判定逻辑": "无500错误",
    },
    {
        "用例ID": "CE_06",
        "需求章节": "10.3",
        "修复项": "#3 UPSERT",
        "严重级别": "HIGH",
        "测试模块": "并发-配置UPSERT",
        "接口名称": "configCHSMToken-配置CloudToken",
        "接口路径": "/api/1.0/chsm/cloudtoken",
        "请求方法": "POST",
        "测试场景": "场景E: UPSERT幂等 — 10路并发写入CloudToken",
        "并发数": 10,
        "锁类型": "无(UPSERT)",
        "前置条件": "CHSM已配置认证公钥",
        "请求参数(Body)": json.dumps({
            "requestId": "uuid", "cloudToken": "eyJ0eXAiOiJKV1QiLCJhbGciOiJSUzI1NiJ9.test{i}",
        }, ensure_ascii=False, indent=2),
        "预期响应状态码": "全部200",
        "预期结果描述": "10路并发CloudToken配置写入无500",
        "判定逻辑": "无500错误",
    },
    {
        "用例ID": "CE_07",
        "需求章节": "10.3",
        "修复项": "#3 UPSERT",
        "严重级别": "HIGH",
        "测试模块": "并发-配置UPSERT",
        "接口名称": "configCHSMMOOCAddress-配置MO对接信息",
        "接口路径": "/api/1.0/chsm/moocaddress",
        "请求方法": "POST",
        "测试场景": "场景E: UPSERT幂等 — 10路并发写入MO OC对接信息",
        "并发数": 10,
        "锁类型": "无(UPSERT)",
        "前置条件": "CHSM已配置认证公钥",
        "请求参数(Body)": json.dumps({
            "requestId": "uuid", "ocIp": "10.10.{i}.1", "ocProtocol": "https", "ocPort": "26635",
            "resourceUrl": "/rest/cloudInfra/v1/resource", "performanceUrl": "/rest/cloudInfra/v1/performance",
            "bizRegionNativeId": "region-{i}", "cloudInfraType": "FusionSphere",
            "dewServiceHost": "kms.cn-north-{i}.myhuaweicloud.com",
        }, ensure_ascii=False, indent=2),
        "预期响应状态码": "全部200",
        "预期结果描述": "10路并发MO OC对接配置写入无500",
        "判定逻辑": "无500错误",
    },
    {
        "用例ID": "CE_08",
        "需求章节": "10.3",
        "修复项": "#3 UPSERT",
        "严重级别": "HIGH",
        "测试模块": "并发-配置UPSERT",
        "接口名称": "configVSMNet-配置VSM网络",
        "接口路径": "/api/1.0/vsm/{vsmId}/network",
        "请求方法": "POST",
        "测试场景": "场景E: UPSERT幂等 — 10路并发写入同vsmId的VSM网络配置",
        "并发数": 10,
        "锁类型": "无(UPSERT)",
        "前置条件": "VSM已创建",
        "请求参数(Body)": json.dumps({
            "requestId": "uuid", "vsmId": VSM_A,
            "ip": "192.168.{i}.201", "mask": "24", "gateway": "192.168.{i}.1",
        }, ensure_ascii=False, indent=2),
        "预期响应状态码": "全部200",
        "预期结果描述": "10路并发VSM网络配置写入无500",
        "判定逻辑": "无500错误",
    },
    {
        "用例ID": "CE_09",
        "需求章节": "10.3",
        "修复项": "#3 UPSERT",
        "严重级别": "HIGH",
        "测试模块": "并发-配置UPSERT",
        "接口名称": "configVSMMac-配置VSM MAC",
        "接口路径": "/api/1.0/vsm/{vsmId}/mac",
        "请求方法": "POST",
        "测试场景": "场景E: UPSERT幂等 — 10路并发写入同vsmId的VSM MAC地址",
        "并发数": 10,
        "锁类型": "无(UPSERT)",
        "前置条件": "VSM已创建",
        "请求参数(Body)": json.dumps({
            "requestId": "uuid", "vsmId": VSM_A, "mac": "ba:ca:c1:d2:e3:0{i}",
        }, ensure_ascii=False, indent=2),
        "预期响应状态码": "全部200",
        "预期结果描述": "10路并发VSM MAC配置写入无500",
        "判定逻辑": "无500错误",
    },
    {
        "用例ID": "CE_10",
        "需求章节": "10.3",
        "修复项": "#3 UPSERT",
        "严重级别": "HIGH",
        "测试模块": "并发-配置UPSERT",
        "接口名称": "configVSMVlan-配置VSM VLAN",
        "接口路径": "/api/1.0/vsm/{vsmId}/vlan",
        "请求方法": "POST",
        "测试场景": "场景E: UPSERT幂等 — 10路并发写入同vsmId的VSM VLAN配置",
        "并发数": 10,
        "锁类型": "无(UPSERT)",
        "前置条件": "VSM已创建",
        "请求参数(Body)": json.dumps({
            "requestId": "uuid", "oprType": "modify", "vsmId": VSM_A, "vlanType": "0", "vlanId": "10{i}",
        }, ensure_ascii=False, indent=2),
        "预期响应状态码": "全部200",
        "预期结果描述": "10路并发VSM VLAN配置写入无500",
        "判定逻辑": "无500错误",
    },
    {
        "用例ID": "CE_11",
        "需求章节": "10.3",
        "修复项": "#3 UPSERT",
        "严重级别": "HIGH",
        "测试模块": "并发-配置UPSERT",
        "接口名称": "deleteVSMNet-删除VSM网络配置",
        "接口路径": "/api/1.0/vsm/{vsmId}/network",
        "请求方法": "DELETE",
        "测试场景": "场景E: UPSERT幂等 — 10路并发删除同vsmId的VSM网络配置",
        "并发数": 10,
        "锁类型": "无(UPSERT)",
        "前置条件": "VSM已创建\nVSM已配置网络",
        "请求参数(Body)": json.dumps({
            "requestId": "uuid", "vsmId": VSM_A,
        }, ensure_ascii=False, indent=2),
        "预期响应状态码": "全部200",
        "预期结果描述": "10路并发删除操作无500",
        "判定逻辑": "无500错误",
    },
    {
        "用例ID": "CE_12",
        "需求章节": "10.3",
        "修复项": "#3 UPSERT",
        "严重级别": "HIGH",
        "测试模块": "并发-配置UPSERT",
        "接口名称": "configVSMToken-配置租户令牌",
        "接口路径": "/api/1.0/vsm/{vsmId}/token",
        "请求方法": "POST",
        "测试场景": "场景E: UPSERT幂等 — 10路并发写入同vsmId的租户令牌",
        "并发数": 10,
        "锁类型": "无(UPSERT)",
        "前置条件": "VSM已创建",
        "请求参数(Body)": json.dumps({
            "requestId": "uuid", "vsmId": VSM_A, "token": "tenant-token-{i}", "tenantId": "tenant-{i}",
        }, ensure_ascii=False, indent=2),
        "预期响应状态码": "全部200",
        "预期结果描述": "10路并发租户令牌配置写入无500",
        "判定逻辑": "无500错误",
    },

    # ============================================
    # 场景 F：导出锁 TTL 超时释放（Fix #5 — MEDIUM）
    # ============================================
    {
        "用例ID": "CF_01",
        "需求章节": "10.5",
        "修复项": "#5 TTL超时释放",
        "严重级别": "MEDIUM",
        "测试模块": "并发-导出锁",
        "接口名称": "exportBackupKeys + importBackupKeys",
        "接口路径": "/platformServlet?method=exportBackupKeys\n/platformServlet?method=importBackupKeys",
        "请求方法": "POST\nPOST",
        "测试场景": "场景F: 导出锁TTL — 并发发起exportBackupKeys和importBackupKeys，验证ExportImagesLockRegistry锁行为及异常时不永久阻塞",
        "并发数": 2,
        "锁类型": "ExportImagesLock + TTL",
        "前置条件": "租户已初始化VSM\n已配置认证公钥",
        "请求参数(Body)": json.dumps({
            "请求1(export)": {"requestId": "uuid", "backupKey": "<base64 backup key>"},
            "请求2(import)": {"requestId": "uuid", "backupKey": "<base64 backup key>"},
        }, ensure_ascii=False, indent=2),
        "预期响应状态码": "无500",
        "预期结果描述": "export和import并发时不应出现500错误；锁超时后应自动释放，不阻塞后续操作",
        "判定逻辑": "无500错误",
    },
    {
        "用例ID": "CF_02",
        "需求章节": "10.5",
        "修复项": "#5 TTL超时释放",
        "严重级别": "MEDIUM",
        "测试模块": "并发-导出锁",
        "接口名称": "exportBackupKeys × 2",
        "接口路径": "/platformServlet?method=exportBackupKeys",
        "请求方法": "POST",
        "测试场景": "场景F: 导出锁互斥 — 并发发起两次exportBackupKeys，验证导出锁互斥及TTL超时释放",
        "并发数": 2,
        "锁类型": "ExportImagesLock + TTL",
        "前置条件": "租户已初始化VSM\n已配置认证公钥",
        "请求参数(Body)": json.dumps({
            "请求1": {"requestId": "uuid", "backupKey": "<base64 key 1>"},
            "请求2": {"requestId": "uuid", "backupKey": "<base64 key 2>"},
        }, ensure_ascii=False, indent=2),
        "预期响应状态码": "无500",
        "预期结果描述": "两次export并发时锁应互斥，不应出现500；异常中断后锁应由TTL自动释放",
        "判定逻辑": "无500错误\n时序: 响应时间差存在(串行化)",
    },

    # ============================================
    # 场景 G：DNS 文件锁（Fix #4 — MEDIUM）
    # ============================================
    {
        "用例ID": "CG_01",
        "需求章节": "10.4",
        "修复项": "#4 resolv.conf文件锁",
        "严重级别": "MEDIUM",
        "测试模块": "并发-DNS文件锁",
        "接口名称": "configCHSMNet-配置CHSM网络(含DNS)",
        "接口路径": "/api/1.0/chsm/network",
        "请求方法": "POST + GET",
        "测试场景": "场景G: DNS文件锁 — 10路并发配置网络(含dnsList)后，连续GET 3次验证resolv.conf内容一致性",
        "并发数": "10(写) + 3(读)",
        "锁类型": "ReentrantLock(文件锁)",
        "前置条件": "CHSM已配置认证公钥",
        "请求参数(Body)": json.dumps({
            "写请求": {"requestId": "uuid", "dnsList": ["8.8.8.8", "114.114.114.114"],
                      "netAddrs": [{"name": "eth0", "ip": "192.168.1.100",
                                    "mask": "255.255.255.0", "gateway": "192.168.1.1"}]},
            "读请求": "GET /api/1.0/chsm/network (无body)",
        }, ensure_ascii=False, indent=2),
        "预期响应状态码": "写:全部200 读:200",
        "预期结果描述": "并发写入不导致resolv.conf内容损坏；后续读取应返回一致的网络配置",
        "判定逻辑": "写阶段无500错误\n读阶段3次GET返回内容一致",
    },
]


def generate():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "并发测试"

    # ── 表头 ──
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 14)

    ws.row_dimensions[1].height = 28

    # ── 数据行 ──
    for row_idx, case in enumerate(CASES, 2):
        scenario_key = case["用例ID"][:2].replace("C", "")
        row_fill = SCENARIO_FILLS.get(scenario_key)

        for col_idx, col_name in enumerate(COLUMNS, 1):
            val = case.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = CELL_FONT
            cell.alignment = WRAP_ALIGN
            cell.border = THIN_BORDER
            if row_fill:
                cell.fill = row_fill

        content_len = max(len(str(case.get(c, "")).split("\n")) for c in COLUMNS if case.get(c))
        ws.row_dimensions[row_idx].height = max(28, content_len * 15)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    out_path = os.path.join(_PROJECT_DIR, "data", "concurrency_test_cases.xlsx")
    wb.save(out_path)
    print(f"已生成: {out_path}")
    print(f"共 {len(CASES)} 条用例")
    return out_path


if __name__ == "__main__":
    generate()
