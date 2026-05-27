#!/usr/bin/env python3
"""生成华为云密码机二期接口测试用例 Excel"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import json

HEADERS = [
    '用例ID', '测试模块', '接口名称', '接口路径', '请求方法',
    '测试场景', '测试类型', '优先级', '前置条件',
    '请求头(Headers)', '请求参数(Query/Body)/步骤',
    '预期响应状态码', '预期响应', '预期结果描述',
    '测试状态', '执行人', '执行时间', '缺陷ID', '备注'
]

MOCK_CALLBACK = "http://10.10.1.207:8000/callback"
API_PREFIX = "/api/1.0"


def std_response(status=200, msg="success", result_desc=""):
    r = {
        "status": status,
        "message": msg,
        "timestamp": "2026-xx-xxTxx:xx:xx.xxx+0800",
        "requestId": "<echo-back>",
        "costMillis": "<int>"
    }
    if result_desc:
        r["result"] = result_desc
    return json.dumps(r, ensure_ascii=False, indent=2)


def auth_header(alg="SHA256withRSA"):
    return (
        f'{{\n'
        f'  "CHSM-AuthPK": "<公钥指纹>",\n'
        f'  "CHSM-SignatureAlg": "{alg}",\n'
        f'  "CHSM-Signature": "<签名值>"\n'
        f'}}'
    )


# ============================================================
# 接口定义：每个接口的元数据和用例模板
# ============================================================

def gen_chsm_info_cases():
    """7.1.1 getCHSMInfo"""
    cases = []
    base = {
        "module": "CHSM配置管理",
        "name": "getCHSMInfo-获取CHSM详细信息",
        "path": f"{API_PREFIX}/chsm",
        "method": "POST",
        "auth": "trusted",
    }
    # P0-正常获取
    cases.append({
        **base, "id": "CHSM_INFO_001",
        "scenario": "正常请求获取CHSM详细信息，oprType=getinfo",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({"requestId": "uuid", "oprType": "getinfo"}, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success", "{id, version, ip, ntpAddr, ...}"),
        "expected_desc": "1. 返回status=200\n2. result中包含id/version/ip等CHSM详细信息",
    })
    # P1-oprType为空
    cases.append({
        **base, "id": "CHSM_INFO_002",
        "scenario": "oprType参数为空字符串",
        "type": "异常测试", "priority": "P1",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({"requestId": "uuid", "oprType": ""}, ensure_ascii=False),
        "expected_status": 400,
        "expected_response": std_response(400, "参数错误"),
        "expected_desc": "返回400，提示oprType参数错误",
    })
    # P1-oprType非法值
    cases.append({
        **base, "id": "CHSM_INFO_003",
        "scenario": "oprType为非法值（如invalid）",
        "type": "异常测试", "priority": "P1",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({"requestId": "uuid", "oprType": "invalid"}, ensure_ascii=False),
        "expected_status": 400,
        "expected_response": std_response(400, "不支持的操作类型"),
        "expected_desc": "返回400，提示不支持的操作类型",
    })
    # P1-缺少requestId
    cases.append({
        **base, "id": "CHSM_INFO_004",
        "scenario": "缺少requestId参数",
        "type": "异常测试", "priority": "P1",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({"oprType": "getinfo"}, ensure_ascii=False),
        "expected_status": 400,
        "expected_response": std_response(400, "缺少requestId"),
        "expected_desc": "返回400，提示缺少必要参数requestId",
    })
    # P1-无认证头
    cases.append({
        **base, "id": "CHSM_INFO_005",
        "scenario": "不携带认证头信息（trusted接口无签名）",
        "type": "异常测试", "priority": "P0",
        "precondition": "CHSM已正常运行，已配置公钥",
        "headers": "无认证头",
        "params": json.dumps({"requestId": "uuid", "oprType": "getinfo"}, ensure_ascii=False),
        "expected_status": 403,
        "expected_response": std_response(403, "认证失败"),
        "expected_desc": "返回403，认证失败",
    })
    # P1-SM2WithSM3签名
    cases.append({
        **base, "id": "CHSM_INFO_006",
        "scenario": "使用SM2WithSM3签名算法获取CHSM信息",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已正常运行，已配置SM2公钥",
        "headers": auth_header("SM2WithSM3"),
        "params": json.dumps({"requestId": "uuid", "oprType": "getinfo"}, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success", "{id, version, ip, ...}"),
        "expected_desc": "1. SM2WithSM3签名认证通过\n2. 返回CHSM详细信息",
    })
    # P2-requestId格式错误
    cases.append({
        **base, "id": "CHSM_INFO_007",
        "scenario": "requestId为非UUID格式（如abc123）",
        "type": "边界测试", "priority": "P2",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({"requestId": "abc123", "oprType": "getinfo"}, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": "",
        "expected_desc": "验证requestId格式校验，可能返回200或400",
        "remark": "待确认requestId是否强制UUID格式",
    })
    return cases


def gen_chsm_status_cases():
    """7.1.2 getCHSMStatus"""
    cases = []
    base = {
        "module": "CHSM配置管理",
        "name": "getCHSMStatus-获取CHSM运行状态",
        "path": f"{API_PREFIX}/chsm/status",
        "method": "GET",
        "auth": "guest",
    }
    cases.append({
        **base, "id": "CHSM_STATUS_001",
        "scenario": "正常获取CHSM运行状态",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已正常运行",
        "headers": "无需认证头（guest接口）",
        "params": "requestId=uuid",
        "expected_status": 200,
        "expected_response": std_response(200, "success", '{"status": "ok"}'),
        "expected_desc": "1. 返回200\n2. result.status为ok或fail",
    })
    cases.append({
        **base, "id": "CHSM_STATUS_002",
        "scenario": "不传requestId参数",
        "type": "异常测试", "priority": "P1",
        "precondition": "CHSM已正常运行",
        "headers": "无需认证头",
        "params": "",
        "expected_status": 400,
        "expected_response": std_response(400, "缺少requestId"),
        "expected_desc": "返回400，缺少必要参数",
    })
    return cases


def gen_chsm_allstatus_cases():
    """7.1.3 getCHSMAllStatus"""
    cases = []
    base = {
        "module": "CHSM配置管理",
        "name": "getCHSMAllStatus-获取CHSM所有状态",
        "path": f"{API_PREFIX}/chsm/allstatus",
        "method": "GET",
        "auth": "guest",
    }
    cases.append({
        **base, "id": "CHSM_ALLSTATUS_001",
        "scenario": "正常获取CHSM及所有VSM运行状态",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已正常运行，至少有1个VSM",
        "headers": "无需认证头（guest接口）",
        "params": "requestId=uuid",
        "expected_status": 200,
        "expected_response": std_response(200, "success", '{"hsmStatus":"ok","vsmStatusMap":{...}}'),
        "expected_desc": "1. 返回200\n2. 包含hsmStatus和vsmStatusMap\n3. vsmStatusMap中包含所有VSM状态",
    })
    cases.append({
        **base, "id": "CHSM_ALLSTATUS_002",
        "scenario": "CHSM下无VSM时获取所有状态",
        "type": "功能测试", "priority": "P1",
        "precondition": "CHSM已运行，无VSM",
        "headers": "无需认证头",
        "params": "requestId=uuid",
        "expected_status": 200,
        "expected_response": std_response(200, "success", '{"hsmStatus":"ok","vsmStatusMap":{}}'),
        "expected_desc": "返回200，vsmStatusMap为空对象",
    })
    return cases


def gen_config_chsm_net_cases():
    """7.1.4 configCHSMNet"""
    cases = []
    base = {
        "module": "CHSM配置管理",
        "name": "configCHSMNet-配置CHSM网络",
        "path": f"{API_PREFIX}/chsm/network",
        "method": "POST",
        "auth": "trusted",
    }
    cases.append({
        **base, "id": "CHSM_NET_001",
        "scenario": "正常配置CHSM的DNS列表",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid",
            "dnsList": ["8.8.8.8", "114.114.114.114"]
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "DNS配置成功，返回200",
    })
    cases.append({
        **base, "id": "CHSM_NET_002",
        "scenario": "配置网口信息（含name/ip/mask/gateway）",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid",
            "netAddrs": [{"name": "eth0", "ip": "192.168.1.100", "mask": "255.255.255.0", "gateway": "192.168.1.1"}]
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "网口配置成功，返回200",
    })
    cases.append({
        **base, "id": "CHSM_NET_003",
        "scenario": "配置与当前一致的网络信息",
        "type": "功能测试", "priority": "P1",
        "precondition": "CHSM已配置网络",
        "headers": auth_header(),
        "params": json.dumps({"requestId": "uuid", "dnsList": ["<当前已配置的DNS>"]}, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "配置与当前一致时直接返回成功",
    })
    cases.append({
        **base, "id": "CHSM_NET_004",
        "scenario": "配置无效的IP地址（如999.999.999.999）",
        "type": "异常测试", "priority": "P1",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid",
            "netAddrs": [{"name": "eth0", "ip": "999.999.999.999", "mask": "255.255.255.0", "gateway": "192.168.1.1"}]
        }, ensure_ascii=False),
        "expected_status": 400,
        "expected_response": std_response(400, "IP地址格式错误"),
        "expected_desc": "返回400，IP格式校验失败",
    })
    cases.append({
        **base, "id": "CHSM_NET_005",
        "scenario": "使用SM2WithSM3签名配置网络",
        "type": "功能测试", "priority": "P0",
        "precondition": "已配置SM2公钥",
        "headers": auth_header("SM2WithSM3"),
        "params": json.dumps({"requestId": "uuid", "dnsList": ["8.8.8.8"]}, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "SM2签名认证通过，配置成功",
    })
    cases.append({
        **base, "id": "CHSM_NET_006",
        "scenario": "配置IPv6地址（mask为子网前缀长度如64）",
        "type": "功能测试", "priority": "P1",
        "precondition": "CHSM支持IPv6",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid",
            "netAddrs": [{"name": "eth0", "ip": "2001:db8::1", "mask": "64", "gateway": "2001:db8::ffff"}]
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "IPv6地址配置成功",
        "remark": "文档提到支持IPv6，mask为前缀长度",
    })
    return cases


def gen_config_chsm_ntp_cases():
    """7.1.5 configCHSMNtp"""
    cases = []
    base = {
        "module": "CHSM配置管理",
        "name": "configCHSMNtp-配置NTP服务器",
        "path": f"{API_PREFIX}/chsm/ntp",
        "method": "POST",
        "auth": "trusted",
    }
    cases.append({
        **base, "id": "CHSM_NTP_001",
        "scenario": "正常配置NTP服务器地址和同步周期",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({"requestId": "uuid", "addr": "10.1.1.1", "syncPeriod": 60}, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "NTP配置成功",
    })
    cases.append({
        **base, "id": "CHSM_NTP_002",
        "scenario": "syncPeriod超过系统最大周期",
        "type": "边界测试", "priority": "P1",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({"requestId": "uuid", "addr": "10.1.1.1", "syncPeriod": 999999}, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "超过最大周期时，配置为系统最大周期，但接口按请求值响应",
        "remark": "文档说明：超过最大周期配成最大值但按请求值响应",
    })
    cases.append({
        **base, "id": "CHSM_NTP_003",
        "scenario": "NTP地址为空字符串",
        "type": "异常测试", "priority": "P1",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({"requestId": "uuid", "addr": "", "syncPeriod": 60}, ensure_ascii=False),
        "expected_status": 400,
        "expected_response": std_response(400, "NTP地址不能为空"),
        "expected_desc": "返回400，NTP地址校验失败",
    })
    cases.append({
        **base, "id": "CHSM_NTP_004",
        "scenario": "syncPeriod为负数",
        "type": "边界测试", "priority": "P2",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({"requestId": "uuid", "addr": "10.1.1.1", "syncPeriod": -1}, ensure_ascii=False),
        "expected_status": 400,
        "expected_response": std_response(400, "同步周期无效"),
        "expected_desc": "返回400，同步周期不能为负数",
    })
    return cases


def gen_config_upload_addr_cases():
    """7.1.6 configCHSMUploadAddress"""
    cases = []
    base = {
        "module": "CHSM配置管理",
        "name": "configCHSMUploadAddress-配置影像上传地址",
        "path": f"{API_PREFIX}/chsm/imageuploader",
        "method": "POST",
        "auth": "trusted",
    }
    cases.append({
        **base, "id": "CHSM_UPLOAD_001",
        "scenario": "正常配置影像上传地址",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({"requestId": "uuid", "url": "http://192.168.0.1/image/upload"}, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "影像上传地址配置成功",
    })
    cases.append({
        **base, "id": "CHSM_UPLOAD_002",
        "scenario": "url参数为空",
        "type": "异常测试", "priority": "P1",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({"requestId": "uuid", "url": ""}, ensure_ascii=False),
        "expected_status": 400,
        "expected_response": std_response(400, "url不能为空"),
        "expected_desc": "返回400",
    })
    return cases


def gen_config_syslog_cases():
    """7.1.7 configSyslogAddr"""
    cases = []
    base = {
        "module": "CHSM配置管理",
        "name": "configSyslogAddr-配置日志上传地址",
        "path": f"{API_PREFIX}/chsm/loguploader",
        "method": "POST",
        "auth": "trusted",
    }
    cases.append({
        **base, "id": "CHSM_SYSLOG_001",
        "scenario": "配置syslog类型的日志上传地址",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid",
            "logServerType": "syslog",
            "logServerAddress": "192.168.1.2"
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "syslog地址配置成功",
    })
    cases.append({
        **base, "id": "CHSM_SYSLOG_002",
        "scenario": "配置logserver类型的日志上传地址（URL格式）",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid",
            "logServerType": "logserver",
            "logServerAddress": "http://192.168.1.2/log"
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "logserver URL配置成功",
    })
    cases.append({
        **base, "id": "CHSM_SYSLOG_003",
        "scenario": "logServerType为非法值",
        "type": "异常测试", "priority": "P1",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid",
            "logServerType": "invalid",
            "logServerAddress": "192.168.1.2"
        }, ensure_ascii=False),
        "expected_status": 400,
        "expected_response": std_response(400, "不支持的日志服务器类型"),
        "expected_desc": "返回400，logServerType只能是syslog或logserver",
    })
    return cases


def gen_export_chsm_cases():
    """7.1.8 exportCHSM"""
    cases = []
    base = {
        "module": "CHSM配置管理",
        "name": "exportCHSM-导出CHSM影像",
        "path": f"{API_PREFIX}/chsm/image",
        "method": "POST",
        "auth": "trusted",
    }
    cases.append({
        **base, "id": "CHSM_EXPORT_001",
        "scenario": "正常导出CHSM影像（oprType=export）",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已正常运行，已配置上传地址",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid", "oprType": "export",
            "callbackUrl": MOCK_CALLBACK
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "返回200，导出任务已提交",
    })
    cases.append({
        **base, "id": "CHSM_EXPORT_002",
        "scenario": "不传callbackUrl参数",
        "type": "异常测试", "priority": "P1",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({"requestId": "uuid", "oprType": "export"}, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": "",
        "expected_desc": "验证callbackUrl是否为必填参数",
        "remark": "待确认callbackUrl是否必填",
    })
    return cases


def gen_import_chsm_cases():
    """7.1.9 importCHSM"""
    cases = []
    base = {
        "module": "CHSM配置管理",
        "name": "importCHSM-导入CHSM影像",
        "path": f"{API_PREFIX}/chsm/image",
        "method": "POST",
        "auth": "trusted",
    }
    cases.append({
        **base, "id": "CHSM_IMPORT_001",
        "scenario": "正常导入CHSM影像（oprType=import，RSA签名）",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已正常运行，有可用影像文件",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid", "oprType": "import",
            "imageUrl": "http://192.168.0.1/image.zip",
            "alg": "RSAWithSHA256", "sign": "<BASE64签名值>",
            "callbackUrl": MOCK_CALLBACK
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "返回200，导入任务已提交",
    })
    cases.append({
        **base, "id": "CHSM_IMPORT_002",
        "scenario": "使用SM2WithSM3签名导入影像",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已正常运行，有可用影像文件",
        "headers": auth_header("SM2WithSM3"),
        "params": json.dumps({
            "requestId": "uuid", "oprType": "import",
            "imageUrl": "http://192.168.0.1/image.zip",
            "alg": "SM2WithSM3", "sign": "<BASE64签名值>",
            "callbackUrl": MOCK_CALLBACK
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "SM2签名验证通过，导入任务已提交",
    })
    cases.append({
        **base, "id": "CHSM_IMPORT_003",
        "scenario": "imageUrl地址不可达",
        "type": "异常测试", "priority": "P1",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid", "oprType": "import",
            "imageUrl": "http://10.0.0.1/notexist.zip",
            "alg": "RSAWithSHA256", "sign": "<BASE64签名值>",
            "callbackUrl": MOCK_CALLBACK
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": "",
        "expected_desc": "接口可能先返回200接受任务，实际失败通过回调通知",
        "remark": "异步接口，错误可能通过回调返回",
    })
    cases.append({
        **base, "id": "CHSM_IMPORT_004",
        "scenario": "sign签名值错误（篡改签名）",
        "type": "异常测试", "priority": "P0",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid", "oprType": "import",
            "imageUrl": "http://192.168.0.1/image.zip",
            "alg": "RSAWithSHA256", "sign": "invalid_sign_value",
            "callbackUrl": MOCK_CALLBACK
        }, ensure_ascii=False),
        "expected_status": 400,
        "expected_response": std_response(400, "签名验证失败"),
        "expected_desc": "返回400或通过回调报告签名验证失败",
    })
    cases.append({
        **base, "id": "CHSM_IMPORT_005",
        "scenario": "缺少alg参数",
        "type": "异常测试", "priority": "P1",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid", "oprType": "import",
            "imageUrl": "http://192.168.0.1/image.zip",
            "sign": "<BASE64签名值>",
            "callbackUrl": MOCK_CALLBACK
        }, ensure_ascii=False),
        "expected_status": 400,
        "expected_response": std_response(400, "缺少alg参数"),
        "expected_desc": "返回400，缺少签名算法参数",
    })
    return cases


def gen_upgrade_chsm_cases():
    """7.1.10 upgradeCHSM"""
    cases = []
    base = {
        "module": "CHSM配置管理",
        "name": "upgradeCHSM-升级CHSM",
        "path": f"{API_PREFIX}/chsm",
        "method": "POST",
        "auth": "trusted",
    }
    cases.append({
        **base, "id": "CHSM_UPGRADE_001",
        "scenario": "正常升级CHSM（oprType=upgrade）",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已运行，升级包已准备",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid", "oprType": "upgrade",
            "packVersion": "v2.0",
            "packUrl": "http://192.168.0.1/update.zip",
            "alg": "RSAWithSHA256", "sign": "<16进制签名值>",
            "callbackUrl": MOCK_CALLBACK
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "返回200，升级任务已提交",
        "remark": "注意：sign为16进制编码（与import的BASE64不同）",
    })
    cases.append({
        **base, "id": "CHSM_UPGRADE_002",
        "scenario": "SM2WithSM3签名升级CHSM",
        "type": "功能测试", "priority": "P0",
        "precondition": "已配置SM2公钥，升级包已准备",
        "headers": auth_header("SM2WithSM3"),
        "params": json.dumps({
            "requestId": "uuid", "oprType": "upgrade",
            "packVersion": "v2.0",
            "packUrl": "http://192.168.0.1/update.zip",
            "alg": "SM2WithSM3", "sign": "<16进制签名值>",
            "callbackUrl": MOCK_CALLBACK
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "SM2签名认证通过，升级任务已提交",
    })
    cases.append({
        **base, "id": "CHSM_UPGRADE_003",
        "scenario": "packUrl地址无效",
        "type": "异常测试", "priority": "P1",
        "precondition": "CHSM已运行",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid", "oprType": "upgrade",
            "packVersion": "v2.0", "packUrl": "",
            "alg": "RSAWithSHA256", "sign": "<16进制签名值>",
            "callbackUrl": MOCK_CALLBACK
        }, ensure_ascii=False),
        "expected_status": 400,
        "expected_response": std_response(400, "升级包地址无效"),
        "expected_desc": "返回400",
    })
    return cases


def gen_restart_chsm_cases():
    """7.1.11 restartCHSM"""
    cases = []
    base = {
        "module": "CHSM配置管理",
        "name": "restartCHSM-重启CHSM",
        "path": f"{API_PREFIX}/chsm",
        "method": "POST",
        "auth": "trusted",
    }
    cases.append({
        **base, "id": "CHSM_RESTART_001",
        "scenario": "正常重启CHSM（oprType=restart）",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid", "oprType": "restart",
            "callbackUrl": MOCK_CALLBACK
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "返回200，重启任务已提交",
    })
    return cases


def gen_backup_chsm_cases():
    """7.1.12 backupCHSM"""
    cases = []
    base = {
        "module": "CHSM配置管理",
        "name": "backupCHSM-备份CHSM",
        "path": f"{API_PREFIX}/chsm",
        "method": "POST",
        "auth": "trusted",
    }
    cases.append({
        **base, "id": "CHSM_BACKUP_001",
        "scenario": "正常备份CHSM敏感数据（oprType=backup）",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid", "oprType": "backup",
            "callbackUrl": MOCK_CALLBACK
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "返回200，备份任务已提交",
        "remark": "可选接口",
    })
    return cases


def gen_restore_chsm_cases():
    """7.1.13 restoreCHSM"""
    cases = []
    base = {
        "module": "CHSM配置管理",
        "name": "restoreCHSM-恢复CHSM",
        "path": f"{API_PREFIX}/chsm",
        "method": "POST",
        "auth": "trusted",
    }
    cases.append({
        **base, "id": "CHSM_RESTORE_001",
        "scenario": "正常恢复CHSM备份数据（oprType=restore，RSA签名）",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已正常运行，备份数据已准备",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid", "oprType": "restore",
            "backupUrl": "http://192.168.0.1/backup.bak",
            "alg": "RSAWithSHA256", "sign": "<BASE64签名值>",
            "callbackUrl": MOCK_CALLBACK
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "返回200，恢复任务已提交",
        "remark": "可选接口，sign为BASE64编码",
    })
    cases.append({
        **base, "id": "CHSM_RESTORE_002",
        "scenario": "SM2WithSM3签名恢复CHSM",
        "type": "功能测试", "priority": "P0",
        "precondition": "已配置SM2公钥",
        "headers": auth_header("SM2WithSM3"),
        "params": json.dumps({
            "requestId": "uuid", "oprType": "restore",
            "backupUrl": "http://192.168.0.1/backup.bak",
            "alg": "SM2WithSM3", "sign": "<BASE64签名值>",
            "callbackUrl": MOCK_CALLBACK
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "SM2签名认证通过，恢复任务已提交",
    })
    return cases


def gen_config_alarm_cases():
    """7.1.14 configCHSMAlarmAddress"""
    cases = []
    base = {
        "module": "CHSM配置管理",
        "name": "configCHSMAlarmAddress-配置告警上传地址",
        "path": f"{API_PREFIX}/chsm/alarmaddress",
        "method": "POST",
        "auth": "trusted",
    }
    cases.append({
        **base, "id": "CHSM_ALARM_001",
        "scenario": "正常配置告警和监控上报地址",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid",
            "url": "http://192.168.0.1:8543/alarm/upload",
            "monitoringUrl": "https://192.168.0.1:8543/v1/metric/upload"
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "告警和监控地址配置成功",
        "remark": "URL已确认为alarmaddress（文档中alramaddress为拼写错误）",
    })
    cases.append({
        **base, "id": "CHSM_ALARM_002",
        "scenario": "只配置告警地址，不配置监控地址",
        "type": "功能测试", "priority": "P1",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid",
            "url": "http://192.168.0.1:8543/alarm/upload"
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": "",
        "expected_desc": "验证monitoringUrl是否为必填参数",
        "remark": "待确认monitoringUrl是否必填",
    })
    return cases


def gen_config_token_cases():
    """7.1.15 configCHSMToken"""
    cases = []
    base = {
        "module": "CHSM配置管理",
        "name": "configCHSMToken-配置云平台Token",
        "path": f"{API_PREFIX}/chsm/cloudtoken",
        "method": "POST",
        "auth": "trusted",
    }
    cases.append({
        **base, "id": "CHSM_TOKEN_001",
        "scenario": "正常配置云平台访问Token",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({"requestId": "uuid", "cloudToken": "valid-iam-token-string"}, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "Token配置成功",
    })
    cases.append({
        **base, "id": "CHSM_TOKEN_002",
        "scenario": "cloudToken为空字符串",
        "type": "异常测试", "priority": "P1",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({"requestId": "uuid", "cloudToken": ""}, ensure_ascii=False),
        "expected_status": 400,
        "expected_response": std_response(400, "Token不能为空"),
        "expected_desc": "返回400",
    })
    return cases


def gen_config_mooc_cases():
    """7.1.16 configCHSMMOOCAddress"""
    cases = []
    base = {
        "module": "CHSM配置管理",
        "name": "configCHSMMOOCAddress-配置MO OC对接信息",
        "path": f"{API_PREFIX}/chsm/moocaddress",
        "method": "POST",
        "auth": "trusted",
    }
    cases.append({
        **base, "id": "CHSM_MOOC_001",
        "scenario": "正常配置ManageOne OC对接信息",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid",
            "ocIp": "10.10.1.100",
            "ocProtocol": "https",
            "ocPort": "26335",
            "resourceUrl": "/rest/resources/v1/instances",
            "performanceUrl": "/V1.0/metric-data",
            "bizRegionNativeId": "region-001",
            "cloudInfraType": "FUSION_CLOUD",
            "dewServiceHost": "dew.example.com"
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "OC对接信息配置成功",
    })
    cases.append({
        **base, "id": "CHSM_MOOC_002",
        "scenario": "缺少ocIp参数",
        "type": "异常测试", "priority": "P1",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid",
            "ocProtocol": "https", "ocPort": "26335"
        }, ensure_ascii=False),
        "expected_status": 400,
        "expected_response": std_response(400, "缺少必要参数"),
        "expected_desc": "返回400，缺少ocIp",
    })
    return cases


def gen_debug_info_cases():
    """7.1.17 getCHSMDebugInfo"""
    cases = []
    base = {
        "module": "CHSM配置管理",
        "name": "getCHSMDebugInfo-获取调试信息",
        "path": f"{API_PREFIX}/chsm/debuginfo",
        "method": "POST",
        "auth": "trusted",
    }
    cases.append({
        **base, "id": "CHSM_DEBUG_001",
        "scenario": "正常获取CHSM及VSM的告警和调试信息",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({"requestId": "uuid"}, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success", '{"chsm":{id,debugInfo}, "vsms":[{id,debugInfo}]}'),
        "expected_desc": "1. 返回200\n2. result包含chsm和vsms的调试信息",
        "remark": "摘要表标注guest但详细定义按trusted处理",
    })
    return cases


def gen_device_info_cases():
    """7.1.18 getCHSMDeviceInfo"""
    cases = []
    base = {
        "module": "CHSM配置管理",
        "name": "getCHSMDeviceInfo-获取设备信息及授权码",
        "path": f"{API_PREFIX}/chsm",
        "method": "POST",
        "auth": "trusted",
    }
    cases.append({
        **base, "id": "CHSM_DEVICE_001",
        "scenario": "正常获取CHSM设备信息（oprType=getDeviceInfo）",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已正常运行",
        "headers": auth_header(),
        "params": json.dumps({"requestId": "uuid", "oprType": "getDeviceInfo"}, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success", '{"version","ip","sn","mac","hwAuthCode"}'),
        "expected_desc": "1. 返回200\n2. result包含version/ip/sn/mac/hwAuthCode",
        "remark": "此接口在摘要表中缺失，待确认是否在二期范围",
    })
    return cases


# ============= VSM 接口 =============

def gen_vsm_info_cases():
    """7.2.1 getVSMInfo"""
    cases = []
    base = {
        "module": "VSM配置管理",
        "name": "getVSMInfo-获取VSM详细信息",
        "path": f"{API_PREFIX}/vsm",
        "method": "POST",
        "auth": "trusted",
    }
    cases.append({
        **base, "id": "VSM_INFO_001",
        "scenario": "正常获取指定VSM详细信息（oprType=getinfo）",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已运行，至少有1个VSM",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid", "oprType": "getinfo",
            "vsmId": "<有效的vsmId>"
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success", '{id,version,token,ip,func,mask,gateway,...}'),
        "expected_desc": "1. 返回200\n2. result包含VSM详细信息（id/version/ip/func等）",
    })
    cases.append({
        **base, "id": "VSM_INFO_002",
        "scenario": "vsmId不存在",
        "type": "异常测试", "priority": "P1",
        "precondition": "CHSM已运行",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid", "oprType": "getinfo",
            "vsmId": "non-exist-vsm-id"
        }, ensure_ascii=False),
        "expected_status": 400,
        "expected_response": std_response(400, "VSM不存在"),
        "expected_desc": "返回400或404，提示VSM不存在",
    })
    cases.append({
        **base, "id": "VSM_INFO_003",
        "scenario": "缺少vsmId参数",
        "type": "异常测试", "priority": "P1",
        "precondition": "CHSM已运行",
        "headers": auth_header(),
        "params": json.dumps({"requestId": "uuid", "oprType": "getinfo"}, ensure_ascii=False),
        "expected_status": 400,
        "expected_response": std_response(400, "缺少vsmId"),
        "expected_desc": "返回400，缺少必要参数vsmId",
    })
    cases.append({
        **base, "id": "VSM_INFO_004",
        "scenario": "SM2WithSM3签名获取VSM信息",
        "type": "功能测试", "priority": "P0",
        "precondition": "已配置SM2公钥",
        "headers": auth_header("SM2WithSM3"),
        "params": json.dumps({
            "requestId": "uuid", "oprType": "getinfo",
            "vsmId": "<有效的vsmId>"
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "SM2签名认证通过，返回VSM信息",
    })
    return cases


def gen_vsm_status_cases():
    """7.2.2 getVSMStatus"""
    cases = []
    base = {
        "module": "VSM配置管理",
        "name": "getVSMStatus-获取VSM运行状态",
        "path": f"{API_PREFIX}/vsm/status",
        "method": "GET",
        "auth": "guest",
    }
    cases.append({
        **base, "id": "VSM_STATUS_001",
        "scenario": "正常获取指定VSM的运行状态",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已运行，至少有1个VSM",
        "headers": "无需认证头（guest接口）",
        "params": "requestId=uuid&vsmId=<有效的vsmId>",
        "expected_status": 200,
        "expected_response": std_response(200, "success", '{"status":"ok"}'),
        "expected_desc": "返回200，result.status为ok或fail",
    })
    cases.append({
        **base, "id": "VSM_STATUS_002",
        "scenario": "vsmId不存在",
        "type": "异常测试", "priority": "P1",
        "precondition": "CHSM已运行",
        "headers": "无需认证头",
        "params": "requestId=uuid&vsmId=non-exist-id",
        "expected_status": 400,
        "expected_response": std_response(400, "VSM不存在"),
        "expected_desc": "返回400，提示VSM不存在",
    })
    return cases


def gen_config_vsm_net_cases():
    """7.2.3 configVSMNet"""
    cases = []
    base = {
        "module": "VSM配置管理",
        "name": "configVSMNet-配置VSM网络",
        "path": f"{API_PREFIX}/vsm/network",
        "method": "POST",
        "auth": "trusted",
    }
    cases.append({
        **base, "id": "VSM_NET_001",
        "scenario": "正常配置VSM网络信息",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已运行，VSM存在",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid", "vsmId": "<有效的vsmId>",
            "ip": "192.168.1.50", "mask": "255.255.255.0", "gateway": "192.168.1.1"
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "VSM网络配置成功",
    })
    cases.append({
        **base, "id": "VSM_NET_002",
        "scenario": "配置IPv6地址（mask为前缀长度）",
        "type": "功能测试", "priority": "P1",
        "precondition": "CHSM已运行，VSM存在",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid", "vsmId": "<有效的vsmId>",
            "ip": "2001:db8::50", "mask": "64", "gateway": "2001:db8::1"
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "IPv6网络配置成功",
    })
    return cases


def gen_config_vsm_token_cases():
    """7.2.4 configVSMToken"""
    cases = []
    base = {
        "module": "VSM配置管理",
        "name": "configVSMToken-配置VSM Token",
        "path": f"{API_PREFIX}/vsm/token",
        "method": "POST",
        "auth": "trusted",
    }
    cases.append({
        **base, "id": "VSM_TOKEN_001",
        "scenario": "正常配置VSM的用户Token（标识租用给某用户）",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已运行，VSM存在",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid", "vsmId": "<有效的vsmId>",
            "token": "user-token-001", "tenantId": "tenant-001"
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "VSM Token配置成功",
    })
    cases.append({
        **base, "id": "VSM_TOKEN_002",
        "scenario": "配置token为0（标识VSM未被使用）",
        "type": "功能测试", "priority": "P1",
        "precondition": "VSM已被某用户租用",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid", "vsmId": "<有效的vsmId>",
            "token": "0", "tenantId": ""
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "VSM释放成功，token置为0",
        "remark": "文档说明token为0表示未被使用",
    })
    return cases


def gen_export_vsm_cases():
    """7.2.5 exportVSM"""
    cases = []
    base = {
        "module": "VSM配置管理",
        "name": "exportVSM-导出VSM影像",
        "path": f"{API_PREFIX}/vsm/image",
        "method": "POST",
        "auth": "trusted",
    }
    cases.append({
        **base, "id": "VSM_EXPORT_001",
        "scenario": "正常导出指定VSM影像（oprType=export）",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已运行，VSM存在",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid", "oprType": "export",
            "vsmId": "<有效的vsmId>",
            "callbackUrl": MOCK_CALLBACK
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "返回200，导出任务已提交",
    })
    return cases


def gen_import_vsm_cases():
    """7.2.6 importVSM"""
    cases = []
    base = {
        "module": "VSM配置管理",
        "name": "importVSM-导入VSM影像",
        "path": f"{API_PREFIX}/vsm/image",
        "method": "POST",
        "auth": "trusted",
    }
    cases.append({
        **base, "id": "VSM_IMPORT_001",
        "scenario": "正常导入VSM影像（RSA签名）",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已运行，VSM存在，影像文件已准备",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid", "oprType": "import",
            "vsmId": "<有效的vsmId>",
            "imageUrl": "http://192.168.0.1/vsm_image.zip",
            "alg": "RSAWithSHA256", "sign": "<BASE64签名值>",
            "callbackUrl": MOCK_CALLBACK
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "返回200，导入任务已提交",
    })
    cases.append({
        **base, "id": "VSM_IMPORT_002",
        "scenario": "SM2WithSM3签名导入VSM影像",
        "type": "功能测试", "priority": "P0",
        "precondition": "已配置SM2公钥",
        "headers": auth_header("SM2WithSM3"),
        "params": json.dumps({
            "requestId": "uuid", "oprType": "import",
            "vsmId": "<有效的vsmId>",
            "imageUrl": "http://192.168.0.1/vsm_image.zip",
            "alg": "SM2WithSM3", "sign": "<BASE64签名值>",
            "callbackUrl": MOCK_CALLBACK
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "SM2签名认证通过，导入任务已提交",
    })
    return cases


def gen_vsm_lifecycle_cases(opr_type, name_cn, case_prefix, extra_note=""):
    """通用 VSM 生命周期接口：startVSM/stopVSM/restartVSM/resetVSM"""
    cases = []
    base = {
        "module": "VSM配置管理",
        "name": f"{opr_type}VSM-{name_cn}",
        "path": f"{API_PREFIX}/vsm",
        "method": "POST",
        "auth": "trusted",
    }
    cases.append({
        **base, "id": f"{case_prefix}_001",
        "scenario": f"正常{name_cn}（oprType={opr_type}）",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已运行，VSM存在",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid", "oprType": opr_type,
            "vsmId": "<有效的vsmId>",
            "callbackUrl": MOCK_CALLBACK
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": f"返回200，{name_cn}任务已提交",
        "remark": extra_note,
    })
    cases.append({
        **base, "id": f"{case_prefix}_002",
        "scenario": f"vsmId不存在时{name_cn}",
        "type": "异常测试", "priority": "P1",
        "precondition": "CHSM已运行",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid", "oprType": opr_type,
            "vsmId": "non-exist-id",
            "callbackUrl": MOCK_CALLBACK
        }, ensure_ascii=False),
        "expected_status": 400,
        "expected_response": std_response(400, "VSM不存在"),
        "expected_desc": "返回400，提示VSM不存在",
    })
    cases.append({
        **base, "id": f"{case_prefix}_003",
        "scenario": f"SM2WithSM3签名{name_cn}",
        "type": "功能测试", "priority": "P0",
        "precondition": "已配置SM2公钥",
        "headers": auth_header("SM2WithSM3"),
        "params": json.dumps({
            "requestId": "uuid", "oprType": opr_type,
            "vsmId": "<有效的vsmId>",
            "callbackUrl": MOCK_CALLBACK
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": f"SM2签名认证通过，{name_cn}任务已提交",
    })
    return cases


def gen_upgrade_vsm_cases():
    """7.2.13 upgradeVSM"""
    cases = []
    base = {
        "module": "VSM配置管理",
        "name": "upgradeVSM-升级VSM",
        "path": f"{API_PREFIX}/vsm",
        "method": "POST",
        "auth": "trusted",
    }
    cases.append({
        **base, "id": "VSM_UPGRADE_001",
        "scenario": "正常升级VSM（oprType=upgrade）",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已运行，VSM存在，升级包已准备",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid", "oprType": "upgrade",
            "vsmId": "<有效的vsmId>",
            "packVersion": "v2.0",
            "packUrl": "http://192.168.0.1/vsm_update.zip",
            "alg": "RSAWithSHA256", "sign": "<16进制签名值>",
            "callbackUrl": MOCK_CALLBACK
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "返回200，升级任务已提交",
        "remark": "注意：sign为16进制编码",
    })
    cases.append({
        **base, "id": "VSM_UPGRADE_002",
        "scenario": "SM2WithSM3签名升级VSM",
        "type": "功能测试", "priority": "P0",
        "precondition": "已配置SM2公钥",
        "headers": auth_header("SM2WithSM3"),
        "params": json.dumps({
            "requestId": "uuid", "oprType": "upgrade",
            "vsmId": "<有效的vsmId>",
            "packVersion": "v2.0",
            "packUrl": "http://192.168.0.1/vsm_update.zip",
            "alg": "SM2WithSM3", "sign": "<16进制签名值>",
            "callbackUrl": MOCK_CALLBACK
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "SM2签名认证通过，升级任务已提交",
    })
    return cases


# ============= 授权配置类接口 =============

def gen_config_pk_cases():
    """7.3.2 configCHSMPk"""
    cases = []
    base = {
        "module": "授权配置",
        "name": "configCHSMPk-配置认证公钥",
        "path": f"{API_PREFIX}/chsm/authpk",
        "method": "POST",
        "auth": "trusted/guest",
    }
    cases.append({
        **base, "id": "AUTH_PK_001",
        "scenario": "首次配置RSA认证公钥（algorithm=rsa）",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM未配置公钥（首次配置，guest权限即可）",
        "headers": "首次配置无需认证头（guest）",
        "params": json.dumps({
            "requestId": "uuid",
            "algorithm": "rsa",
            "pks": ["<RSA公钥BASE64>"]
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "RSA公钥配置成功",
    })
    cases.append({
        **base, "id": "AUTH_PK_002",
        "scenario": "配置SM2认证公钥（algorithm=sm2）",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已运行",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid",
            "algorithm": "sm2",
            "pks": ["<SM2公钥BASE64>"]
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "SM2公钥配置成功",
    })
    cases.append({
        **base, "id": "AUTH_PK_003",
        "scenario": "配置多个公钥",
        "type": "功能测试", "priority": "P1",
        "precondition": "CHSM已运行",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid",
            "algorithm": "rsa",
            "pks": ["<公钥1>", "<公钥2>"]
        }, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "多个公钥配置成功",
    })
    cases.append({
        **base, "id": "AUTH_PK_004",
        "scenario": "algorithm为非法值（如aes）",
        "type": "异常测试", "priority": "P1",
        "precondition": "CHSM已运行",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid",
            "algorithm": "aes",
            "pks": ["<公钥>"]
        }, ensure_ascii=False),
        "expected_status": 400,
        "expected_response": std_response(400, "不支持的算法"),
        "expected_desc": "返回400，algorithm只支持rsa和sm2",
    })
    cases.append({
        **base, "id": "AUTH_PK_005",
        "scenario": "pks为空列表",
        "type": "异常测试", "priority": "P1",
        "precondition": "CHSM已运行",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid", "algorithm": "rsa", "pks": []
        }, ensure_ascii=False),
        "expected_status": 400,
        "expected_response": std_response(400, "公钥列表不能为空"),
        "expected_desc": "返回400",
    })
    cases.append({
        **base, "id": "AUTH_PK_006",
        "scenario": "公钥格式错误（非BASE64编码）",
        "type": "异常测试", "priority": "P1",
        "precondition": "CHSM已运行",
        "headers": auth_header(),
        "params": json.dumps({
            "requestId": "uuid", "algorithm": "rsa",
            "pks": ["invalid-key-format!!!"]
        }, ensure_ascii=False),
        "expected_status": 400,
        "expected_response": std_response(400, "公钥格式错误"),
        "expected_desc": "返回400，公钥格式校验失败",
    })
    return cases


def gen_get_pk_cases():
    """7.3.1 getCHSMPk"""
    cases = []
    base = {
        "module": "授权配置",
        "name": "getCHSMPk-获取认证公钥指纹",
        "path": f"{API_PREFIX}/chsm/authpk",
        "method": "GET",
        "auth": "guest",
    }
    cases.append({
        **base, "id": "AUTH_GETPK_001",
        "scenario": "正常获取已配置的认证公钥指纹",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已配置公钥",
        "headers": "无需认证头（guest接口）",
        "params": "requestId=uuid",
        "expected_status": 200,
        "expected_response": std_response(200, "success", '{"algorithm":"sm3/sha256","fingerprints":["fp1","fp2"]}'),
        "expected_desc": "1. 返回200\n2. result包含algorithm和fingerprints列表\n3. fingerprints为BASE64编码",
    })
    cases.append({
        **base, "id": "AUTH_GETPK_002",
        "scenario": "CHSM未配置公钥时获取指纹",
        "type": "功能测试", "priority": "P1",
        "precondition": "CHSM未配置任何公钥",
        "headers": "无需认证头",
        "params": "requestId=uuid",
        "expected_status": 200,
        "expected_response": std_response(200, "success", '{"fingerprints":[]}'),
        "expected_desc": "返回200，fingerprints为空列表",
    })
    return cases


def gen_clear_pk_cases():
    """7.3.3 clearCHSMPk"""
    cases = []
    base = {
        "module": "授权配置",
        "name": "clearCHSMPk-清空认证公钥",
        "path": f"{API_PREFIX}/chsm/authpk",
        "method": "DELETE",
        "auth": "trusted",
    }
    cases.append({
        **base, "id": "AUTH_CLEARPK_001",
        "scenario": "正常清空所有认证公钥",
        "type": "功能测试", "priority": "P0",
        "precondition": "CHSM已配置公钥",
        "headers": auth_header(),
        "params": json.dumps({"requestId": "uuid"}, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "公钥清空成功",
    })
    cases.append({
        **base, "id": "AUTH_CLEARPK_002",
        "scenario": "无认证头清空公钥（trusted接口）",
        "type": "异常测试", "priority": "P0",
        "precondition": "CHSM已配置公钥",
        "headers": "无认证头",
        "params": json.dumps({"requestId": "uuid"}, ensure_ascii=False),
        "expected_status": 403,
        "expected_response": std_response(403, "认证失败"),
        "expected_desc": "返回403，认证失败",
    })
    cases.append({
        **base, "id": "AUTH_CLEARPK_003",
        "scenario": "SM2WithSM3签名清空公钥",
        "type": "功能测试", "priority": "P0",
        "precondition": "已配置SM2公钥",
        "headers": auth_header("SM2WithSM3"),
        "params": json.dumps({"requestId": "uuid"}, ensure_ascii=False),
        "expected_status": 200,
        "expected_response": std_response(200, "success"),
        "expected_desc": "SM2签名认证通过，公钥清空成功",
    })
    return cases


# ============= 通用异常场景（适用于所有trusted接口）=============

def gen_common_auth_cases():
    """通用认证异常测试"""
    cases = []
    base = {
        "module": "通用认证测试",
        "name": "通用-认证异常场景",
        "path": "<适用于所有trusted接口>",
        "method": "POST",
        "auth": "trusted",
    }
    cases.append({
        **base, "id": "COMMON_AUTH_001",
        "scenario": "签名值被篡改（CHSM-Signature错误）",
        "type": "异常测试", "priority": "P0",
        "precondition": "已配置公钥，签名认证已启用",
        "headers": '{\n  "CHSM-AuthPK": "<正确指纹>",\n  "CHSM-SignatureAlg": "SHA256withRSA",\n  "CHSM-Signature": "invalid_tampered_signature"\n}',
        "params": "任意合法请求体",
        "expected_status": 403,
        "expected_response": std_response(403, "签名验证失败"),
        "expected_desc": "返回403，签名验证失败",
    })
    cases.append({
        **base, "id": "COMMON_AUTH_002",
        "scenario": "公钥指纹不匹配（CHSM-AuthPK错误）",
        "type": "异常测试", "priority": "P0",
        "precondition": "已配置公钥",
        "headers": '{\n  "CHSM-AuthPK": "wrong_fingerprint_value",\n  "CHSM-SignatureAlg": "SHA256withRSA",\n  "CHSM-Signature": "<正确签名>"\n}',
        "params": "任意合法请求体",
        "expected_status": 403,
        "expected_response": std_response(403, "公钥指纹不匹配"),
        "expected_desc": "返回403，找不到对应的公钥",
    })
    cases.append({
        **base, "id": "COMMON_AUTH_003",
        "scenario": "签名算法标识不支持（CHSM-SignatureAlg=unknown）",
        "type": "异常测试", "priority": "P1",
        "precondition": "已配置公钥",
        "headers": '{\n  "CHSM-AuthPK": "<正确指纹>",\n  "CHSM-SignatureAlg": "UnknownAlg",\n  "CHSM-Signature": "<签名>"\n}',
        "params": "任意合法请求体",
        "expected_status": 403,
        "expected_response": std_response(403, "不支持的签名算法"),
        "expected_desc": "返回403，不支持的签名算法",
    })
    cases.append({
        **base, "id": "COMMON_AUTH_004",
        "scenario": "缺少CHSM-Signature头",
        "type": "异常测试", "priority": "P0",
        "precondition": "已配置公钥",
        "headers": '{\n  "CHSM-AuthPK": "<正确指纹>",\n  "CHSM-SignatureAlg": "SHA256withRSA"\n}',
        "params": "任意合法请求体",
        "expected_status": 403,
        "expected_response": std_response(403, "缺少签名头"),
        "expected_desc": "返回403，缺少必要的认证头",
    })
    cases.append({
        **base, "id": "COMMON_AUTH_005",
        "scenario": "缺少CHSM-AuthPK头",
        "type": "异常测试", "priority": "P0",
        "precondition": "已配置公钥",
        "headers": '{\n  "CHSM-SignatureAlg": "SHA256withRSA",\n  "CHSM-Signature": "<签名>"\n}',
        "params": "任意合法请求体",
        "expected_status": 403,
        "expected_response": std_response(403, "缺少公钥指纹头"),
        "expected_desc": "返回403",
    })
    return cases


# ============= 汇总生成 =============

def generate_all_cases():
    all_cases = []
    all_cases.extend(gen_chsm_info_cases())
    all_cases.extend(gen_chsm_status_cases())
    all_cases.extend(gen_chsm_allstatus_cases())
    all_cases.extend(gen_config_chsm_net_cases())
    all_cases.extend(gen_config_chsm_ntp_cases())
    all_cases.extend(gen_config_upload_addr_cases())
    all_cases.extend(gen_config_syslog_cases())
    all_cases.extend(gen_export_chsm_cases())
    all_cases.extend(gen_import_chsm_cases())
    all_cases.extend(gen_upgrade_chsm_cases())
    all_cases.extend(gen_restart_chsm_cases())
    all_cases.extend(gen_backup_chsm_cases())
    all_cases.extend(gen_restore_chsm_cases())
    all_cases.extend(gen_config_alarm_cases())
    all_cases.extend(gen_config_token_cases())
    all_cases.extend(gen_config_mooc_cases())
    all_cases.extend(gen_debug_info_cases())
    all_cases.extend(gen_device_info_cases())
    # VSM
    all_cases.extend(gen_vsm_info_cases())
    all_cases.extend(gen_vsm_status_cases())
    all_cases.extend(gen_config_vsm_net_cases())
    all_cases.extend(gen_config_vsm_token_cases())
    all_cases.extend(gen_export_vsm_cases())
    all_cases.extend(gen_import_vsm_cases())
    all_cases.extend(gen_vsm_lifecycle_cases("start", "启动VSM", "VSM_START"))
    all_cases.extend(gen_vsm_lifecycle_cases("stop", "停止VSM", "VSM_STOP"))
    all_cases.extend(gen_vsm_lifecycle_cases("restart", "重启VSM", "VSM_RESTART"))
    all_cases.extend(gen_vsm_lifecycle_cases("reset", "重置VSM", "VSM_RESET", "高危操作，清除所有密钥/token/ukey信息"))
    all_cases.extend(gen_upgrade_vsm_cases())
    # 授权配置
    all_cases.extend(gen_config_pk_cases())
    all_cases.extend(gen_get_pk_cases())
    all_cases.extend(gen_clear_pk_cases())
    # 通用认证
    all_cases.extend(gen_common_auth_cases())
    return all_cases


def write_excel(cases, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "管理接口"

    # Style definitions
    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Write headers
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    for row_idx, case in enumerate(cases, 2):
        row_data = [
            case.get("id", ""),
            case.get("module", ""),
            case.get("name", ""),
            case.get("path", ""),
            case.get("method", ""),
            case.get("scenario", ""),
            case.get("type", ""),
            case.get("priority", ""),
            case.get("precondition", ""),
            case.get("headers", ""),
            case.get("params", ""),
            case.get("expected_status", ""),
            case.get("expected_response", ""),
            case.get("expected_desc", ""),
            "",  # 测试状态
            "",  # 执行人
            "",  # 执行时间
            "",  # 缺陷ID
            case.get("remark", ""),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")
            cell.alignment = cell_alignment
            cell.border = thin_border

    # Column widths
    col_widths = [18, 14, 30, 25, 8, 40, 8, 6, 25, 35, 50, 8, 40, 30, 8, 8, 10, 8, 30]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze first row
    ws.freeze_panes = 'A2'

    wb.save(output_path)
    return len(cases)


if __name__ == '__main__':
    cases = generate_all_cases()
    output = '/home/dreamer/utimaco/2026II/华为云密码机二期-接口测试用例.xlsx'
    count = write_excel(cases, output)
    print(f"生成完成！共 {count} 条测试用例")
    print(f"文件路径: {output}")

    # 统计
    modules = {}
    priorities = {}
    types = {}
    for c in cases:
        m = c.get("module", "")
        modules[m] = modules.get(m, 0) + 1
        p = c.get("priority", "")
        priorities[p] = priorities.get(p, 0) + 1
        t = c.get("type", "")
        types[t] = types.get(t, 0) + 1

    print("\n=== 按模块统计 ===")
    for k, v in sorted(modules.items()):
        print(f"  {k}: {v}条")
    print("\n=== 按优先级统计 ===")
    for k, v in sorted(priorities.items()):
        print(f"  {k}: {v}条")
    print("\n=== 按类型统计 ===")
    for k, v in sorted(types.items()):
        print(f"  {k}: {v}条")
